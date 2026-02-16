"""Formula recalculation command with Excel/LibreOffice backend auto-detection."""

import time
from typing import Any

import typer

from agent_xlsx.cli import app
from agent_xlsx.formatters.json_formatter import output
from agent_xlsx.utils.errors import ExcelRequiredError, NoRenderingBackendError, handle_error
from agent_xlsx.utils.validation import validate_file

# Formula error values to scan for
_FORMULA_ERRORS = {"#REF!", "#DIV/0!", "#NAME?", "#NULL!", "#N/A", "#VALUE!", "#NUM!"}


def _check_formula_errors(filepath: str) -> dict[str, Any]:
    """Scan for formula error values without recalculating."""
    from openpyxl import load_workbook

    start = time.perf_counter()
    wb = load_workbook(filepath, data_only=True, read_only=True)

    error_summary: dict[str, dict[str, Any]] = {}
    total_errors = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val in _FORMULA_ERRORS:
                    total_errors += 1
                    if val not in error_summary:
                        error_summary[val] = {"count": 0, "locations": []}
                    error_summary[val]["count"] += 1
                    coord = f"{ws.title}!{cell.coordinate}"
                    if len(error_summary[val]["locations"]) < 20:
                        error_summary[val]["locations"].append(coord)

    wb.close()

    # Count formulas via a separate pass
    total_formulas = 0
    wb2 = load_workbook(filepath, read_only=True)
    for ws in wb2.worksheets:
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell.data_type == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    total_formulas += 1
    wb2.close()

    elapsed_ms = round((time.perf_counter() - start) * 1000, 1)
    status = "errors_found" if total_errors > 0 else "success"

    result: dict[str, Any] = {
        "status": status,
        "mode": "check_only",
        "total_formulas": total_formulas,
        "total_errors": total_errors,
        "check_time_ms": elapsed_ms,
    }
    if error_summary:
        result["error_summary"] = error_summary
    return result


@app.command()
@handle_error
def recalc(
    file: str = typer.Argument(..., help="Path to the Excel file"),
    check_only: bool = typer.Option(
        False, "--check-only", help="Scan for formula errors without recalculating"
    ),
    timeout: int = typer.Option(
        60, "--timeout", help="Timeout in seconds (LibreOffice backend only)"
    ),
    engine: str = typer.Option(
        "auto",
        "--engine",
        "-e",
        help="Rendering backend: auto (default), excel, aspose, or libreoffice",
    ),
) -> None:
    """Recalculate all formulas and report errors.

    Uses Excel (via xlwings) when available, otherwise falls back to LibreOffice.
    Use --engine to force a specific backend. Use --check-only to scan for formula
    errors (#REF!, #DIV/0!, etc.) without recalculating.
    """
    path = validate_file(file)
    filepath = str(path)

    if check_only:
        result = _check_formula_errors(filepath)
    else:
        # Backend selection: explicit --engine or auto-detect
        from agent_xlsx.adapters.libreoffice_adapter import is_libreoffice_available
        from agent_xlsx.adapters.xlwings_adapter import is_excel_available

        use_engine = None
        engine_lower = engine.lower()

        if engine_lower == "excel":
            if not is_excel_available():
                raise ExcelRequiredError("recalc")
            use_engine = "excel"
        elif engine_lower == "aspose":
            from agent_xlsx.adapters.aspose_adapter import is_aspose_available
            from agent_xlsx.utils.errors import AsposeNotInstalledError

            if not is_aspose_available():
                raise AsposeNotInstalledError()
            use_engine = "aspose"
        elif engine_lower in ("libreoffice", "lo"):
            if not is_libreoffice_available():
                from agent_xlsx.utils.errors import LibreOfficeNotFoundError

                raise LibreOfficeNotFoundError()
            use_engine = "libreoffice"
        elif engine_lower == "auto":
            if is_excel_available():
                use_engine = "excel"
            else:
                from agent_xlsx.adapters.aspose_adapter import is_aspose_available

                if is_aspose_available():
                    use_engine = "aspose"
                elif is_libreoffice_available():
                    use_engine = "libreoffice"
                else:
                    raise NoRenderingBackendError("recalc")
        else:
            raise NoRenderingBackendError("recalc")

        if use_engine == "excel":
            from agent_xlsx.adapters.xlwings_adapter import recalculate

            result = recalculate(filepath)
        elif use_engine == "aspose":
            from agent_xlsx.adapters.aspose_adapter import recalculate

            result = recalculate(filepath)
        else:
            from agent_xlsx.adapters.libreoffice_adapter import recalculate

            result = recalculate(filepath, timeout=timeout)

    output(result)
