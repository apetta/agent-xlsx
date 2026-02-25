"""Read data from Excel ranges — fast path via Polars, openpyxl fallback for formulas."""

import time
from pathlib import Path
from typing import Any, Optional

import polars as pl
import typer
from openpyxl.cell.read_only import EmptyCell

from agent_xlsx.adapters.polars_adapter import (
    get_sheet_headers,
    get_sheet_names,
    read_exact_range,
    read_sheet_data,
)
from agent_xlsx.cli import app
from agent_xlsx.formatters.json_formatter import output_spreadsheet_data, should_include_meta
from agent_xlsx.utils.constants import DEFAULT_LIMIT, DEFAULT_OFFSET, MAX_READ_ROWS
from agent_xlsx.utils.dataframe import apply_compact
from agent_xlsx.utils.dates import detect_date_column_indices, excel_serial_to_isodate
from agent_xlsx.utils.errors import SheetNotFoundError, handle_error
from agent_xlsx.utils.validation import (
    ParsedRange,
    col_letter_to_index,
    file_size_human,
    parse_multi_range,
    parse_range,
    validate_file,
)


@app.command()
@handle_error
def read(
    file: str = typer.Argument(..., help="Path to the Excel file"),
    range_: Optional[str] = typer.Argument(
        None,
        metavar="RANGE",
        help="Range e.g. 'Sheet1!A1:C10' or comma-separated 'Sheet1!A1:C10,E1:G10'",
    ),
    sheet: Optional[str] = typer.Option(None, "--sheet", "-s", help="Sheet name"),
    limit: int = typer.Option(DEFAULT_LIMIT, "--limit", "-l", help="Maximum rows to return"),
    offset: int = typer.Option(DEFAULT_OFFSET, "--offset", help="Rows to skip"),
    format_: str = typer.Option("json", "--format", "-f", help="Output format: json or csv"),
    formulas: bool = typer.Option(
        False,
        "--formulas",
        help="Include formula strings (slower)",
    ),
    sort: Optional[str] = typer.Option(
        None,
        "--sort",
        help="Sort by column name",
    ),
    descending: bool = typer.Option(
        False,
        "--descending",
        help="Sort in descending order",
    ),
    no_header: bool = typer.Option(
        False,
        "--no-header",
        help="Treat row 1 as data, use column letters (A, B, C) as headers. "
        "Use for non-tabular sheets like P&L reports and dashboards.",
    ),
    headers: bool = typer.Option(
        False,
        "--headers",
        help="Resolve column letters to row-1 header names in range reads. "
        "Adds column_map to output.",
    ),
    compact: bool = typer.Option(
        True,
        "--compact/--no-compact",
        help="Drop fully-null columns from output to reduce token waste (default: on).",
    ),
    all_sheets: bool = typer.Option(
        False,
        "--all-sheets",
        help="Read the same range(s) from every sheet in the workbook.",
    ),
    precision: Optional[int] = typer.Option(
        None,
        "--precision",
        "-p",
        help="Round float values to N decimal places (default: full precision).",
    ),
) -> None:
    """Read data from an Excel range or sheet.

    Default fast path uses Polars + fastexcel (7-10x faster than openpyxl).
    Speed scales with file size; check file_size_human in the output.
    Use --formulas to fall back to openpyxl for formula string extraction.

    Supports comma-separated ranges (e.g. 'Sheet1!A1:C10,E1:G10') for
    multi-range reads, and --all-sheets to read the same range(s) from
    every sheet.
    """
    path = validate_file(file)
    start = time.perf_counter()

    # Cap limit to prevent massive outputs
    effective_limit = min(limit, MAX_READ_ROWS)

    if formulas:
        _read_with_formulas(path, range_, sheet, effective_limit, offset, compact, precision)
        return

    # --- Determine ranges to read ---
    is_multi_range = range_ is not None and "," in range_

    if is_multi_range:
        assert range_ is not None  # Guaranteed by is_multi_range check above
        ranges = parse_multi_range(range_)
    elif range_:
        ranges = [parse_range(range_)]
    else:
        ranges = []  # No range — full sheet read

    # --- Determine target sheets ---
    available = get_sheet_names(str(path))

    if all_sheets:
        target_sheets: list[str | int] = list(available)
    elif ranges and ranges[0]["sheet"]:
        ts = ranges[0]["sheet"]
        if ts not in available:
            raise SheetNotFoundError(ts, available)
        target_sheets = [ts]
    elif sheet:
        if sheet not in available:
            raise SheetNotFoundError(sheet, available)
        target_sheets = [sheet]
    else:
        target_sheets = [0]  # Default to first sheet

    # --- Multi-result path (multi-range OR all-sheets) ---
    is_multi = is_multi_range or all_sheets

    if is_multi:
        # Auto-resolve headers for multi-range reads — agents need header
        # names (e.g. "2014"), not column letters (e.g. "BG"). --no-header
        # is the explicit opt-out.
        if not no_header:
            headers = True

        results = []
        _formula_check_cache: dict[str, bool] = {}  # per-sheet formula detection cache

        # Header cache — populated lazily per-sheet. The batch multi-range
        # reader returns headers from its single file load (no extra open).
        # For non-batch paths, we open a single reader to fetch headers.
        _header_cache: dict[str, list[str]] = {}
        _header_cache_loaded = False

        def _ensure_header_cache() -> None:
            nonlocal _header_cache_loaded
            if _header_cache_loaded or not (headers and not no_header):
                return
            import fastexcel

            _hdr_reader = fastexcel.read_excel(str(path))
            for ts in target_sheets:
                cache_key = str(ts)
                try:
                    resolved = ts if isinstance(ts, str) else available[ts]
                    _hdr_sheet = _hdr_reader.load_sheet(resolved, n_rows=0)
                    _header_cache[cache_key] = [c.name for c in _hdr_sheet.available_columns()]
                except Exception:
                    _header_cache[cache_key] = []
            _header_cache_loaded = True

        for target_sheet in target_sheets:
            sheet_name = target_sheet if isinstance(target_sheet, str) else available[target_sheet]
            if ranges:
                # Batch optimization: load the sheet once and slice each range
                # from memory, instead of opening N separate readers. The batch
                # reader also returns row-1 headers, eliminating a separate open.
                batch_dfs: list[tuple[pl.DataFrame, str | None]] | None = None
                if len(ranges) > 1:
                    from agent_xlsx.adapters.polars_adapter import read_multi_ranges

                    batch_dfs, batch_headers = read_multi_ranges(path, target_sheet, ranges)
                    if batch_headers:
                        _header_cache[str(target_sheet)] = batch_headers
                        _header_cache_loaded = True
                else:
                    _ensure_header_cache()

                for ri_idx, ri in enumerate(ranges):
                    if batch_dfs is not None:
                        df, oob_warning = batch_dfs[ri_idx]
                    else:
                        df, oob_warning = _read_single_range(
                            path, target_sheet, ri, no_header, effective_limit, offset
                        )
                    df = apply_compact(df, compact)
                    if sort and sort in df.columns:
                        df = df.sort(sort, descending=descending)

                    # Resolve column letters to row-1 header names
                    column_map = None
                    if headers and not no_header and ri.get("start"):
                        sheet_headers = _header_cache.get(str(target_sheet), [])
                        if sheet_headers:
                            column_map = {}
                            for col_letter in df.columns:
                                idx = col_letter_to_index(col_letter)
                                if idx < len(sheet_headers):
                                    column_map[col_letter] = sheet_headers[idx]
                            rename_map = {
                                letter: name for letter, name in column_map.items() if name
                            }
                            df = df.rename(rename_map)

                    rows = _df_to_serialisable_rows(df, precision)
                    rows = _apply_date_conversion(rows, df, path, target_sheet)

                    range_str = (
                        f"{ri['start']}:{ri['end']}" if ri.get("end") else ri.get("start", "")
                    )
                    entry = {
                        "range": range_str,
                        "sheet": sheet_name,
                        "headers": df.columns,
                        "data": rows,
                        "row_count": len(df),
                    }
                    if column_map:
                        entry["column_map"] = column_map
                    if oob_warning:
                        entry["warning"] = oob_warning
                    # Detect uncached formulas (cached per-sheet)
                    if not formulas and len(df) > 0:
                        _add_formula_hint_cached(
                            entry, df, path, target_sheet, available, _formula_check_cache
                        )
                    results.append(entry)
            else:
                # No range — full sheet read per sheet
                df = read_sheet_data(
                    filepath=path,
                    sheet_name=target_sheet,
                    skip_rows=offset,
                    n_rows=effective_limit,
                    no_header=no_header,
                )
                df = apply_compact(df, compact)
                if sort and sort in df.columns:
                    df = df.sort(sort, descending=descending)

                rows = _df_to_serialisable_rows(df, precision)
                rows = _apply_date_conversion(rows, df, path, target_sheet)

                results.append(
                    {
                        "range": sheet_name,
                        "sheet": sheet_name,
                        "headers": df.columns,
                        "data": rows,
                        "row_count": len(df),
                    }
                )

        elapsed_ms = round((time.perf_counter() - start) * 1000, 1)
        result: dict[str, Any] = {
            "results": results,
            "total_ranges": len(results),
            "compact": compact,
            "read_time_ms": elapsed_ms,
        }
        if should_include_meta():
            result["file_size_human"] = file_size_human(path)
        output_spreadsheet_data(result)
        return

    # --- Single-range path (existing behaviour, backward compatible) ---
    target_sheet = target_sheets[0]
    range_info = ranges[0] if ranges else None

    oob_warning = None
    if range_info and range_info["start"]:
        df, oob_warning = _read_single_range(
            path, target_sheet, range_info, no_header, effective_limit, offset
        )
    else:
        df = read_sheet_data(
            filepath=path,
            sheet_name=target_sheet,
            skip_rows=offset,
            n_rows=effective_limit,
            no_header=no_header,
        )

    df = apply_compact(df, compact)

    # Resolve column letters to row-1 header names when --headers is used
    column_map = None
    if headers and not no_header and range_info and range_info.get("start"):
        try:
            sheet_headers = get_sheet_headers(path, target_sheet)
            column_map = {}
            for col_letter in df.columns:
                idx = col_letter_to_index(col_letter)
                if idx < len(sheet_headers):
                    column_map[col_letter] = sheet_headers[idx]
            rename_map = {letter: name for letter, name in column_map.items() if name}
            df = df.rename(rename_map)
        except Exception:
            pass  # Header resolution is best-effort; don't break reads

    if sort and sort in df.columns:
        df = df.sort(sort, descending=descending)

    elapsed_ms = round((time.perf_counter() - start) * 1000, 1)

    if format_ == "csv":
        _output_csv(df)
        return

    sheet_name_str = target_sheet if isinstance(target_sheet, str) else available[target_sheet]
    range_str = range_ or f"{sheet_name_str}"

    rows = _df_to_serialisable_rows(df, precision)
    rows = _apply_date_conversion(rows, df, path, target_sheet)

    result: dict[str, Any] = {
        "range": range_str,
        "dimensions": {"rows": len(df), "cols": len(df.columns)},
        "headers": df.columns,
        "data": rows,
        "row_count": len(df),
        "truncated": len(df) >= effective_limit,
        "backend": "polars+fastexcel",
        "read_time_ms": elapsed_ms,
    }
    if should_include_meta():
        result["file_size_human"] = file_size_human(path)
    if column_map:
        result["column_map"] = column_map
    if oob_warning:
        result["warning"] = oob_warning

    # Detect uncached formulas: flag empty cells that actually hold formulas
    # so agents don't misinterpret them as missing data
    if not formulas and len(df) > 0:
        _add_formula_hint(result, df, path, target_sheet, available)

    output_spreadsheet_data(result)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _find_empty_col_indices(df: pl.DataFrame) -> list[int]:
    """Return column indices that contain null or empty-string values."""
    empty_indices = []
    for i, col in enumerate(df.columns):
        null_count = df[col].null_count()
        empty_str_count = 0
        if df[col].dtype in (pl.Utf8, pl.String):
            empty_str_count = (df[col] == "").sum()
        if null_count > 0 or empty_str_count > 0:
            empty_indices.append(i)
    return empty_indices


def _add_formula_hint(
    result: dict[str, Any],
    df: pl.DataFrame,
    path: Path,
    target_sheet: str | int,
    available: list[str],
) -> None:
    """Add has_uncached_formulas hint to a read result if applicable."""
    empty_cols = _find_empty_col_indices(df)
    if not empty_cols:
        return
    from agent_xlsx.adapters.polars_adapter import detect_uncached_formulas

    sheet_name = target_sheet if isinstance(target_sheet, str) else available[target_sheet]
    if detect_uncached_formulas(path, sheet_name, empty_cols):
        result["has_uncached_formulas"] = True
        result["hint"] = (
            "Some cells contain formulas whose cached values are empty "
            "(not yet recalculated). Use --formulas to see formula strings, "
            "or run 'recalc' to compute values."
        )


def _add_formula_hint_cached(
    entry: dict[str, Any],
    df: pl.DataFrame,
    path: Path,
    target_sheet: str | int,
    available: list[str],
    cache: dict[str, bool],
) -> None:
    """Like _add_formula_hint but caches the result per-sheet."""
    sheet_name = target_sheet if isinstance(target_sheet, str) else available[target_sheet]
    if sheet_name not in cache:
        empty_cols = _find_empty_col_indices(df)
        if not empty_cols:
            cache[sheet_name] = False
        else:
            from agent_xlsx.adapters.polars_adapter import detect_uncached_formulas

            cache[sheet_name] = detect_uncached_formulas(path, sheet_name, empty_cols)
    if cache[sheet_name]:
        entry["has_uncached_formulas"] = True
        entry["hint"] = (
            "Some cells contain formulas whose cached values are empty. "
            "Use --formulas to see formula strings."
        )


def _read_single_range(
    path: Path,
    target_sheet: str | int,
    range_info: ParsedRange,
    no_header: bool,
    effective_limit: int,
    offset: int,
) -> tuple[pl.DataFrame, str | None]:
    """Read a single parsed range from a sheet.

    Returns (DataFrame, warning) where warning is set when the sheet has
    fewer columns than the requested range (out-of-bounds columns).
    """
    warning = None

    if range_info and range_info["start"]:
        start_col = "".join(c for c in range_info["start"] if c.isalpha())
        start_row = int("".join(c for c in range_info["start"] if c.isdigit()))

        if range_info["end"]:
            end_col = "".join(c for c in range_info["end"] if c.isalpha())
            end_row = int("".join(c for c in range_info["end"] if c.isdigit()))
        else:
            end_col = start_col
            end_row = start_row

        start_col_idx = col_letter_to_index(start_col)
        end_col_idx = col_letter_to_index(end_col)

        # Clamp end column to actual sheet width to avoid fastexcel OOB errors
        from agent_xlsx.adapters.polars_adapter import get_sheet_dimensions
        from agent_xlsx.utils.validation import index_to_col_letter

        try:
            dims = get_sheet_dimensions(path, target_sheet)
            sheet_max_col_idx = dims["cols"] - 1  # 0-based
            if end_col_idx > sheet_max_col_idx:
                clamped_end_idx = max(sheet_max_col_idx, start_col_idx)
                omitted = end_col_idx - clamped_end_idx
                if omitted > 0:
                    actual_last_letter = index_to_col_letter(clamped_end_idx)
                    warning = (
                        f"Requested through column {end_col} but sheet only has data "
                        f"through {actual_last_letter}. {omitted} column(s) omitted."
                    )
                end_col_idx = clamped_end_idx
        except Exception:
            pass  # Dimension check is best-effort; let read_exact_range handle it

        df = read_exact_range(
            filepath=path,
            sheet_name=target_sheet,
            start_col_idx=start_col_idx,
            end_col_idx=end_col_idx,
            start_row=start_row,
            end_row=end_row,
        )

        return df, warning

    return read_sheet_data(
        filepath=path,
        sheet_name=target_sheet,
        skip_rows=offset,
        n_rows=effective_limit,
        no_header=no_header,
    ), None


def _apply_date_conversion(
    rows: list[list],
    df: pl.DataFrame,
    path: Path,
    target_sheet: str | int,
) -> list[list]:
    """Best-effort conversion of Excel serial numbers to ISO dates.

    Uses index-based detection so it works in both normal and --no-header mode.
    """
    try:
        sheet_arg = target_sheet if isinstance(target_sheet, str) else None
        date_indices = detect_date_column_indices(str(path), sheet_arg)
        if not date_indices:
            return rows
        for row in rows:
            for idx in date_indices:
                if idx >= len(row):
                    continue
                val = row[idx]
                # --no-header makes all columns String; coerce numeric strings
                if isinstance(val, str):
                    try:
                        val = float(val)
                    except (ValueError, TypeError):
                        continue
                if isinstance(val, (int, float)) and val == val and val > 0:
                    converted = excel_serial_to_isodate(float(val))
                    if converted is not None:
                        row[idx] = converted
    except Exception:
        pass  # date detection is best-effort; don't break reads
    return rows


def _read_with_formulas(
    path: Path,
    range_str: Optional[str],
    sheet: Optional[str],
    limit: int,
    offset: int,
    compact: bool = True,
    precision: int | None = None,
) -> None:
    """Read with formula strings via openpyxl (slower path)."""
    start = time.perf_counter()

    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=False)

    target_sheet = sheet
    range_info = None

    if range_str:
        range_info = parse_range(range_str)
        if range_info["sheet"]:
            target_sheet = range_info["sheet"]

    if target_sheet:
        if target_sheet not in wb.sheetnames:
            raise SheetNotFoundError(target_sheet, wb.sheetnames)
        ws = wb[target_sheet]
    else:
        ws = wb.active
        target_sheet = ws.title

    # Determine cell range
    if range_info and range_info["start"]:
        if range_info["end"]:
            cell_range = f"{range_info['start']}:{range_info['end']}"
        else:
            # Single cell — expand to self-range for consistent openpyxl
            # tuple-of-tuples return (ws["A1"] returns a bare Cell object)
            cell_range = f"{range_info['start']}:{range_info['start']}"
        rows = list(ws[cell_range])
    else:
        rows = list(ws.iter_rows(min_row=1 + offset, max_row=1 + offset + limit))

    cells: list[dict] = []
    for row in rows:
        for cell in row:
            # Skip empty placeholders from openpyxl's read_only mode
            if isinstance(cell, EmptyCell):
                continue
            value = cell.value
            formula = None
            if isinstance(value, str) and value.startswith("="):
                formula = value
                # Try to get computed value from data_only workbook
                value = formula  # We don't have the computed value in non-data_only mode

            if precision is not None and isinstance(value, float):
                value = round(value, precision)

            cells.append(
                {
                    "cell": cell.coordinate,
                    "value": value,
                    "formula": formula,
                }
            )

    wb.close()

    # Strip blank cells (both value and formula null) in compact mode
    if compact:
        cells = [c for c in cells if c["value"] is not None or c["formula"] is not None]

    elapsed_ms = round((time.perf_counter() - start) * 1000, 1)

    result: dict[str, Any] = {
        "range": range_str or target_sheet,
        "cells": cells[: limit * 20],  # Cap cell output
        "cell_count": len(cells),
        "truncated": len(cells) > limit * 20,
        "backend": "openpyxl",
        "read_time_ms": elapsed_ms,
    }
    if should_include_meta():
        result["file_size_human"] = file_size_human(path)

    output_spreadsheet_data(result)


def _output_csv(df) -> None:
    """Output DataFrame as CSV to stdout."""
    import sys

    csv_str = df.write_csv()
    sys.stdout.write(csv_str)


def _df_to_serialisable_rows(df, precision: int | None = None) -> list[list]:
    """Convert DataFrame rows to JSON-serialisable lists."""
    rows = df.rows()
    result = []
    for row in rows:
        clean = []
        for val in row:
            if val is None:
                clean.append(None)
            elif hasattr(val, "isoformat"):
                clean.append(val.isoformat())
            elif isinstance(val, float) and val != val:
                clean.append(None)
            elif precision is not None and isinstance(val, float):
                clean.append(round(val, precision))
            else:
                clean.append(val)
        result.append(clean)
    return result
