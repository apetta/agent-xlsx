"""Read and apply cell formatting."""

import json as json_mod
from typing import Optional

import typer

from agent_xlsx.cli import app
from agent_xlsx.formatters import json_formatter
from agent_xlsx.formatters.json_formatter import output_spreadsheet_data
from agent_xlsx.utils.errors import AgentExcelError, handle_error
from agent_xlsx.utils.validation import ParsedRange, _normalise_shell_ref, validate_file


@app.command("format")
@handle_error
def format_cmd(
    file: str = typer.Argument(..., help="Path to the Excel file"),
    cell: str = typer.Argument(
        ...,
        help="Cell, range, or comma-separated ranges (e.g. 'A1', 'A1:D10', 'A1:C1,B4')",
    ),
    read: bool = typer.Option(False, "--read", help="Read formatting at the cell"),
    font: Optional[str] = typer.Option(
        None,
        "--font",
        help='Font options as JSON (e.g. \'{"bold": true, "size": 14}\')',
    ),
    fill: Optional[str] = typer.Option(
        None,
        "--fill",
        help='Fill options as JSON (e.g. \'{"color": "FFFF00"}\')',
    ),
    border: Optional[str] = typer.Option(
        None,
        "--border",
        help='Border options as JSON (e.g. \'{"style": "thin"}\')',
    ),
    number_format: Optional[str] = typer.Option(
        None,
        "--number-format",
        "--number",
        help='Number format string (e.g. "#,##0.00")',
    ),
    # Shorthand flags — avoid JSON for common formatting operations
    bold: Optional[bool] = typer.Option(None, "--bold/--no-bold", help="Set font bold"),
    italic: Optional[bool] = typer.Option(None, "--italic/--no-italic", help="Set font italic"),
    font_size: Optional[float] = typer.Option(None, "--font-size", help="Font size (points)"),
    font_color: Optional[str] = typer.Option(
        None, "--font-color", help="Font color hex (e.g. FF0000)"
    ),
    fill_color: Optional[str] = typer.Option(
        None, "--fill-color", help="Fill color hex (e.g. FFFF00)"
    ),
    copy_from: Optional[str] = typer.Option(
        None,
        "--copy-from",
        help="Copy formatting from another cell",
    ),
    output: Optional[str] = typer.Option(
        None,
        "--output",
        "-o",
        help="Save to a new file instead of overwriting",
    ),
    sheet: Optional[str] = typer.Option(None, "--sheet", "-s", help="Target sheet name"),
) -> None:
    """Read or apply cell formatting."""
    path = validate_file(file)

    cell = _normalise_shell_ref(cell)

    from agent_xlsx.adapters import openpyxl_adapter as oxl

    # Detect multi-range (comma-separated)
    is_multi = "," in cell
    if is_multi:
        from agent_xlsx.utils.validation import parse_multi_range

        ranges = parse_multi_range(cell)
    else:
        ranges = None

    # --- Read mode ---
    if read:
        if is_multi and ranges:
            target_sheet = _resolve_sheet(cell, sheet, str(path), ranges)
            results = []
            for ri in ranges:
                # Use start cell of each range as representative
                cell_ref = ri["start"]
                result = oxl.get_cell_formatting(str(path), target_sheet, cell_ref)
                range_str = f"{ri['start']}:{ri['end']}" if ri.get("end") else ri["start"]
                results.append({"range": range_str, "formatting": result})
            output_spreadsheet_data({"results": results, "total_ranges": len(results)})
            return

        read_sheet = sheet
        read_cell = cell
        if "!" in cell:
            parts = cell.split("!", 1)
            read_sheet = parts[0]
            read_cell = parts[1]
        if not read_sheet:
            read_sheet = _default_sheet(str(path))
        result = oxl.get_cell_formatting(str(path), read_sheet, read_cell)
        output_spreadsheet_data(result)
        return

    # --- Copy mode ---
    if copy_from:
        if is_multi and ranges:
            target_sheet = _resolve_sheet(cell, sheet, str(path), ranges)
            # When --output is used, subsequent iterations must read from the output
            # file (not the original source) to preserve earlier range formatting.
            # Use the adapter's actual save path (output_file) since non-writable
            # extensions like .xls are auto-converted to .xlsx by the adapter.
            working_path = str(path)
            for ri in ranges:
                range_str = f"{ri['start']}:{ri['end']}" if ri.get("end") else ri["start"]
                copy_result = oxl.copy_formatting(
                    working_path,
                    sheet_name=target_sheet,
                    source_ref=copy_from,
                    target_ref=range_str,
                    output_path=output,
                )
                if output:
                    working_path = copy_result.get("output_file", output)
            json_formatter.output(
                {
                    "status": "success",
                    "source": copy_from,
                    "ranges_formatted": len(ranges),
                }
            )
            return

        target_sheet = sheet
        target_cell = cell
        if "!" in cell:
            parts = cell.split("!", 1)
            target_sheet = parts[0]
            target_cell = parts[1]
        if not target_sheet:
            target_sheet = _default_sheet(str(path))
        result = oxl.copy_formatting(
            str(path),
            sheet_name=target_sheet,
            source_ref=copy_from,
            target_ref=target_cell,
            output_path=output,
        )
        json_formatter.output(result)
        return

    # --- Apply mode ---
    has_shorthand = any(v is not None for v in [bold, italic, font_size, font_color, fill_color])
    has_formatting = any([font, fill, border, number_format]) or has_shorthand
    if not has_formatting:
        raise AgentExcelError(
            "MISSING_FORMAT",
            "No formatting options provided",
            [
                "Use --read to read formatting",
                "Use --bold, --italic, --font-size, --font-color, --fill-color for common styles",
                "Use --font, --fill, --border, or --number-format for full JSON control",
                "Use --copy-from to copy formatting from another cell",
            ],
        )

    font_opts = _parse_json_opt(font, "font") if font else None
    fill_opts = _parse_json_opt(fill, "fill") if fill else None
    border_opts = _parse_json_opt(border, "border") if border else None

    # Merge shorthand flags into parsed JSON opts (shorthands layer on top)
    if any(v is not None for v in [bold, italic, font_size, font_color]):
        if font_opts is None:
            font_opts = {}
        if bold is not None:
            font_opts["bold"] = bold
        if italic is not None:
            font_opts["italic"] = italic
        if font_size is not None:
            font_opts["size"] = font_size
        if font_color is not None:
            font_opts["color"] = font_color

    if fill_color is not None:
        if fill_opts is None:
            fill_opts = {}
        fill_opts["color"] = fill_color
        fill_opts.setdefault("fill_type", "solid")

    if is_multi and ranges:
        target_sheet = _resolve_sheet(cell, sheet, str(path), ranges)
        total_formatted = 0
        # When --output is used, subsequent iterations must read from the output
        # file (not the original source) to preserve earlier range formatting.
        # Use the adapter's actual save path (output_file) since non-writable
        # extensions like .xls are auto-converted to .xlsx by the adapter.
        working_path = str(path)
        for ri in ranges:
            range_str = f"{ri['start']}:{ri['end']}" if ri.get("end") else ri["start"]
            result = oxl.apply_formatting(
                working_path,
                sheet_name=target_sheet,
                cell_ref=range_str,
                font_opts=font_opts,
                fill_opts=fill_opts,
                border_opts=border_opts,
                number_format=number_format,
                output_path=output,
            )
            if output:
                working_path = result.get("output_file", output)
            total_formatted += result.get("cells_formatted", 0)
        json_formatter.output(
            {
                "status": "success",
                "ranges_formatted": len(ranges),
                "total_cells_formatted": total_formatted,
            }
        )
        return

    target_sheet = sheet
    target_cell = cell
    if "!" in cell:
        parts = cell.split("!", 1)
        target_sheet = parts[0]
        target_cell = parts[1]
    if not target_sheet:
        target_sheet = _default_sheet(str(path))

    result = oxl.apply_formatting(
        str(path),
        sheet_name=target_sheet,
        cell_ref=target_cell,
        font_opts=font_opts,
        fill_opts=fill_opts,
        border_opts=border_opts,
        number_format=number_format,
        output_path=output,
    )
    json_formatter.output(result)


def _resolve_sheet(
    cell: str,
    sheet: str | None,
    filepath: str,
    ranges: list[ParsedRange],
) -> str:
    """Resolve the target sheet for multi-range operations.

    Priority: explicit --sheet flag > sheet prefix in first range > default (first sheet).
    """
    if sheet:
        return sheet
    first_sheet = ranges[0].get("sheet") if ranges else None
    if first_sheet:
        return str(first_sheet)
    return _default_sheet(filepath)


def _parse_json_opt(json_str: str, label: str) -> dict:
    """Parse a JSON string option, raising a clean error on failure."""
    try:
        data = json_mod.loads(json_str)
    except json_mod.JSONDecodeError as exc:
        raise AgentExcelError(
            "INVALID_JSON",
            f"Failed to parse --{label} JSON: {exc}",
            [f"Provide valid JSON for --{label}"],
        )
    if not isinstance(data, dict):
        raise AgentExcelError(
            "INVALID_JSON",
            f"--{label} must be a JSON object, not {type(data).__name__}",
            [f"e.g. --{label} '{{\"bold\": true}}'"],
        )
    return data


def _default_sheet(filepath: str) -> str:
    """Return the name of the first sheet in the workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True)
    try:
        return wb.sheetnames[0]
    finally:
        wb.close()
