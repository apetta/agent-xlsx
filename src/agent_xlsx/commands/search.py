"""Search across workbook for values using Polars, with openpyxl fallback for formulas."""

import time
from pathlib import Path
from typing import Dict, List, Optional

import typer

from agent_xlsx.adapters.polars_adapter import search_values
from agent_xlsx.cli import app
from agent_xlsx.formatters.json_formatter import output, output_spreadsheet_data
from agent_xlsx.utils.constants import MAX_SEARCH_RESULTS
from agent_xlsx.utils.errors import InvalidRegexError, SheetNotFoundError, handle_error
from agent_xlsx.utils.validation import validate_file


@app.command()
@handle_error
def search(
    file: str = typer.Argument(..., help="Path to the Excel file"),
    query: str = typer.Argument(..., help="Search term or pattern"),
    regex: bool = typer.Option(False, "--regex", "-r", help="Treat query as regex pattern"),
    ignore_case: bool = typer.Option(False, "--ignore-case", "-i", help="Case-insensitive search"),
    sheet: Optional[str] = typer.Option(None, "--sheet", "-s", help="Search in specific sheet"),
    in_formulas: bool = typer.Option(
        False, "--in-formulas", help="Search in formula strings (uses openpyxl, slower)"
    ),
    no_header: bool = typer.Option(
        False,
        "--no-header",
        help="Treat row 1 as data, use column letters. "
        "Use for non-tabular sheets like P&L reports.",
    ),
) -> None:
    """Search for values across the workbook.

    Uses Polars for fast value searching by default.
    Use --in-formulas to search formula strings via openpyxl.
    """
    path = validate_file(file)

    if regex:
        import re

        try:
            re.compile(query)
        except re.error as e:
            raise InvalidRegexError(query, str(e))

    start = time.perf_counter()

    if in_formulas:
        matches = _search_formulas(path, query, sheet, regex, ignore_case)
    else:
        try:
            matches = search_values(
                filepath=path,
                query=query,
                sheet_name=sheet,
                regex=regex,
                ignore_case=ignore_case,
                no_header=no_header,
            )
        except Exception as e:
            # Polars uses a Rust regex engine that rejects some Python-valid
            # patterns (e.g. backreferences, lookaheads). Surface these as
            # structured errors rather than raw tracebacks.
            if regex and "regex" in str(e).lower():
                reason = str(e).strip().rsplit("\n", 1)[-1].strip()
                raise InvalidRegexError(query, reason) from None
            raise

    elapsed_ms = round((time.perf_counter() - start) * 1000, 1)

    result = {
        "query": query,
        "match_count": len(matches),
        "matches": matches,
        "truncated": len(matches) >= MAX_SEARCH_RESULTS,
        "search_time_ms": elapsed_ms,
    }

    output_spreadsheet_data(result)


def _search_formulas(
    path: Path,
    query: str,
    sheet_name: Optional[str],
    regex: bool,
    ignore_case: bool,
) -> List[Dict]:
    """Search formula strings via openpyxl (slower path)."""
    import re

    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=False)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(sheet_name, wb.sheetnames)
        target_sheets = [sheet_name]
    else:
        target_sheets = wb.sheetnames

    flags = re.IGNORECASE if ignore_case else 0
    if regex:
        pattern = re.compile(query, flags)
    else:
        search_term = query.lower() if ignore_case else query

    matches: list[dict] = []

    for ws_name in target_sheets:
        ws = wb[ws_name]
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if not (cell.value and isinstance(cell.value, str) and cell.value.startswith("=")):
                    continue

                formula = cell.value
                matched = False

                if regex:
                    matched = bool(pattern.search(formula))
                elif ignore_case:
                    matched = search_term in formula.lower()
                else:
                    matched = query in formula

                if matched:
                    matches.append(
                        {
                            "sheet": ws_name,
                            "cell": cell.coordinate,
                            "formula": formula,
                        }
                    )
                    if len(matches) >= MAX_SEARCH_RESULTS:
                        wb.close()
                        return matches

    wb.close()
    return matches
