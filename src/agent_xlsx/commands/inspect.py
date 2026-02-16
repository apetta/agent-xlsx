"""Detailed workbook inspection command.

Drills into specific workbook elements: sheets, ranges, named ranges,
charts, VBA presence, formatting, comments, conditional formatting,
data validation, and hyperlinks.
"""

from typing import Optional

import typer

from agent_xlsx.cli import app
from agent_xlsx.formatters import json_formatter
from agent_xlsx.formatters.token_optimizer import cap_list, summarise_formulas
from agent_xlsx.utils.constants import MAX_FORMULA_CELLS, MAX_LOCATIONS
from agent_xlsx.utils.errors import handle_error
from agent_xlsx.utils.validation import parse_range, validate_file


@app.command("inspect")
@handle_error
def inspect_cmd(
    file: str = typer.Argument(..., help="Path to the Excel file"),
    sheet: Optional[str] = typer.Option(None, "--sheet", "-s", help="Inspect a specific sheet"),
    range_: Optional[str] = typer.Option(
        None,
        "--range",
        "-r",
        help="Inspect a cell range (e.g. 'A1:C10')",
    ),
    names: bool = typer.Option(False, "--names", help="Inspect named ranges"),
    charts: bool = typer.Option(False, "--charts", help="Inspect charts"),
    vba: bool = typer.Option(False, "--vba", help="Inspect VBA modules"),
    format_cell: Optional[str] = typer.Option(
        None,
        "--format",
        "-f",
        help="Inspect formatting at a cell",
    ),
    comments: bool = typer.Option(False, "--comments", help="Inspect comments/notes"),
    conditional: Optional[str] = typer.Option(
        None,
        "--conditional",
        help="Inspect conditional formatting for a range",
    ),
    validation: Optional[str] = typer.Option(
        None,
        "--validation",
        help="Inspect data validation for a sheet",
    ),
    hyperlinks: Optional[str] = typer.Option(
        None,
        "--hyperlinks",
        help="Inspect hyperlinks for a sheet",
    ),
) -> None:
    """Detailed inspection of workbook elements."""
    path = validate_file(file)

    from agent_xlsx.adapters import openpyxl_adapter as oxl

    # --- Specific inspection modes ---

    if format_cell:
        # Determine the sheet for formatting inspection
        fmt_sheet = sheet
        if "!" in format_cell:
            parts = format_cell.split("!", 1)
            fmt_sheet = parts[0]
            format_cell = parts[1]
        if not fmt_sheet:
            fmt_sheet = _default_sheet(str(path))
        result = oxl.get_cell_formatting(str(path), fmt_sheet, format_cell)
        json_formatter.output(result)
        return

    if range_:
        parsed = parse_range(range_)
        r_sheet = parsed["sheet"] or sheet or _default_sheet(str(path))
        formulas = oxl.get_range_formulas(str(path), r_sheet, parsed["start"], parsed["end"])
        summary = summarise_formulas(formulas, MAX_FORMULA_CELLS)
        summary["range"] = range_
        summary["sheet"] = r_sheet
        json_formatter.output(summary)
        return

    if names:
        meta = oxl.get_workbook_metadata(str(path))
        json_formatter.output(
            {
                "named_ranges": meta["named_ranges"],
                "count": meta["named_range_count"],
            }
        )
        return

    if charts:
        meta = oxl.get_workbook_metadata(str(path))
        charts_data: list[dict] = []
        for s in meta["sheets"]:
            if s["chart_count"] > 0:
                charts_data.append(
                    {
                        "sheet": s["name"],
                        "chart_count": s["chart_count"],
                    }
                )
        json_formatter.output(
            {
                "charts": charts_data,
                "total_chart_count": meta["total_chart_count"],
            }
        )
        return

    if vba:
        meta = oxl.get_workbook_metadata(str(path))
        json_formatter.output(
            {
                "has_vba": meta["has_vba"],
            }
        )
        return

    if comments:
        target_sheet = sheet or _default_sheet(str(path))
        comment_list = oxl.get_comments(str(path), target_sheet)
        capped = cap_list(comment_list, MAX_LOCATIONS)
        json_formatter.output(
            {
                "sheet": target_sheet,
                "comments": capped["items"],
                "comment_count": capped["total"],
                "truncated": capped["truncated"],
            }
        )
        return

    if conditional:
        # conditional can be "Sheet1!A1:D100" or just a sheet name
        cf_sheet = conditional
        if "!" in conditional:
            cf_sheet = conditional.split("!", 1)[0]
        cf_rules = oxl.get_conditional_formatting(str(path), cf_sheet)
        capped = cap_list(cf_rules, MAX_LOCATIONS)
        json_formatter.output(
            {
                "sheet": cf_sheet,
                "rules": capped["items"],
                "rule_count": capped["total"],
                "truncated": capped["truncated"],
            }
        )
        return

    if validation:
        val_rules = oxl.get_data_validations(str(path), validation)
        capped = cap_list(val_rules, MAX_LOCATIONS)
        json_formatter.output(
            {
                "sheet": validation,
                "validations": capped["items"],
                "validation_count": capped["total"],
                "truncated": capped["truncated"],
            }
        )
        return

    if hyperlinks:
        link_list = oxl.get_hyperlinks(str(path), hyperlinks)
        capped = cap_list(link_list, MAX_LOCATIONS)
        json_formatter.output(
            {
                "sheet": hyperlinks,
                "hyperlinks": capped["items"],
                "hyperlink_count": capped["total"],
                "truncated": capped["truncated"],
            }
        )
        return

    if sheet:
        # Full sheet inspection — everything in one pass
        result = oxl.get_full_sheet_inspection(str(path), sheet)
        json_formatter.output(result)
        return

    # --- Default: inspect all sheets with summary metadata ---
    meta = oxl.get_workbook_metadata(str(path))
    json_formatter.output(meta)


def _default_sheet(filepath: str) -> str:
    """Return the name of the first sheet in the workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True)
    try:
        return wb.sheetnames[0]
    finally:
        wb.close()
