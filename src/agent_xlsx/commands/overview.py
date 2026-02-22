"""Structural metadata overview using openpyxl for formula/chart/VBA detection."""

import re
import time
from pathlib import Path

import typer

from agent_xlsx.cli import app
from agent_xlsx.formatters.json_formatter import output, output_spreadsheet_data
from agent_xlsx.utils.constants import MAX_FORMULA_PATTERNS, VBA_EXTENSIONS
from agent_xlsx.utils.errors import handle_error
from agent_xlsx.utils.validation import file_size_bytes, validate_file


@app.command()
@handle_error
def overview(
    file: str = typer.Argument(..., help="Path to the Excel file"),
    include_formulas: bool = typer.Option(
        False, "--include-formulas", help="Include sample formulas per sheet"
    ),
    include_formatting: bool = typer.Option(
        False, "--include-formatting", help="Include formatting presence info"
    ),
) -> None:
    """Structural metadata overview of a workbook.

    Focuses on elements fastexcel cannot detect: formulas, charts, VBA, tables.
    For data-focused profiling, prefer 'probe' — it is faster and richer.
    """
    path = validate_file(file)
    start = time.perf_counter()

    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=False, data_only=False)

    import fastexcel

    try:
        fe_reader = fastexcel.read_excel(str(path))
    except Exception:
        fe_reader = None

    sheets_result: list[dict] = []
    total_formula_count = 0
    total_chart_count = 0

    for idx, ws_name in enumerate(wb.sheetnames):
        ws = wb[ws_name]
        sheet_info: dict = {
            "name": ws_name,
            "index": idx,
            "dimensions": ws.dimensions or "A1",
            "row_count": ws.max_row or 0,
            "col_count": ws.max_column or 0,
        }

        # Cross-reference with fastexcel for actual data dimensions
        if fe_reader is not None:
            try:
                fe_sheet = fe_reader.load_sheet(ws_name, n_rows=0)
                sheet_info["data_rows"] = fe_sheet.total_height
                sheet_info["data_cols"] = fe_sheet.width
            except Exception:
                pass

        # Formula detection — scan cells for strings starting with '='
        formula_count = 0
        # Pattern deduplication: normalise row numbers to group repetitive formulas
        pattern_map: dict[str, dict] = {}  # normalised_pattern → info dict
        if ws.max_row and ws.max_column:
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
                        if include_formulas and len(pattern_map) < MAX_FORMULA_PATTERNS:
                            formula_text = cell.value
                            # Normalise by stripping row numbers → pattern shape
                            normalised = re.sub(r"\d+", "{N}", formula_text)
                            if normalised not in pattern_map:
                                pattern_map[normalised] = {
                                    "pattern": normalised,
                                    "example_cell": cell.coordinate,
                                    "example": formula_text,
                                    "count": 1,
                                }
                            else:
                                pattern_map[normalised]["count"] += 1

        sheet_info["has_formulas"] = formula_count > 0
        sheet_info["formula_count"] = formula_count
        total_formula_count += formula_count

        if include_formulas and pattern_map:
            sheet_info["sample_formulas"] = list(pattern_map.values())

        # Charts
        chart_count = len(ws._charts) if hasattr(ws, "_charts") else 0
        sheet_info["has_charts"] = chart_count > 0
        sheet_info["chart_count"] = chart_count
        total_chart_count += chart_count

        # Tables
        tables = list(ws.tables.values()) if hasattr(ws, "tables") else []
        sheet_info["has_tables"] = len(tables) > 0
        if tables:
            sheet_info["table_names"] = [t.name for t in tables]

        # Formatting presence (lightweight check)
        if include_formatting:
            has_merged = bool(ws.merged_cells.ranges) if hasattr(ws, "merged_cells") else False
            sheet_info["has_merged_cells"] = has_merged
            if has_merged:
                sheet_info["merged_cell_count"] = len(ws.merged_cells.ranges)

        sheets_result.append(sheet_info)

    # Named ranges
    named_ranges = list(wb.defined_names.keys()) if hasattr(wb.defined_names, "keys") else []
    named_range_count = len(named_ranges)

    # VBA detection
    has_vba = path.suffix.lower() in VBA_EXTENSIONS
    vba_module_count = 0
    if has_vba and wb.vba_archive:
        vba_module_count = len(
            [
                n
                for n in wb.vba_archive.namelist()
                if n.endswith(".bas") or n.endswith(".cls") or n.endswith(".frm")
            ]
        )

    wb.close()

    elapsed_ms = round((time.perf_counter() - start) * 1000, 1)

    result = {
        "file": Path(file).name,
        "size_bytes": file_size_bytes(path),
        "overview_time_ms": elapsed_ms,
        "sheets": sheets_result,
        "named_ranges": named_ranges,
        "named_range_count": named_range_count,
        "has_vba": has_vba,
        "vba_module_count": vba_module_count,
        "total_formula_count": total_formula_count,
        "total_chart_count": total_chart_count,
    }

    output_spreadsheet_data(result)
