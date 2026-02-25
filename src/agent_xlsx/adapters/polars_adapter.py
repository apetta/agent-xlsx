"""Primary data adapter using Polars + fastexcel (Calamine) for blazingly fast Excel reads."""

from __future__ import annotations

import contextlib
import os
import time
from pathlib import Path
from typing import Any

import fastexcel
import polars as pl

from agent_xlsx.utils.constants import (
    CHUNK_SIZE_ROWS,
    CHUNK_THRESHOLD_BYTES,
    MAX_SAMPLE_ROWS,
    MAX_SEARCH_RESULTS,
    SEARCH_FULLLOAD_FILE_SIZE_THRESHOLD,
)
from agent_xlsx.utils.dates import (
    detect_date_column_indices_batch,
    detect_date_columns,
    excel_serial_to_isodate,
)
from agent_xlsx.utils.errors import SheetNotFoundError
from agent_xlsx.utils.validation import (
    ParsedRange,
    col_letter_to_index,
    file_size_bytes,
    file_size_human,
    index_to_col_letter,
)


def get_sheet_names(filepath: str | Path) -> list[str]:
    """Return sheet names from a workbook via fastexcel (instant, no data parsed)."""
    reader = fastexcel.read_excel(str(filepath))
    return reader.sheet_names


def get_sheet_dimensions(filepath: str | Path, sheet_name: str | int = 0) -> dict[str, Any]:
    """Return sheet dimensions without loading data (n_rows=0 trick)."""
    reader = fastexcel.read_excel(str(filepath))
    sheet = reader.load_sheet(sheet_name, n_rows=0)
    return {
        "name": sheet.name,
        "rows": sheet.total_height,
        "cols": sheet.width,
        "headers": [c.name for c in sheet.available_columns()],
        "visible": sheet.visible == "visible",
    }


def get_sheet_headers(filepath: str | Path, sheet_name: str | int = 0) -> list[str]:
    """Return row-1 header names for a sheet (zero data parsing)."""
    fpath = str(filepath)
    resolved = _resolve_sheet_name(fpath, sheet_name)
    reader = fastexcel.read_excel(fpath)
    sheet = reader.load_sheet(resolved, n_rows=0)
    return [c.name for c in sheet.available_columns()]


def read_sheet_data(
    filepath: str | Path,
    sheet_name: str | int = 0,
    skip_rows: int = 0,
    n_rows: int | None = None,
    use_columns: str | list[str] | None = None,
    no_header: bool = False,
) -> pl.DataFrame:
    """Read sheet data into a Polars DataFrame via fastexcel (zero-copy Arrow).

    For large files (>100MB), automatically chunks the read to stay within memory budget.
    When *no_header* is True, row 1 is treated as data and columns are named
    with Excel letters (A, B, C, …).
    """
    fpath = str(filepath)
    size = file_size_bytes(fpath)

    # Polars read_excel requires string sheet names — resolve integer indices
    resolved_name = _resolve_sheet_name(fpath, sheet_name)

    read_opts: dict[str, Any] = {}
    if skip_rows:
        read_opts["skip_rows"] = skip_rows
    if n_rows is not None:
        read_opts["n_rows"] = n_rows
    if use_columns is not None:
        read_opts["use_columns"] = use_columns
    if no_header:
        read_opts["header_row"] = None

    with _suppress_stderr():
        if size < CHUNK_THRESHOLD_BYTES or n_rows is not None:
            # Direct read — fastexcel handles this efficiently
            df = pl.read_excel(fpath, sheet_name=resolved_name, read_options=read_opts or None)
        else:
            # Chunked read for very large files
            df = _read_chunked(fpath, resolved_name, skip_rows, n_rows)

    if no_header:
        col_letters = [index_to_col_letter(i) for i in range(len(df.columns))]
        df = df.rename(dict(zip(df.columns, col_letters)))

    return df


def read_exact_range(
    filepath: str | Path,
    sheet_name: str | int,
    start_col_idx: int,
    end_col_idx: int,
    start_row: int,
    end_row: int,
) -> pl.DataFrame:
    """Read an exact Excel range with no header assumption.

    All row params use Excel 1-based conventions. Column indices are 0-based
    (A=0, B=1, ...). Returns a DataFrame with Excel column-letter headers.
    """
    fpath = str(filepath)
    resolved_name = _resolve_sheet_name(fpath, sheet_name)

    # With header_row=None, fastexcel names columns __UNNAMED__N so letter-based
    # column selection doesn't work.  Use 0-based integer indices instead.
    use_cols: list[int] = list(range(start_col_idx, end_col_idx + 1))

    reader = fastexcel.read_excel(fpath)
    with _suppress_stderr():
        sheet = reader.load_sheet(
            resolved_name,
            header_row=None,
            skip_rows=start_row - 1,
            n_rows=end_row - start_row + 1,
            use_columns=use_cols,
        )
        df = pl.DataFrame(sheet)

    # Rename generic __UNNAMED__N columns to Excel letters
    num_cols = end_col_idx - start_col_idx + 1
    col_letters = [index_to_col_letter(start_col_idx + i) for i in range(num_cols)]
    if len(df.columns) <= len(col_letters):
        df = df.rename(dict(zip(df.columns, col_letters[: len(df.columns)])))

    return df


def read_multi_ranges(
    filepath: str | Path,
    sheet_name: str | int,
    ranges: list[ParsedRange],
) -> tuple[list[tuple[pl.DataFrame, str | None]], list[str]]:
    """Read multiple ranges from a single sheet with one file open.

    Opens the fastexcel reader once, loads the full sheet DataFrame once,
    then slices each range from memory. O(1) per range after initial load.

    Each range dict has keys: start (e.g. "A184608"), end (e.g. "D184608").
    Returns (results, col_headers) where results is a list of
    (DataFrame, warning_or_None) in input order, and col_headers is the
    row-1 header names for column_map resolution.
    """
    import re as _re

    fpath = str(filepath)
    reader = fastexcel.read_excel(fpath)
    resolved_name = _resolve_sheet_name(fpath, sheet_name)

    # Load full sheet with headers — significantly faster than header_row=None
    # on large files (~3s vs ~5s). Save header names before renaming to letters.
    with _suppress_stderr():
        full_sheet = reader.load_sheet(resolved_name)
        full_df = pl.DataFrame(full_sheet)

    col_headers = list(full_df.columns)
    sheet_width = len(full_df.columns)

    # Rename to Excel letters for positional range slicing
    all_letters = [index_to_col_letter(i) for i in range(sheet_width)]
    full_df = full_df.rename(dict(zip(full_df.columns, all_letters)))

    results: list[tuple[pl.DataFrame, str | None]] = []

    for ri in ranges:
        warning = None
        start_str = ri.get("start", "")
        end_str = ri.get("end", "")

        if not start_str:
            # No range specified — return full sheet
            results.append((full_df, None))
            continue

        m_start = _re.match(r"^([A-Z]+)(\d+)$", start_str.upper())
        if not m_start:
            results.append((pl.DataFrame(), f"Invalid range start: {start_str}"))
            continue

        start_col_idx = col_letter_to_index(m_start.group(1))
        start_row = int(m_start.group(2))

        if end_str:
            m_end = _re.match(r"^([A-Z]+)(\d+)$", end_str.upper())
            if m_end:
                end_col_idx = col_letter_to_index(m_end.group(1))
                end_row = int(m_end.group(2))
            else:
                end_col_idx = start_col_idx
                end_row = start_row
        else:
            end_col_idx = start_col_idx
            end_row = start_row

        # Clamp end column to actual sheet width
        sheet_max_col_idx = sheet_width - 1
        if end_col_idx > sheet_max_col_idx:
            clamped_end_idx = max(sheet_max_col_idx, start_col_idx)
            omitted = end_col_idx - clamped_end_idx
            if omitted > 0:
                actual_last_letter = index_to_col_letter(clamped_end_idx)
                end_letter = m_end.group(1) if end_str and m_end else start_str
                warning = (
                    f"Requested through column {end_letter} but sheet only has data "
                    f"through {actual_last_letter}. {omitted} column(s) omitted."
                )
            end_col_idx = clamped_end_idx

        # Slice rows — header row was consumed by fastexcel, so Excel row N
        # maps to df row (N - 2). Row 1 = header, row 2 = df[0].
        n_rows = end_row - start_row + 1
        df_slice = full_df.slice(max(start_row - 2, 0), n_rows)

        # Select columns
        col_letters_range = [index_to_col_letter(i) for i in range(start_col_idx, end_col_idx + 1)]
        available_cols = [c for c in col_letters_range if c in df_slice.columns]
        if available_cols:
            df_slice = df_slice.select(available_cols)

        results.append((df_slice, warning))

    return results, col_headers


def _read_chunked(
    filepath: str,
    sheet_name: str | int,
    skip_rows: int,
    n_rows: int | None,
) -> pl.DataFrame:
    """Read in chunks using fastexcel's skip_rows + n_rows for very large files."""
    reader = fastexcel.read_excel(filepath)
    probe = reader.load_sheet(sheet_name, n_rows=0)
    total_rows = probe.total_height

    target_rows = n_rows if n_rows is not None else total_rows - skip_rows
    chunks: list[pl.DataFrame] = []
    rows_read = 0

    offset = skip_rows
    while rows_read < target_rows and offset < total_rows:
        chunk_n = min(CHUNK_SIZE_ROWS, target_rows - rows_read)
        sheet = reader.load_sheet(
            sheet_name,
            skip_rows=offset,
            n_rows=chunk_n,
        )
        chunk_df = pl.DataFrame(sheet)
        chunks.append(chunk_df)
        rows_read += len(chunk_df)
        offset += chunk_n

    if not chunks:
        # Return empty DataFrame with proper schema
        sheet = reader.load_sheet(sheet_name, n_rows=0)
        return pl.DataFrame(sheet)

    return pl.concat(chunks)


def probe_workbook(
    filepath: str | Path,
    sheet_name: str | None = None,
    sample_rows: int = 0,
    stats: bool = False,
    include_types: bool = False,
    no_header: bool = False,
    max_columns: int | None = None,
) -> dict[str, Any]:
    """Ultra-fast workbook profiling via fastexcel + Polars.

    Default is a lean metadata-only probe (sheet names, dimensions, headers) with
    zero data parsing.  Pass ``include_types``, ``sample_rows``, or ``stats`` to
    opt into progressively richer detail.

    When *no_header* is True, row 1 is treated as data and column names
    become Excel letters (A, B, C, …).  This is the correct mode for
    non-tabular spreadsheets like P&L reports and dashboards.
    """
    start = time.perf_counter()
    fpath = str(filepath)
    reader = fastexcel.read_excel(fpath)

    all_sheet_names = reader.sheet_names
    target_sheets = [sheet_name] if sheet_name else all_sheet_names

    # Validate requested sheet exists
    if sheet_name and sheet_name not in all_sheet_names:
        raise SheetNotFoundError(sheet_name, all_sheet_names)

    needs_data = include_types or sample_rows > 0 or stats
    sheets_result: list[dict[str, Any]] = []

    for idx, name in enumerate(all_sheet_names):
        if name not in target_sheets:
            continue

        # Fast metadata via n_rows=0 — no data parsed
        meta = reader.load_sheet(name, n_rows=0)
        sheet_info: dict[str, Any] = {
            "name": name,
            "index": idx,
            "visible": meta.visible == "visible",
            "rows": meta.total_height,
            "cols": meta.width,
        }

        if not needs_data:
            # Lean path — headers from metadata only, zero data parsing
            if no_header:
                n_cols = meta.width
                sheet_info["headers"] = [index_to_col_letter(i) for i in range(n_cols)]
            else:
                sheet_info["headers"] = [c.name for c in meta.available_columns()]
            headers = sheet_info["headers"]
            sheet_info["last_col"] = index_to_col_letter(len(headers) - 1) if headers else "A"
            if not no_header:
                sheet_info["column_map"] = {
                    h: index_to_col_letter(i) for i, h in enumerate(headers)
                }
            sheets_result.append(sheet_info)
            continue

        # Load full sheet data for profiling — suppress fastexcel's stderr
        # "Could not determine dtype" warnings for entirely-null columns.
        with _suppress_stderr():
            if no_header:
                sheet = reader.load_sheet(name, header_row=None)
                df = pl.DataFrame(sheet)
                col_letters = [index_to_col_letter(i) for i in range(len(df.columns))]
                df = df.rename(dict(zip(df.columns, col_letters)))
            else:
                sheet = reader.load_sheet(name)
                df = pl.DataFrame(sheet)

        # Always expose the full schema (all headers)
        sheet_info["headers"] = df.columns
        headers = sheet_info["headers"]
        sheet_info["last_col"] = index_to_col_letter(len(headers) - 1) if headers else "A"
        if not no_header:
            sheet_info["column_map"] = {h: index_to_col_letter(i) for i, h in enumerate(headers)}

        # Optionally limit profiling detail to first N columns
        profile_df = df
        if max_columns is not None and max_columns < len(df.columns):
            profile_df = df.select(df.columns[:max_columns])
            sheet_info["profiled_columns"] = max_columns
            sheet_info["columns_truncated"] = True

        if include_types:
            # Null counts per column — compute first so we can filter other sections
            fully_null_set: set[str] = set()
            if len(profile_df) > 0:
                null_row = profile_df.null_count().row(0)
                all_null_counts = dict(zip(profile_df.columns, null_row))
                sheet_info["null_counts"] = {
                    col: count for col, count in all_null_counts.items() if count < len(profile_df)
                }
                fully_null_set = set(all_null_counts.keys()) - set(sheet_info["null_counts"].keys())
                if fully_null_set:
                    sheet_info["fully_null_columns"] = len(fully_null_set)
            else:
                sheet_info["null_counts"] = {col: 0 for col in profile_df.columns}

            # Column types — omit fully-null columns (zero information)
            sheet_info["column_types"] = {
                col: _polars_dtype_to_str(dtype)
                for col, dtype in zip(profile_df.columns, profile_df.dtypes)
                if col not in fully_null_set
            }

            # Detect date columns masquerading as float64
            float_cols = [
                col for col, t in sheet_info["column_types"].items() if t in ("float64", "float32")
            ]
            if float_cols:
                date_col_names = detect_date_columns(fpath, name)
                for col in float_cols:
                    if col in date_col_names:
                        sheet_info["column_types"][col] = "date"

            # --no-header: all columns are String, so detect dates by index
            if no_header:
                try:
                    date_idx_map = detect_date_column_indices_batch(fpath, [name])
                    for idx in date_idx_map.get(name, set()):
                        col_letter = index_to_col_letter(idx)
                        if col_letter in sheet_info["column_types"]:
                            sheet_info["column_types"][col_letter] = "date"
                except Exception:
                    pass

            # Potential header detection for non-tabular sheets
            if no_header:
                potential = _detect_potential_headers(profile_df)
                if potential:
                    sheet_info["potential_headers"] = potential

        # Sample data (head + tail) — sparse dict format to reduce token waste
        capped_sample = min(sample_rows, MAX_SAMPLE_ROWS)
        if capped_sample > 0 and len(profile_df) > 0:
            date_col_set = {
                col for col, t in sheet_info.get("column_types", {}).items() if t == "date"
            }
            head_rows = _df_to_sparse_rows(profile_df.head(capped_sample))
            tail_rows = _df_to_sparse_rows(profile_df.tail(capped_sample))
            # Convert Excel serial numbers to ISO dates in sparse dicts
            if date_col_set:
                for row in head_rows + tail_rows:
                    for col in date_col_set:
                        val = row.get(col)
                        # --no-header yields string values; coerce numeric strings
                        if isinstance(val, str):
                            try:
                                val = float(val)
                            except (ValueError, TypeError):
                                continue
                        if isinstance(val, (int, float)):
                            row[col] = excel_serial_to_isodate(float(val))
            sheet_info["sample"] = {
                "head": head_rows,
                "tail": tail_rows,
            }

        # Extended statistics via Polars describe()
        if stats and len(profile_df) > 0:
            # Identify fully-null columns to exclude from stats (noise reduction)
            null_counts = sheet_info.get("null_counts", {})
            fully_null_cols = {
                col for col, count in null_counts.items() if count >= len(profile_df)
            }

            date_col_set = {
                col for col, t in sheet_info.get("column_types", {}).items() if t == "date"
            }
            numeric_cols = [
                col
                for col, dtype in zip(profile_df.columns, profile_df.dtypes)
                if dtype.is_numeric() and col not in date_col_set and col not in fully_null_cols
            ]
            if numeric_cols:
                sheet_info["numeric_summary"] = _build_numeric_summary(profile_df, numeric_cols)

            string_cols = [
                col
                for col, dtype in zip(profile_df.columns, profile_df.dtypes)
                if dtype == pl.Utf8
                and col not in fully_null_cols
                and null_counts.get(col, 0) < 0.5 * len(profile_df)
            ]
            if string_cols:
                sheet_info["string_summary"] = _build_string_summary(profile_df, string_cols)

            # Date column summary (min/max as ISO dates)
            if date_col_set:
                date_summary = {}
                for col in date_col_set:
                    series = profile_df[col].drop_nulls()
                    if len(series) > 0:
                        min_val = _safe_scalar(series.min())
                        max_val = _safe_scalar(series.max())
                        if isinstance(min_val, (int, float)) and min_val == min_val:
                            date_summary[col] = {
                                "min": excel_serial_to_isodate(float(min_val)),
                                "max": excel_serial_to_isodate(float(max_val)),
                                "count": len(series),
                            }
                if date_summary:
                    sheet_info["date_summary"] = date_summary

        sheets_result.append(sheet_info)

    elapsed_ms = round((time.perf_counter() - start) * 1000, 1)

    # Workbook-level metadata from fastexcel
    result: dict[str, Any] = {
        "file": Path(filepath).name,
        "size_bytes": file_size_bytes(fpath),
        "file_size_human": file_size_human(fpath),
        "format": Path(filepath).suffix.lstrip(".").lower(),
        "probe_time_ms": elapsed_ms,
        "sheets": sheets_result,
    }

    # Auto-suggest --no-header when most headers are unnamed (lean path only)
    if not no_header and not needs_data:
        all_headers = [h for s in sheets_result for h in s.get("headers", [])]
        if all_headers:
            unnamed_count = sum(1 for h in all_headers if str(h).startswith("__UNNAMED__"))
            if unnamed_count / len(all_headers) > 0.5:
                result["hint"] = (
                    "Most headers are unnamed. Consider --no-header for column-letter headers."
                )

    # Named ranges and tables (fast via fastexcel)
    try:
        defined = reader.defined_names()
        result["named_ranges"] = [n.name for n in defined] if defined else []
    except Exception:
        result["named_ranges"] = []

    try:
        result["tables"] = reader.table_names() or []
    except Exception:
        result["tables"] = []

    # VBA detection — check file extension
    result["has_vba"] = Path(filepath).suffix.lower() in {".xlsm", ".xlsb"}

    return result


def search_values(
    filepath: str | Path,
    query: str,
    sheet_name: str | None = None,
    regex: bool = False,
    ignore_case: bool = False,
    no_header: bool = False,
    columns: str | None = None,
    limit: int = MAX_SEARCH_RESULTS,
    range_spec: ParsedRange | None = None,
) -> list[dict[str, Any]]:
    """Search for values across sheets using Polars string matching.

    Returns list of match dicts: {sheet, column, row, cell, value}.
    When *no_header* is True, columns are named with Excel letters and
    cell references use simple 1-based row numbering (no header offset).

    Optional filters:
    - columns: comma-separated column letters or header names to restrict search
    - limit: max results to return (default: MAX_SEARCH_RESULTS)
    - range_spec: parsed range dict {sheet, start, end} to restrict row/col scope
    """
    fpath = str(filepath)
    reader = fastexcel.read_excel(fpath)
    all_sheets = reader.sheet_names

    if sheet_name:
        if sheet_name not in all_sheets:
            raise SheetNotFoundError(sheet_name, all_sheets)
        target_sheets = [sheet_name]
    else:
        target_sheets = all_sheets

    matches: list[dict[str, Any]] = []

    # Pre-parse range bounds if provided
    range_row_offset = None
    r_start_col_idx = None
    r_start_row = None
    r_end_row = None
    r_use_cols = None
    if range_spec:
        import re as _re

        start_str = range_spec["start"]
        m = _re.match(r"^([A-Z]+)(\d+)$", start_str.upper())
        if m:
            r_start_col_idx = col_letter_to_index(m.group(1))
            r_start_row = int(m.group(2))
        end_str = range_spec["end"]
        if end_str:
            m2 = _re.match(r"^([A-Z]+)(\d+)$", end_str.upper())
            if m2:
                r_end_col_idx = col_letter_to_index(m2.group(1))
                r_end_row = int(m2.group(2))
                if r_start_col_idx is not None:
                    r_use_cols = list(range(r_start_col_idx, r_end_col_idx + 1))
        else:
            # Single cell range
            r_end_row = r_start_row
            r_use_cols = [r_start_col_idx] if r_start_col_idx is not None else None

    # Best-effort date column detection — single workbook open for all sheets
    try:
        all_date_cols = detect_date_column_indices_batch(fpath, target_sheets)
    except Exception:
        all_date_cols = {}

    for name in target_sheets:
        # Cache row-1 header names from the full load so column filter
        # resolution doesn't need an additional load_sheet call (~2.6s saved).
        _cached_col_headers: list[str] | None = None

        with _suppress_stderr():
            if range_spec and r_start_row is not None and r_end_row is not None:
                n_range_rows = r_end_row - r_start_row + 1

                # Full-load-then-slice heuristic: for large files, loading the
                # full sheet and slicing in memory is faster than fastexcel's
                # sequential skip_rows. Calamine must parse the entire xlsx
                # regardless (no row index in the format), so skip_rows just
                # adds per-row overhead with no benefit.
                use_fullload = file_size_bytes(fpath) >= SEARCH_FULLLOAD_FILE_SIZE_THRESHOLD

                if use_fullload:
                    # Loading with headers is ~40% faster than header_row=None
                    # on large files, so prefer it and adjust the row offset.
                    if no_header:
                        sheet = reader.load_sheet(name, header_row=None)
                        df = pl.DataFrame(sheet)
                        row_offset = r_start_row - 1
                    else:
                        sheet = reader.load_sheet(name)
                        df = pl.DataFrame(sheet)
                        _cached_col_headers = list(df.columns)
                        row_offset = r_start_row - 2  # header row consumed
                    # Rename columns to Excel letters for positional matching
                    all_letters = [index_to_col_letter(i) for i in range(len(df.columns))]
                    df = df.rename(dict(zip(df.columns, all_letters)))
                    # Slice to the requested row range
                    df = df.slice(max(row_offset, 0), n_range_rows)
                    # Select only the requested columns
                    if r_use_cols is not None:
                        col_subset = [index_to_col_letter(c) for c in r_use_cols]
                        col_subset = [c for c in col_subset if c in df.columns]
                        if col_subset:
                            df = df.select(col_subset)
                else:
                    # Standard range-scoped load via fastexcel skip_rows
                    sheet = reader.load_sheet(
                        name,
                        header_row=None,
                        skip_rows=r_start_row - 1,
                        n_rows=n_range_rows,
                        use_columns=r_use_cols,
                    )
                    df = pl.DataFrame(sheet)
                    base_col = r_start_col_idx or 0
                    n_cols = len(df.columns)
                    col_letters = [index_to_col_letter(base_col + i) for i in range(n_cols)]
                    df = df.rename(dict(zip(df.columns, col_letters[: len(df.columns)])))

                range_row_offset = r_start_row
            elif no_header:
                sheet = reader.load_sheet(name, header_row=None)
                df = pl.DataFrame(sheet)
                col_letters = [index_to_col_letter(i) for i in range(len(df.columns))]
                df = df.rename(dict(zip(df.columns, col_letters)))
            else:
                sheet = reader.load_sheet(name)
                df = pl.DataFrame(sheet)

        if len(df) == 0:
            continue

        date_col_indices = all_date_cols.get(name, set())

        # Apply column filter if specified
        if columns:
            from agent_xlsx.utils.validation import resolve_column_filter

            # When range or no-header is active, df columns are letters —
            # use cached headers from the full load if available, otherwise
            # load row-1 headers from the already-open reader.
            col_headers = _cached_col_headers
            if col_headers is None and (range_spec or no_header):
                try:
                    hdr_sheet = reader.load_sheet(name, n_rows=0)
                    col_headers = [c.name for c in hdr_sheet.available_columns()]
                except Exception:
                    col_headers = None
            search_cols = resolve_column_filter(columns, list(df.columns), headers=col_headers)
        else:
            search_cols = list(df.columns)

        for col in search_cols:
            # Cast column to string for searching
            str_col = df[col].cast(pl.Utf8, strict=False)

            if regex:
                if ignore_case:
                    mask = str_col.str.contains(f"(?i){query}")
                else:
                    mask = str_col.str.contains(query)
            elif ignore_case:
                mask = str_col.str.to_lowercase().str.contains(query.lower(), literal=True)
            else:
                mask = str_col.str.contains(query, literal=True)

            # Get matching row indices
            match_indices = df.with_row_index("__row_idx__").filter(mask)["__row_idx__"].to_list()

            for row_idx in match_indices:
                cell_value = df[col][row_idx]
                # Convert to Python native
                if hasattr(cell_value, "item"):
                    cell_value = cell_value.item()

                col_idx = df.columns.index(col)
                # For range-scoped search, column letter comes from the DataFrame
                # (already renamed to correct Excel letters). For normal search,
                # compute from the DataFrame position.
                if range_spec and r_start_col_idx is not None:
                    col_letter = index_to_col_letter(r_start_col_idx + col_idx)
                    abs_col_idx = r_start_col_idx + col_idx
                else:
                    col_letter = index_to_col_letter(col_idx)
                    abs_col_idx = col_idx

                # Convert date serial numbers to ISO strings
                # --no-header makes all columns String; coerce numeric strings
                if abs_col_idx in date_col_indices and isinstance(cell_value, str):
                    try:
                        cell_value = float(cell_value)
                    except (ValueError, TypeError):
                        pass
                if (
                    abs_col_idx in date_col_indices
                    and isinstance(cell_value, (int, float))
                    and cell_value == cell_value  # not NaN
                    and cell_value > 0
                ):
                    converted = excel_serial_to_isodate(float(cell_value))
                    if converted is not None:
                        cell_value = converted

                # Compute Excel row number
                if range_row_offset is not None:
                    excel_row = range_row_offset + row_idx
                elif no_header:
                    excel_row = row_idx + 1
                else:
                    excel_row = row_idx + 2
                cell_ref = f"{col_letter}{excel_row}"

                matches.append(
                    {
                        "sheet": name,
                        "column": col_letter,
                        "row": excel_row,
                        "cell": cell_ref,
                        "value": cell_value,
                    }
                )

                if len(matches) >= limit:
                    return matches

    return matches


def detect_uncached_formulas(
    filepath: str | Path,
    sheet_name: str,
    empty_col_indices: list[int],
) -> bool:
    """Check if a sheet has formula cells with empty cached values.

    Opens the file with openpyxl read_only=True, data_only=False and samples
    cells at the intersection of empty columns and the first rows. Returns
    True if any cell contains a formula (starts with "=").

    Performance: <100ms for typical files (read_only mode + early exit).
    """
    from agent_xlsx.utils.constants import FORMULA_CHECK_SAMPLE_SIZE

    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(filepath), read_only=True, data_only=False)
        try:
            if sheet_name not in wb.sheetnames:
                return False
            ws = wb[sheet_name]
            empty_set = set(empty_col_indices)
            checked = 0
            for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row or 100, 100)):
                for cell in row:
                    if cell.column is not None and (cell.column - 1) in empty_set:
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            return True
                        checked += 1
                        if checked >= FORMULA_CHECK_SAMPLE_SIZE:
                            return False
            return False
        finally:
            wb.close()
    except Exception:
        return False  # Formula detection is best-effort


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


@contextlib.contextmanager
def _suppress_stderr():
    """Suppress stderr output from Rust/fastexcel dtype warnings."""
    devnull_fd = os.open(os.devnull, os.O_WRONLY)
    old_stderr_fd = os.dup(2)
    try:
        os.dup2(devnull_fd, 2)
        yield
    finally:
        os.dup2(old_stderr_fd, 2)
        os.close(devnull_fd)
        os.close(old_stderr_fd)


def _resolve_sheet_name(filepath: str, sheet_name: str | int) -> str:
    """Resolve an integer sheet index to a string name. Polars requires string names."""
    if isinstance(sheet_name, str):
        return sheet_name
    reader = fastexcel.read_excel(filepath)
    names = reader.sheet_names
    if sheet_name < 0 or sheet_name >= len(names):
        raise SheetNotFoundError(str(sheet_name), names)
    return names[sheet_name]


def _polars_dtype_to_str(dtype: pl.DataType) -> str:
    """Convert Polars dtype to a short human-readable string."""
    mapping = {
        pl.Utf8: "string",
        pl.String: "string",
        pl.Int8: "int8",
        pl.Int16: "int16",
        pl.Int32: "int32",
        pl.Int64: "int64",
        pl.UInt8: "uint8",
        pl.UInt16: "uint16",
        pl.UInt32: "uint32",
        pl.UInt64: "uint64",
        pl.Float32: "float32",
        pl.Float64: "float64",
        pl.Boolean: "boolean",
        pl.Date: "date",
        pl.Time: "time",
        pl.Null: "null",
    }
    # Handle parameterised types like Datetime(...)
    for base_type, label in mapping.items():
        if dtype == base_type:
            return label

    dtype_str = str(dtype).lower()
    if "datetime" in dtype_str:
        return "datetime"
    if "duration" in dtype_str:
        return "duration"

    return str(dtype).lower()


def _df_to_rows(df: pl.DataFrame) -> list[list[Any]]:
    """Convert DataFrame to a list of rows, handling special types for JSON serialisation."""
    rows = df.rows()
    result = []
    for row in rows:
        clean_row = []
        for val in row:
            if val is None:
                clean_row.append(None)
            elif hasattr(val, "isoformat"):
                clean_row.append(val.isoformat())
            elif isinstance(val, float) and val != val:
                clean_row.append(None)  # NaN → null
            else:
                clean_row.append(val)
        result.append(clean_row)
    return result


def _is_numeric_string(val: str) -> bool:
    """Return True if the string represents a number (e.g. '0', '71847.38', '-5')."""
    try:
        float(val.replace(",", ""))
        return True
    except (ValueError, AttributeError):
        return False


def _detect_potential_headers(df: pl.DataFrame, max_rows: int = 10) -> list[dict[str, Any]]:
    """Detect rows that look like headers in headerless (non-tabular) data.

    Scans the first *max_rows* rows and returns candidates where ≥30% of
    columns are non-null AND ≥60% of those non-null cells are short strings
    (≤20 chars) — the typical signature of month names, column labels, etc.
    """
    candidates: list[dict[str, Any]] = []
    scan_rows = min(max_rows, len(df))
    columns = df.columns

    for i in range(scan_rows):
        row = df.row(i)
        non_null_vals = [
            (col, val)
            for col, val in zip(columns, row)
            if val is not None and not (isinstance(val, float) and val != val)
        ]

        if not non_null_vals:
            continue

        # At least 30% of columns must be non-null
        if len(non_null_vals) / len(columns) < 0.3:
            continue

        # At least 60% of non-null cells must be short non-numeric strings
        # (headers are text labels like "Dec", "% sales", not "0" or "71847.38")
        short_strings = sum(
            1
            for _, val in non_null_vals
            if isinstance(val, str) and len(val) <= 20 and not _is_numeric_string(val)
        )
        if short_strings / len(non_null_vals) < 0.6:
            continue

        # Build sparse dict of values
        values: dict[str, Any] = {}
        for col, val in non_null_vals:
            if hasattr(val, "isoformat"):
                values[col] = val.isoformat()
            else:
                values[col] = val
        # Row number is 1-based (Excel convention for no_header mode)
        candidates.append({"row": i + 1, "values": values})

    return candidates


def _df_to_sparse_rows(df: pl.DataFrame) -> list[dict[str, Any]]:
    """Convert DataFrame rows to sparse dicts — only non-null cells are included.

    This drastically reduces token output for wide sheets with many null
    separator columns (e.g. 46 elements → 5 keys).  Long string values
    are truncated to cap sample section size.
    """
    from agent_xlsx.utils.constants import SAMPLE_VALUE_MAX_CHARS

    columns = df.columns
    rows = df.rows()
    result: list[dict[str, Any]] = []
    for row in rows:
        sparse: dict[str, Any] = {}
        for col_name, val in zip(columns, row):
            if val is None:
                continue
            if isinstance(val, float) and val != val:
                continue  # NaN → skip
            if hasattr(val, "isoformat"):
                sparse[col_name] = val.isoformat()
            elif isinstance(val, str) and len(val) > SAMPLE_VALUE_MAX_CHARS:
                sparse[col_name] = val[:SAMPLE_VALUE_MAX_CHARS] + "..."
            else:
                sparse[col_name] = val
        result.append(sparse)
    return result


def _build_numeric_summary(df: pl.DataFrame, numeric_cols: list[str]) -> dict[str, Any]:
    """Build numeric summary statistics using Polars."""
    summary: dict[str, Any] = {}
    for col in numeric_cols:
        series = df[col].drop_nulls()
        if len(series) == 0:
            continue
        summary[col] = {
            "min": _safe_scalar(series.min()),
            "max": _safe_scalar(series.max()),
            "mean": _safe_scalar(series.mean()),
            "median": _safe_scalar(series.median()),
            "std": _safe_scalar(series.std()),
        }
    return summary


def _build_string_summary(df: pl.DataFrame, string_cols: list[str]) -> dict[str, Any]:
    """Build string column summary with unique counts and top values.

    Free-text columns (avg length > threshold) get a compact summary with
    unique count and avg length instead of full top-5 values, since listing
    3K-char paragraphs as top values wastes tokens without useful insight.
    Shorter categorical columns get truncated top values.
    """
    from agent_xlsx.utils.constants import FREETEXT_AVG_LENGTH_THRESHOLD, STRING_VALUE_MAX_CHARS

    summary: dict[str, Any] = {}
    for col in string_cols:
        series = df[col].drop_nulls()
        if len(series) == 0:
            continue

        n_unique = series.n_unique()
        # str.len_chars().mean() always returns int/float/None for a UInt32 series,
        # but Polars' generic .mean() return type includes date/timedelta etc.
        avg_len_raw = series.str.len_chars().mean()
        avg_len: float | None = None
        if isinstance(avg_len_raw, (int, float)):
            avg_len = float(avg_len_raw)

        # Free-text columns: emit a compact descriptor instead of top values
        if avg_len is not None and avg_len > FREETEXT_AVG_LENGTH_THRESHOLD:
            summary[col] = {
                "unique": n_unique,
                "avg_length": round(avg_len),
                "type": "free_text",
            }
            continue

        # Categorical columns: top-5 values, truncated to cap token output
        top_raw = series.value_counts(sort=True).head(5).get_column(col).to_list()
        top = [
            (v[:STRING_VALUE_MAX_CHARS] + "...")
            if isinstance(v, str) and len(v) > STRING_VALUE_MAX_CHARS
            else v
            for v in top_raw
        ]
        summary[col] = {
            "unique": n_unique,
            "top_values": top,
        }
    return summary


def _safe_scalar(val: Any) -> Any:
    """Convert Polars scalar to JSON-safe Python type."""
    if val is None:
        return None
    if isinstance(val, float) and val != val:
        return None
    if hasattr(val, "item"):
        return val.item()
    if isinstance(val, float) and val == int(val):
        return int(val)
    if isinstance(val, float):
        return round(val, 6)
    return val
