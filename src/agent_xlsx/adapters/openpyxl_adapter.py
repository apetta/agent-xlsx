"""openpyxl adapter for metadata reads, formatting inspection, and write operations.

This adapter is the fallback for metadata that fastexcel cannot provide (formulas,
formatting, charts, comments, conditional formatting, data validation) and the
primary backend for all write operations.

Uses read_only=True where possible for memory efficiency, but falls back to normal
mode for metadata that read-only worksheets do not expose (charts, tables, merged
cells, freeze panes, conditional formatting, data validation, hyperlinks).

Use keep_vba=True when working with .xlsm files.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from agent_xlsx.utils.constants import (
    MAX_FORMULA_CELLS,
    MAX_LOCATIONS,
    WRITABLE_EXTENSIONS,
)

# ---------------------------------------------------------------------------
# Workbook-level metadata
# ---------------------------------------------------------------------------


def get_workbook_metadata(filepath: str | Path) -> dict[str, Any]:
    """Return full structural metadata for a workbook.

    Covers: sheets, formulas, charts, VBA presence, tables, named ranges,
    merged cells count per sheet.

    Uses normal (non-read-only) mode because read-only worksheets do not
    expose charts, tables, or merged cell metadata.
    """
    filepath = Path(filepath)
    wb = load_workbook(str(filepath), data_only=False)
    try:
        sheets_info: list[dict[str, Any]] = []
        total_formulas = 0
        total_charts = 0

        for ws in wb.worksheets:
            formula_count = 0
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1

            chart_count = len(ws._charts)
            table_names = [t.name for t in ws.tables.values()]

            total_formulas += formula_count
            total_charts += chart_count

            sheet_data: dict[str, Any] = {
                "name": ws.title,
                "dimensions": ws.dimensions or "",
                "has_formulas": formula_count > 0,
                "formula_count": formula_count,
                "has_charts": chart_count > 0,
                "chart_count": chart_count,
                "has_tables": len(table_names) > 0,
                "table_names": table_names,
            }
            sheets_info.append(sheet_data)

        defined_names = list(wb.defined_names)

        result: dict[str, Any] = {
            "sheets": sheets_info,
            "named_ranges": defined_names,
            "named_range_count": len(defined_names),
            "has_vba": wb.vba_archive is not None,
            "total_formula_count": total_formulas,
            "total_chart_count": total_charts,
        }
        return result
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Sheet-level metadata
# ---------------------------------------------------------------------------


def _extract_chart_title(chart: Any) -> str | None:
    """Extract plain text from an openpyxl chart title object."""
    if not chart.title:
        return None
    title_obj = chart.title
    # Try to get text from rich text paragraphs
    if hasattr(title_obj, "text") and title_obj.text:
        tx = title_obj.text
        if hasattr(tx, "rich") and tx.rich and hasattr(tx.rich, "paragraphs"):
            parts = []
            for para in tx.rich.paragraphs:
                for run in para.r:
                    if run.t:
                        parts.append(run.t)
            text = "".join(parts)
            if text and text != "None":
                return text
    # Fallback: body attribute
    if hasattr(title_obj, "body") and title_obj.body:
        return str(title_obj.body)
    return None


def get_sheet_metadata(filepath: str | Path, sheet_name: str) -> dict[str, Any]:
    """Return detailed metadata for a single sheet.

    Includes: formulas (sampled), merged cells, tables, charts, freeze panes.

    Uses normal mode (not read-only) to access charts, tables, merged cells,
    and freeze panes.
    """
    wb = load_workbook(str(filepath), data_only=False)
    try:
        ws = wb[sheet_name]
        formula_cells: list[dict[str, str]] = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells.append(
                        {
                            "cell": f"{get_column_letter(cell.column)}{cell.row}",
                            "formula": cell.value,
                        }
                    )

        chart_info = []
        for chart in ws._charts:
            chart_info.append(
                {
                    "title": _extract_chart_title(chart),
                    "type": chart.__class__.__name__,
                }
            )

        table_info = []
        for t in ws.tables.values():
            table_info.append(
                {
                    "name": t.name,
                    "range": t.ref,
                    "header_row": t.headerRowCount != 0 if hasattr(t, "headerRowCount") else True,
                    "total_row": (
                        t.totalsRowCount is not None and t.totalsRowCount > 0
                        if hasattr(t, "totalsRowCount")
                        else False
                    ),
                }
            )

        merged = [str(m) for m in ws.merged_cells.ranges]

        freeze = None
        if ws.freeze_panes:
            freeze = str(ws.freeze_panes)

        result: dict[str, Any] = {
            "sheet": sheet_name,
            "dimensions": ws.dimensions or "",
            "formula_count": len(formula_cells),
            "sample_formulas": formula_cells[:MAX_FORMULA_CELLS],
            "formulas_truncated": len(formula_cells) > MAX_FORMULA_CELLS,
            "merged_regions": merged[:MAX_LOCATIONS],
            "merged_count": len(merged),
            "tables": table_info,
            "charts": chart_info,
            "freeze_panes": freeze,
        }
        return result
    finally:
        wb.close()


def get_full_sheet_inspection(filepath: str | Path, sheet_name: str) -> dict[str, Any]:
    """Return comprehensive inspection data for a single sheet in one pass.

    Combines the data from get_sheet_metadata, get_comments,
    get_conditional_formatting, get_data_validations, and get_hyperlinks
    into a single workbook open/close for efficiency.
    """
    wb = load_workbook(str(filepath), read_only=False, data_only=False)
    try:
        ws = wb[sheet_name]

        # --- Formulas ---
        formula_cells: list[dict[str, str]] = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells.append(
                        {
                            "cell": f"{get_column_letter(cell.column)}{cell.row}",
                            "formula": cell.value,
                        }
                    )

        # --- Charts ---
        chart_info = []
        for chart in ws._charts:
            chart_info.append(
                {
                    "title": _extract_chart_title(chart),
                    "type": chart.__class__.__name__,
                }
            )

        # --- Tables ---
        table_info = []
        for t in ws.tables.values():
            table_info.append(
                {
                    "name": t.name,
                    "range": t.ref,
                    "header_row": t.headerRowCount != 0 if hasattr(t, "headerRowCount") else True,
                    "total_row": (
                        t.totalsRowCount is not None and t.totalsRowCount > 0
                        if hasattr(t, "totalsRowCount")
                        else False
                    ),
                }
            )

        # --- Merged cells ---
        merged = [str(m) for m in ws.merged_cells.ranges]

        # --- Freeze panes ---
        freeze = str(ws.freeze_panes) if ws.freeze_panes else None

        # --- Comments ---
        comment_items: list[dict[str, Any]] = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    comment_items.append(
                        {
                            "cell": f"{get_column_letter(cell.column)}{cell.row}",
                            "author": cell.comment.author,
                            "text": cell.comment.text,
                        }
                    )

        # --- Conditional formatting ---
        cf_rules: list[dict[str, Any]] = []
        for cf_range, rule_list in ws.conditional_formatting._cf_rules:
            for rule in rule_list:
                rule_data: dict[str, Any] = {
                    "range": str(cf_range),
                    "type": rule.type,
                    "priority": rule.priority,
                }
                if rule.operator:
                    rule_data["operator"] = rule.operator
                if rule.formula:
                    rule_data["formula"] = list(rule.formula)
                cf_rules.append(rule_data)

        # --- Data validation ---
        dv_rules: list[dict[str, Any]] = []
        if ws.data_validations and ws.data_validations.dataValidation:
            for dv in ws.data_validations.dataValidation:
                val_data: dict[str, Any] = {
                    "range": str(dv.sqref),
                    "type": dv.type,
                    "allow_blank": dv.allow_blank,
                }
                if dv.operator:
                    val_data["operator"] = dv.operator
                if dv.formula1:
                    val_data["formula1"] = dv.formula1
                if dv.formula2:
                    val_data["formula2"] = dv.formula2
                if dv.error:
                    val_data["error_message"] = dv.error
                if dv.errorTitle:
                    val_data["error_title"] = dv.errorTitle
                if dv.prompt:
                    val_data["prompt"] = dv.prompt
                dv_rules.append(val_data)

        # --- Hyperlinks ---
        link_items: list[dict[str, Any]] = []
        for hl in ws._hyperlinks:
            link_items.append(
                {
                    "cell": hl.ref,
                    "target": hl.target,
                    "display": hl.display,
                    "tooltip": hl.tooltip,
                }
            )

        # --- Build result with capping ---
        result: dict[str, Any] = {
            "sheet": sheet_name,
            "dimensions": ws.dimensions or "",
            "formulas": {
                "count": len(formula_cells),
                "sample": formula_cells[:MAX_FORMULA_CELLS],
                "truncated": len(formula_cells) > MAX_FORMULA_CELLS,
            },
            "merged_cells": {
                "count": len(merged),
                "regions": merged[:MAX_LOCATIONS],
                "truncated": len(merged) > MAX_LOCATIONS,
            },
            "tables": table_info,
            "charts": chart_info,
            "freeze_panes": freeze,
            "comments": {
                "count": len(comment_items),
                "items": comment_items[:MAX_LOCATIONS],
                "truncated": len(comment_items) > MAX_LOCATIONS,
            },
            "conditional_formatting": {
                "count": len(cf_rules),
                "rules": cf_rules[:MAX_LOCATIONS],
                "truncated": len(cf_rules) > MAX_LOCATIONS,
            },
            "data_validation": {
                "count": len(dv_rules),
                "rules": dv_rules[:MAX_LOCATIONS],
                "truncated": len(dv_rules) > MAX_LOCATIONS,
            },
            "hyperlinks": {
                "count": len(link_items),
                "items": link_items[:MAX_LOCATIONS],
                "truncated": len(link_items) > MAX_LOCATIONS,
            },
        }
        return result
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Cell formatting
# ---------------------------------------------------------------------------


def _colour_to_hex(colour: Any) -> str | None:
    """Extract hex colour string from an openpyxl Color object."""
    if colour is None:
        return None
    if colour.type == "rgb" and colour.rgb:
        rgb = str(colour.rgb)
        # Strip leading alpha if present (AARRGGBB -> RRGGBB)
        if len(rgb) == 8:
            return rgb[2:]
        return rgb
    if colour.type == "theme":
        return f"theme:{colour.theme}"
    if colour.type == "indexed" and colour.indexed is not None:
        return f"indexed:{colour.indexed}"
    return None


def _border_side(side: Any) -> dict[str, Any] | None:
    """Convert an openpyxl Side to a dict."""
    if side is None or side.style is None:
        return None
    return {
        "style": side.style,
        "color": _colour_to_hex(side.color),
    }


def get_cell_formatting(filepath: str | Path, sheet_name: str, cell_ref: str) -> dict[str, Any]:
    """Return detailed formatting for a single cell."""
    wb = load_workbook(str(filepath), data_only=True)
    try:
        ws = wb[sheet_name]
        cell = ws[cell_ref]

        font = cell.font
        fill = cell.fill
        alignment = cell.alignment
        border = cell.border

        result: dict[str, Any] = {
            "cell": cell_ref,
            "value": cell.value,
            "font": {
                "name": font.name,
                "size": font.size,
                "bold": font.bold,
                "italic": font.italic,
                "underline": font.underline,
                "color": _colour_to_hex(font.color),
            },
            "fill": {
                "type": fill.fill_type,
                "color": _colour_to_hex(fill.fgColor) if fill.fgColor else None,
            },
            "border": {
                "top": _border_side(border.top),
                "bottom": _border_side(border.bottom),
                "left": _border_side(border.left),
                "right": _border_side(border.right),
            },
            "alignment": {
                "horizontal": alignment.horizontal,
                "vertical": alignment.vertical,
                "wrap_text": alignment.wrap_text,
                "text_rotation": alignment.text_rotation,
            },
            "number_format": cell.number_format,
        }
        return result
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Range formulas
# ---------------------------------------------------------------------------


def get_range_formulas(
    filepath: str | Path, sheet_name: str, start: str, end: str | None
) -> list[dict[str, Any]]:
    """Return formula strings for cells in a range."""
    wb = load_workbook(str(filepath), data_only=False)
    try:
        ws = wb[sheet_name]
        range_str = f"{start}:{end}" if end else start
        formula_cells: list[dict[str, Any]] = []
        for row in ws[range_str]:
            if not isinstance(row, tuple):
                row = (row,)
            for cell in row:
                val = cell.value
                formula_cells.append(
                    {
                        "cell": f"{get_column_letter(cell.column)}{cell.row}",
                        "value": val
                        if not (isinstance(val, str) and val.startswith("="))
                        else None,
                        "formula": val if isinstance(val, str) and val.startswith("=") else None,
                    }
                )
        return formula_cells
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Comments
# ---------------------------------------------------------------------------


def get_comments(filepath: str | Path, sheet_name: str) -> list[dict[str, Any]]:
    """Return all cell comments for a sheet."""
    wb = load_workbook(str(filepath))
    try:
        ws = wb[sheet_name]
        comments: list[dict[str, Any]] = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    comments.append(
                        {
                            "cell": f"{get_column_letter(cell.column)}{cell.row}",
                            "author": cell.comment.author,
                            "text": cell.comment.text,
                        }
                    )
        return comments
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Conditional formatting
# ---------------------------------------------------------------------------


def get_conditional_formatting(filepath: str | Path, sheet_name: str) -> list[dict[str, Any]]:
    """Return conditional formatting rules for a sheet."""
    wb = load_workbook(str(filepath))
    try:
        ws = wb[sheet_name]
        rules: list[dict[str, Any]] = []
        for cf_range, rule_list in ws.conditional_formatting._cf_rules:
            for rule in rule_list:
                rule_data: dict[str, Any] = {
                    "range": str(cf_range),
                    "type": rule.type,
                    "priority": rule.priority,
                }
                if rule.operator:
                    rule_data["operator"] = rule.operator
                if rule.formula:
                    rule_data["formula"] = list(rule.formula)
                rules.append(rule_data)
        return rules
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Data validations
# ---------------------------------------------------------------------------


def get_data_validations(filepath: str | Path, sheet_name: str) -> list[dict[str, Any]]:
    """Return data validation rules for a sheet."""
    wb = load_workbook(str(filepath))
    try:
        ws = wb[sheet_name]
        validations: list[dict[str, Any]] = []
        if ws.data_validations and ws.data_validations.dataValidation:
            for dv in ws.data_validations.dataValidation:
                val_data: dict[str, Any] = {
                    "range": str(dv.sqref),
                    "type": dv.type,
                    "allow_blank": dv.allow_blank,
                }
                if dv.operator:
                    val_data["operator"] = dv.operator
                if dv.formula1:
                    val_data["formula1"] = dv.formula1
                if dv.formula2:
                    val_data["formula2"] = dv.formula2
                if dv.error:
                    val_data["error_message"] = dv.error
                if dv.errorTitle:
                    val_data["error_title"] = dv.errorTitle
                if dv.prompt:
                    val_data["prompt"] = dv.prompt
                validations.append(val_data)
        return validations
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Hyperlinks
# ---------------------------------------------------------------------------


def get_hyperlinks(filepath: str | Path, sheet_name: str) -> list[dict[str, Any]]:
    """Return all hyperlinks for a sheet."""
    wb = load_workbook(str(filepath))
    try:
        ws = wb[sheet_name]
        links: list[dict[str, Any]] = []
        for hl in ws._hyperlinks:
            links.append(
                {
                    "cell": hl.ref,
                    "target": hl.target,
                    "display": hl.display,
                    "tooltip": hl.tooltip,
                }
            )
        return links
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Write operations
# ---------------------------------------------------------------------------


def write_cells(
    filepath: str | Path,
    sheet_name: str | None,
    data: list[dict[str, Any]],
    output_path: str | Path | None = None,
) -> dict[str, Any]:
    """Write values or formulas to cells.

    Each item in `data` should have: cell (str), value (any), and optionally
    number_format (str).

    Preserves existing workbook content. For .xlsm files, VBA is preserved.
    """
    filepath = Path(filepath)
    keep_vba = filepath.suffix.lower() == ".xlsm"
    wb = load_workbook(str(filepath), keep_vba=keep_vba)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active

        cells_written = 0
        for item in data:
            cell_ref = item["cell"]
            value = item["value"]
            cell = ws[cell_ref]
            cell.value = value
            if "number_format" in item and item["number_format"]:
                cell.number_format = item["number_format"]
            cells_written += 1

        save_path = Path(output_path) if output_path else filepath
        if save_path.suffix.lower() not in WRITABLE_EXTENSIONS:
            save_path = save_path.with_suffix(".xlsx")
        wb.save(str(save_path))

        return {
            "status": "success",
            "cells_written": cells_written,
            "output_file": str(save_path),
        }
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Apply formatting
# ---------------------------------------------------------------------------


def apply_formatting(
    filepath: str | Path,
    sheet_name: str | None,
    cell_ref: str,
    font_opts: dict[str, Any] | None = None,
    fill_opts: dict[str, Any] | None = None,
    border_opts: dict[str, Any] | None = None,
    number_format: str | None = None,
    output_path: str | Path | None = None,
) -> dict[str, Any]:
    """Apply formatting to a cell or range.

    Accepts optional font, fill, border dicts and number_format string.
    """
    from openpyxl.styles import Border, Font, PatternFill, Side

    filepath = Path(filepath)
    keep_vba = filepath.suffix.lower() == ".xlsm"
    wb = load_workbook(str(filepath), keep_vba=keep_vba)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active

        # Determine if range or single cell
        cells = _resolve_cells(ws, cell_ref)

        for cell in cells:
            if font_opts:
                # Preserve existing colour object unless a new hex colour is provided
                font_colour = cell.font.color
                if "color" in font_opts:
                    font_colour = font_opts["color"]
                cell.font = Font(
                    name=font_opts.get("name", cell.font.name),
                    size=font_opts.get("size", cell.font.size),
                    bold=font_opts.get("bold", cell.font.bold),
                    italic=font_opts.get("italic", cell.font.italic),
                    underline=font_opts.get("underline", cell.font.underline),
                    color=font_colour,
                )

            if fill_opts:
                cell.fill = PatternFill(
                    fill_type=fill_opts.get("type", "solid"),
                    fgColor=fill_opts.get("color", "FFFFFF"),
                )

            if border_opts:
                style = border_opts.get("style", "thin")
                colour = border_opts.get("color", "000000")
                side = Side(style=style, color=colour)
                cell.border = Border(
                    top=side if border_opts.get("top", True) else cell.border.top,
                    bottom=side if border_opts.get("bottom", True) else cell.border.bottom,
                    left=side if border_opts.get("left", True) else cell.border.left,
                    right=side if border_opts.get("right", True) else cell.border.right,
                )

            if number_format:
                cell.number_format = number_format

        save_path = Path(output_path) if output_path else filepath
        wb.save(str(save_path))

        return {
            "status": "success",
            "cells_formatted": len(cells),
            "output_file": str(save_path),
        }
    finally:
        wb.close()


def copy_formatting(
    filepath: str | Path,
    sheet_name: str | None,
    source_ref: str,
    target_ref: str,
    output_path: str | Path | None = None,
) -> dict[str, Any]:
    """Copy formatting from a source cell to a target cell or range."""
    from copy import copy

    filepath = Path(filepath)
    keep_vba = filepath.suffix.lower() == ".xlsm"
    wb = load_workbook(str(filepath), keep_vba=keep_vba)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        source = ws[source_ref]

        targets = _resolve_cells(ws, target_ref)
        for cell in targets:
            cell.font = copy(source.font)
            cell.fill = copy(source.fill)
            cell.border = copy(source.border)
            cell.alignment = copy(source.alignment)
            cell.number_format = source.number_format

        save_path = Path(output_path) if output_path else filepath
        wb.save(str(save_path))

        return {
            "status": "success",
            "cells_formatted": len(targets),
            "source": source_ref,
            "output_file": str(save_path),
        }
    finally:
        wb.close()


def _resolve_cells(ws: Any, cell_ref: str) -> list[Any]:
    """Resolve a cell reference (single or range) to a flat list of cell objects."""
    if ":" in cell_ref:
        cells = []
        for row in ws[cell_ref]:
            if isinstance(row, tuple):
                cells.extend(row)
            else:
                cells.append(row)
        return cells
    return [ws[cell_ref]]


# ---------------------------------------------------------------------------
# Sheet management
# ---------------------------------------------------------------------------


def manage_sheet(
    filepath: str | Path,
    action: str,
    sheet_name: str,
    new_name: str | None = None,
    output_path: str | Path | None = None,
) -> dict[str, Any]:
    """Create, delete, rename, copy, hide, or unhide a sheet.

    Actions: create, delete, rename, copy, hide, unhide
    """
    filepath = Path(filepath)
    keep_vba = filepath.suffix.lower() == ".xlsm"
    wb = load_workbook(str(filepath), keep_vba=keep_vba)
    try:
        result: dict[str, Any] = {"status": "success", "action": action}

        if action == "create":
            wb.create_sheet(title=sheet_name)
            result["sheet"] = sheet_name

        elif action == "delete":
            if sheet_name not in wb.sheetnames:
                from agent_xlsx.utils.errors import SheetNotFoundError

                raise SheetNotFoundError(sheet_name, wb.sheetnames)
            del wb[sheet_name]
            result["sheet"] = sheet_name

        elif action == "rename":
            if sheet_name not in wb.sheetnames:
                from agent_xlsx.utils.errors import SheetNotFoundError

                raise SheetNotFoundError(sheet_name, wb.sheetnames)
            ws = wb[sheet_name]
            ws.title = new_name or sheet_name
            result["old_name"] = sheet_name
            result["new_name"] = ws.title

        elif action == "copy":
            if sheet_name not in wb.sheetnames:
                from agent_xlsx.utils.errors import SheetNotFoundError

                raise SheetNotFoundError(sheet_name, wb.sheetnames)
            source = wb[sheet_name]
            copied = wb.copy_worksheet(source)
            if new_name:
                copied.title = new_name
            result["source"] = sheet_name
            result["copy"] = copied.title

        elif action == "hide":
            if sheet_name not in wb.sheetnames:
                from agent_xlsx.utils.errors import SheetNotFoundError

                raise SheetNotFoundError(sheet_name, wb.sheetnames)
            wb[sheet_name].sheet_state = "hidden"
            result["sheet"] = sheet_name

        elif action == "unhide":
            if sheet_name not in wb.sheetnames:
                from agent_xlsx.utils.errors import SheetNotFoundError

                raise SheetNotFoundError(sheet_name, wb.sheetnames)
            wb[sheet_name].sheet_state = "visible"
            result["sheet"] = sheet_name

        elif action == "list":
            sheets = []
            for ws in wb.worksheets:
                sheets.append(
                    {
                        "name": ws.title,
                        "state": ws.sheet_state,
                        "dimensions": ws.dimensions or "",
                    }
                )
            result["sheets"] = sheets
            result["count"] = len(sheets)
            # No save needed for list
            return result

        else:
            from agent_xlsx.utils.errors import AgentExcelError

            raise AgentExcelError(
                "INVALID_ACTION",
                f"Unknown sheet action: '{action}'",
                ["Valid actions: create, delete, rename, copy, hide, unhide, list"],
            )

        save_path = Path(output_path) if output_path else filepath
        wb.save(str(save_path))
        result["output_file"] = str(save_path)
        return result
    finally:
        wb.close()
