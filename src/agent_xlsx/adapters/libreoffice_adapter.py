"""LibreOffice adapter for screenshots and recalculation on headless/Linux systems.

Provides a fallback rendering pipeline when Microsoft Excel is not available.
Uses LibreOffice in headless mode to convert sheets to PDF, then PyMuPDF + Pillow
for high-fidelity PNG rendering with automatic content-bound cropping.
"""

from __future__ import annotations

import os
import shutil
import subprocess
import sys
import tempfile
import time
from pathlib import Path
from typing import Any

from PIL import Image, ImageChops

from agent_xlsx.utils.validation import index_to_col_letter

# ---------------------------------------------------------------------------
# LibreOffice discovery
# ---------------------------------------------------------------------------


def _find_soffice() -> str | None:
    """Locate the soffice binary on PATH or in standard install locations."""
    # Check PATH first
    for name in ("soffice", "libreoffice"):
        found = shutil.which(name)
        if found:
            return found

    # macOS standard install location
    if sys.platform == "darwin":
        mac_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
        if os.path.isfile(mac_path):
            return mac_path

    # Windows common locations
    if sys.platform == "win32":
        for base in (os.environ.get("PROGRAMFILES", ""), os.environ.get("PROGRAMFILES(X86)", "")):
            if base:
                candidate = os.path.join(base, "LibreOffice", "program", "soffice.exe")
                if os.path.isfile(candidate):
                    return candidate

    return None


def is_libreoffice_available() -> bool:
    """Check whether LibreOffice is installed and reachable."""
    return _find_soffice() is not None


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _headless_env() -> dict[str, str]:
    """Environment variables for headless LibreOffice."""
    env = os.environ.copy()
    env["SAL_USE_VCLPLUGIN"] = "gen"
    return env


def _unique_profile_dir() -> Path:
    """Create a temporary user profile to avoid LO single-instance locking."""
    return Path(tempfile.mkdtemp(prefix="agent_xlsx_lo_"))


def _autocrop_whitespace(img: Image.Image, padding: int = 10) -> Image.Image:
    """Crop whitespace from a rendered page image, keeping a small padding."""
    bg = Image.new(img.mode, img.size, (255, 255, 255))
    diff = ImageChops.difference(img, bg)
    bbox = diff.getbbox()
    if bbox is None:
        return img  # Entirely white — return as-is
    left = max(0, bbox[0] - padding)
    top = max(0, bbox[1] - padding)
    right = min(img.width, bbox[2] + padding)
    bottom = min(img.height, bbox[3] + padding)
    return img.crop((left, top, right, bottom))


def _prepare_sheet_for_export(
    filepath: str | Path,
    sheet_name: str,
    range_str: str | None,
    dest_path: Path,
) -> str:
    """Prepare an xlsx with only the target sheet, optimised for PDF export.

    Uses openpyxl to set print area, zero margins, landscape orientation,
    fitToWidth=1, fitToHeight=1, and removes all other sheets.

    Returns the resolved print-area range string (e.g. ``"A1:AT156"``).
    """
    from openpyxl import load_workbook
    from openpyxl.worksheet.page import PageMargins, PrintPageSetup

    wb = load_workbook(str(filepath))

    # Delete all sheets except the target
    for name in wb.sheetnames:
        if name != sheet_name:
            del wb[name]

    ws = wb[sheet_name]

    # Set print area to the exact data range so LO doesn't render
    # formatted-but-empty cells beyond it.  Safe with fitToHeight=1
    # (everything fits on one page regardless).
    if range_str:
        ws.print_area = range_str
        resolved = range_str
    else:
        # Scan for actual data extent (max_row/max_column include formatted-but-empty cells)
        max_data_row = 0
        max_data_col = 0
        for cell in ws._cells.values():
            if cell.value is not None:
                if cell.row > max_data_row:
                    max_data_row = cell.row
                if cell.column > max_data_col:
                    max_data_col = cell.column
        max_data_row = max_data_row or 1
        max_data_col = max_data_col or 1
        col_letter = index_to_col_letter(max_data_col - 1)
        resolved = f"A1:{col_letter}{max_data_row}"
        ws.print_area = resolved

    # Optimise for PDF export
    ws.page_margins = PageMargins(left=0.1, right=0.1, top=0.1, bottom=0.1, header=0, footer=0)
    ws.page_setup = PrintPageSetup(
        orientation="landscape",
        fitToWidth=1,
        fitToHeight=1,
    )
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Suppress headers/footers (page numbers prevent autocrop from trimming whitespace)
    for attr in (
        "oddHeader",
        "evenHeader",
        "firstHeader",
        "oddFooter",
        "evenFooter",
        "firstFooter",
    ):
        hf = getattr(ws, attr, None)
        if hf:
            hf.left.text = ""
            hf.center.text = ""
            hf.right.text = ""

    wb.save(str(dest_path))
    wb.close()

    return resolved


def _libreoffice_convert(
    soffice: str,
    input_path: Path,
    output_dir: Path,
    fmt: str = "pdf",
    filter_name: str | None = None,
    timeout: int = 30,
) -> Path:
    """Run LibreOffice headless conversion and return the output file path."""
    profile_dir = _unique_profile_dir()
    try:
        convert_arg = fmt if not filter_name else f"{fmt}:{filter_name}"
        cmd = [
            soffice,
            "--headless",
            "--norestore",
            "--nolockcheck",
            f"-env:UserInstallation=file://{profile_dir}",
            "--convert-to",
            convert_arg,
            "--outdir",
            str(output_dir),
            str(input_path),
        ]

        subprocess.run(
            cmd,
            env=_headless_env(),
            timeout=timeout,
            capture_output=True,
            check=True,
        )

        # Find the output file
        expected_ext = fmt.split(":")[0] if ":" in fmt else fmt
        output_file = output_dir / f"{input_path.stem}.{expected_ext}"
        if not output_file.exists():
            # Sometimes LO changes the extension slightly
            candidates = list(output_dir.glob(f"{input_path.stem}.*"))
            if candidates:
                output_file = candidates[0]
            else:
                msg = f"LibreOffice conversion produced no output for {input_path.name}"
                raise RuntimeError(msg)

        return output_file

    finally:
        shutil.rmtree(profile_dir, ignore_errors=True)


# ---------------------------------------------------------------------------
# Public API: screenshot
# ---------------------------------------------------------------------------


def screenshot(
    filepath: str | Path,
    sheet_name: str | None = None,
    range_str: str | None = None,
    output_path: str | Path | None = None,
    dpi: int = 200,
    timeout: int = 30,
) -> dict[str, Any]:
    """Export workbook sheet(s) to PNG via LibreOffice.

    Pipeline: openpyxl pre-process -> LO headless -> PDF -> PyMuPDF render -> Pillow crop -> PNG
    """
    import fitz  # PyMuPDF

    from agent_xlsx.utils.errors import LibreOfficeNotFoundError

    soffice = _find_soffice()
    if soffice is None:
        raise LibreOfficeNotFoundError()

    filepath = Path(filepath).resolve()
    start = time.perf_counter()

    if output_path is None:
        output_dir = Path("/tmp/agent-xlsx")
    else:
        output_dir = Path(output_path)
        if output_dir.suffix:
            output_dir = output_dir.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    stem = filepath.stem

    # Determine target sheets
    from openpyxl import load_workbook

    wb_info = load_workbook(str(filepath), read_only=True)
    if sheet_name:
        target_sheets = [s.strip() for s in sheet_name.split(",")]
    else:
        target_sheets = wb_info.sheetnames
    wb_info.close()

    # -------------------------------------------------------------------
    # PNG path — per-sheet rendering via PDF intermediate
    # -------------------------------------------------------------------
    sheets_result: list[dict[str, Any]] = []

    with tempfile.TemporaryDirectory(prefix="agent_xlsx_lo_") as tmpdir:
        tmp_path = Path(tmpdir)

        for target in target_sheets:
            # Prepare single-sheet workbook
            prepared = tmp_path / f"{stem}_{target}.xlsx"
            resolved_range = _prepare_sheet_for_export(filepath, target, range_str, prepared)

            # Convert to PDF
            pdf_file = _libreoffice_convert(
                soffice,
                prepared,
                tmp_path,
                fmt="pdf",
                timeout=timeout,
            )

            # Render all PDF pages via PyMuPDF and stitch vertically
            doc = fitz.open(str(pdf_file))
            page_images: list[Image.Image] = []
            for page in doc:
                pix = page.get_pixmap(dpi=dpi)
                page_images.append(Image.frombytes("RGB", (pix.width, pix.height), pix.samples))
            doc.close()

            if len(page_images) == 1:
                combined = page_images[0]
            else:
                total_width = max(im.width for im in page_images)
                total_height = sum(im.height for im in page_images)
                combined = Image.new("RGB", (total_width, total_height), (255, 255, 255))
                y_offset = 0
                for im in page_images:
                    combined.paste(im, (0, y_offset))
                    y_offset += im.height

            cropped = _autocrop_whitespace(combined)

            safe_name = target.replace("/", "_").replace("\\", "_").replace(" ", "_")
            if range_str:
                safe_range = range_str.replace(":", "-").replace("$", "")
                png_path = output_dir / f"{stem}_{safe_name}_{safe_range}.png"
            else:
                png_path = output_dir / f"{stem}_{safe_name}.png"
            cropped.save(str(png_path))

            sheets_result.append(
                {
                    "name": target,
                    "path": str(png_path),
                    "size_bytes": png_path.stat().st_size,
                    "width": cropped.width,
                    "height": cropped.height,
                    "resolved_range": resolved_range,
                }
            )

    elapsed_ms = round((time.perf_counter() - start) * 1000, 1)

    if len(sheets_result) == 1:
        result: dict[str, Any] = {
            "status": "success",
            "format": "png",
            "path": sheets_result[0]["path"],
            "sheet": sheets_result[0]["name"],
            "size_bytes": sheets_result[0]["size_bytes"],
            "width": sheets_result[0]["width"],
            "height": sheets_result[0]["height"],
            "resolved_range": sheets_result[0]["resolved_range"],
            "dpi": dpi,
            "capture_time_ms": elapsed_ms,
            "engine": "libreoffice",
        }
        if range_str:
            result["range"] = range_str
        return result
    else:
        return {
            "status": "success",
            "format": "png",
            "dpi": dpi,
            "sheets": sheets_result,
            "capture_time_ms": elapsed_ms,
            "engine": "libreoffice",
        }


# ---------------------------------------------------------------------------
# Public API: recalculate
# ---------------------------------------------------------------------------


def recalculate(
    filepath: str | Path,
    output_path: str | Path | None = None,
    timeout: int = 60,
) -> dict[str, Any]:
    """Recalculate all formulas by opening/saving through LibreOffice.

    LO recalculates all formulas on open, then we export back to xlsx.
    """
    from agent_xlsx.utils.errors import LibreOfficeNotFoundError

    soffice = _find_soffice()
    if soffice is None:
        raise LibreOfficeNotFoundError()

    filepath = Path(filepath).resolve()
    start = time.perf_counter()

    with tempfile.TemporaryDirectory(prefix="agent_xlsx_lo_") as tmpdir:
        tmp_path = Path(tmpdir)

        # Convert xlsx -> xlsx through LO (triggers recalc on open)
        output_file = _libreoffice_convert(
            soffice,
            filepath,
            tmp_path,
            fmt="xlsx",
            filter_name="Calc MS Excel 2007 XML",
            timeout=timeout,
        )

        target = Path(output_path) if output_path else filepath
        shutil.copy2(str(output_file), str(target))

    elapsed_ms = round((time.perf_counter() - start) * 1000, 1)
    return {
        "status": "success",
        "engine": "libreoffice",
        "output_file": str(target),
        "recalc_time_ms": elapsed_ms,
    }
