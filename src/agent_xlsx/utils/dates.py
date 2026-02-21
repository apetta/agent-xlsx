"""Centralised date utilities for detecting and converting Excel date columns."""

from __future__ import annotations

import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

# Regex to detect date-like number formats in Excel (contains year or day tokens)
_DATE_FORMAT_RE = re.compile(r"[yYdD]")


def detect_date_columns(filepath: str | Path, sheet_name: str | None = None) -> dict[str, bool]:
    """Detect columns with date number formats by inspecting openpyxl cell formats.

    Opens the workbook in read-only mode, checks the number_format of
    the first data row (row 2) for each column, and returns a dict of
    ``{header_name: True}`` for columns whose format contains date tokens.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(filepath), read_only=True)
    try:
        if sheet_name is not None:
            if sheet_name not in wb.sheetnames:
                return {}
            ws = wb[sheet_name]
        else:
            ws = wb.active
            if ws is None:
                return {}

        date_cols: dict[str, bool] = {}

        # Read rows 1 (header) and 2 (first data row)
        rows_iter = ws.iter_rows(min_row=1, max_row=2)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return {}

        try:
            data_row = next(rows_iter)
        except StopIteration:
            return {}

        for header_cell, data_cell in zip(header_row, data_row):
            header_name = header_cell.value
            if header_name is None:
                continue
            header_name = str(header_name)

            fmt = data_cell.number_format or "General"
            if _DATE_FORMAT_RE.search(fmt):
                date_cols[header_name] = True

        return date_cols
    finally:
        wb.close()


def excel_serial_to_isodate(serial: float) -> str | None:
    """Convert an Excel serial number to an ISO date string.

    Returns date-only (``"2024-02-15"``) for whole numbers, or
    datetime (``"2024-02-15T14:30:00"``) when a fractional time
    component is present.  Returns ``None`` for NaN values and the
    serial as-is for non-positive values.
    """
    if serial != serial:  # NaN check
        return None
    if serial <= 0:
        return serial  # type: ignore[return-value]

    base = datetime(1899, 12, 30)
    int_part = int(serial)
    frac_part = serial - int_part

    dt = base + timedelta(days=int_part)

    if frac_part > 1e-9:
        # Has a time component
        dt = dt + timedelta(days=frac_part)
        return dt.strftime("%Y-%m-%dT%H:%M:%S")

    return dt.strftime("%Y-%m-%d")


def detect_date_column_indices(
    filepath: str | Path, sheet_name: str | None = None
) -> set[int]:
    """Detect columns with date number formats. Returns 0-based column indices.

    Unlike :func:`detect_date_columns` which returns header names, this
    returns indices — making it safe for ``--no-header`` mode where
    DataFrame columns are letters (A, B, C) rather than header values.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(filepath), read_only=True)
    try:
        if sheet_name is not None:
            if sheet_name not in wb.sheetnames:
                return set()
            ws = wb[sheet_name]
        else:
            ws = wb.active
            if ws is None:
                return set()

        rows_iter = ws.iter_rows(min_row=1, max_row=2)
        try:
            next(rows_iter)  # row 1 (header or first data row)
        except StopIteration:
            return set()
        try:
            data_row = next(rows_iter)  # row 2 — carries number formats
        except StopIteration:
            return set()

        return {
            i
            for i, cell in enumerate(data_row)
            if _DATE_FORMAT_RE.search(cell.number_format or "General")
        }
    finally:
        wb.close()


def detect_date_column_indices_batch(
    filepath: str | Path, sheet_names: list[str]
) -> dict[str, set[int]]:
    """Detect date column indices for multiple sheets in a single workbook open.

    Returns ``{sheet_name: set[int]}`` mapping each sheet to its 0-based
    date column indices.  More efficient than calling
    :func:`detect_date_column_indices` per sheet.
    """
    import openpyxl

    result: dict[str, set[int]] = {}
    wb = openpyxl.load_workbook(str(filepath), read_only=True)
    try:
        for name in sheet_names:
            if name not in wb.sheetnames:
                result[name] = set()
                continue
            ws = wb[name]
            rows_iter = ws.iter_rows(min_row=1, max_row=2)
            try:
                next(rows_iter)  # row 1
            except StopIteration:
                result[name] = set()
                continue
            try:
                data_row = next(rows_iter)  # row 2
            except StopIteration:
                result[name] = set()
                continue
            result[name] = {
                i
                for i, cell in enumerate(data_row)
                if _DATE_FORMAT_RE.search(cell.number_format or "General")
            }
        return result
    finally:
        wb.close()


def convert_date_values(
    rows: list[list[Any]],
    headers: list[str],
    date_columns: set[str],
) -> list[list[Any]]:
    """Convert Excel serial numbers to ISO dates in-place for nominated columns.

    For each row, values at header indices that are in ``date_columns``
    are converted via :func:`excel_serial_to_isodate` when they are floats.
    Strings and ``None`` values pass through unchanged.
    """
    # Pre-compute column indices for date columns
    date_indices = [i for i, h in enumerate(headers) if h in date_columns]
    if not date_indices:
        return rows

    for row in rows:
        for idx in date_indices:
            if idx >= len(row):
                continue
            val = row[idx]
            if isinstance(val, (int, float)) and val == val:  # not NaN
                row[idx] = excel_serial_to_isodate(float(val))
    return rows
