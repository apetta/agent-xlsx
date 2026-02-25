"""Tests for the format command — multi-range support."""

import json

import pytest
from openpyxl import Workbook
from typer.testing import CliRunner

from agent_xlsx.cli import app

runner = CliRunner()


@pytest.fixture()
def styled_xlsx(tmp_path):
    """Workbook with data for formatting tests."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Header A"
    ws["B1"] = "Header B"
    ws["C1"] = "Header C"
    ws["A2"] = 1
    ws["B2"] = 2
    ws["C2"] = 3
    ws["A3"] = 4
    ws["B3"] = 5
    ws["C3"] = 6
    ws["A4"] = ""
    ws["B4"] = "Summary"
    ws["C4"] = ""
    ws["A5"] = 10
    ws["B5"] = 20
    ws["C5"] = 30
    p = tmp_path / "styled.xlsx"
    wb.save(p)
    return p


# ---------------------------------------------------------------------------
# Multi-range apply
# ---------------------------------------------------------------------------


def test_format_multi_range_apply_bold(styled_xlsx):
    """Multi-range applies same formatting to all ranges."""
    result = runner.invoke(
        app,
        ["format", str(styled_xlsx), "A1:C1,B4", "--font", '{"bold": true}'],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert data["ranges_formatted"] == 2
    assert data["total_cells_formatted"] == 4  # 3 cells in A1:C1 + 1 in B4

    # Verify formatting was applied to both ranges
    r1 = runner.invoke(app, ["format", str(styled_xlsx), "A1", "--read"])
    assert r1.exit_code == 0
    assert json.loads(r1.stdout)["font"]["bold"] is True

    r2 = runner.invoke(app, ["format", str(styled_xlsx), "B4", "--read"])
    assert r2.exit_code == 0
    assert json.loads(r2.stdout)["font"]["bold"] is True


def test_format_multi_range_apply_number_format(styled_xlsx):
    """Multi-range number format applies to all ranges."""
    result = runner.invoke(
        app,
        ["format", str(styled_xlsx), "A2:C2,A5:C5", "--number-format", "#,##0.00"],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["ranges_formatted"] == 2
    assert data["total_cells_formatted"] == 6


# ---------------------------------------------------------------------------
# Multi-range read
# ---------------------------------------------------------------------------


def test_format_multi_range_read(styled_xlsx):
    """Multi-range read returns formatting per range."""
    # Apply bold to A1 first
    runner.invoke(
        app,
        ["format", str(styled_xlsx), "A1", "--font", '{"bold": true}'],
    )

    result = runner.invoke(
        app,
        ["format", str(styled_xlsx), "A1,B4", "--read"],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["total_ranges"] == 2
    assert len(data["results"]) == 2
    # A1 should be bold, B4 should not
    assert data["results"][0]["formatting"]["font"]["bold"] is True
    assert data["results"][1]["formatting"]["font"]["bold"] is not True


# ---------------------------------------------------------------------------
# Multi-range copy
# ---------------------------------------------------------------------------


def test_format_multi_range_copy(styled_xlsx):
    """Multi-range copy applies source formatting to all target ranges."""
    # Make A1 bold
    runner.invoke(
        app,
        ["format", str(styled_xlsx), "A1", "--font", '{"bold": true}'],
    )

    # Copy A1's formatting to B4 and C5
    result = runner.invoke(
        app,
        ["format", str(styled_xlsx), "B4,C5", "--copy-from", "A1"],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert data["ranges_formatted"] == 2

    # Verify both cells got bold
    for cell_ref in ("B4", "C5"):
        r = runner.invoke(app, ["format", str(styled_xlsx), cell_ref, "--read"])
        assert json.loads(r.stdout)["font"]["bold"] is True


# ---------------------------------------------------------------------------
# Multi-range with sheet prefix
# ---------------------------------------------------------------------------


def test_format_multi_range_with_sheet_prefix(styled_xlsx):
    """Sheet prefix in first range carries forward to subsequent ranges."""
    result = runner.invoke(
        app,
        ["format", str(styled_xlsx), "Sheet1!A1:C1,B4", "--font", '{"bold": true}'],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["ranges_formatted"] == 2


# ---------------------------------------------------------------------------
# Single-range backward compatibility
# ---------------------------------------------------------------------------


def test_format_single_range_unchanged(styled_xlsx):
    """Single-range formatting still works (backward compat)."""
    result = runner.invoke(
        app,
        ["format", str(styled_xlsx), "A1:C1", "--font", '{"bold": true}'],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert data["cells_formatted"] == 3


# ---------------------------------------------------------------------------
# Issue #2 — --number alias for --number-format
# ---------------------------------------------------------------------------


def test_format_number_alias(styled_xlsx):
    """--number works as alias for --number-format."""
    result = runner.invoke(
        app,
        ["format", str(styled_xlsx), "A2:C2", "--number", "0.00"],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert data["cells_formatted"] == 3


# ---------------------------------------------------------------------------
# Phase 3 — Format shorthand flags
# ---------------------------------------------------------------------------


def test_format_bold_shorthand(styled_xlsx):
    """--bold shorthand sets font bold without JSON."""
    result = runner.invoke(
        app,
        ["format", str(styled_xlsx), "A1", "--bold"],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"

    # Verify bold was applied
    r = runner.invoke(app, ["format", str(styled_xlsx), "A1", "--read"])
    assert r.exit_code == 0
    assert json.loads(r.stdout)["font"]["bold"] is True


def test_format_italic_shorthand(styled_xlsx):
    """--italic shorthand sets font italic without JSON."""
    result = runner.invoke(
        app,
        ["format", str(styled_xlsx), "A1", "--italic"],
    )
    assert result.exit_code == 0, result.stdout

    r = runner.invoke(app, ["format", str(styled_xlsx), "A1", "--read"])
    assert r.exit_code == 0
    assert json.loads(r.stdout)["font"]["italic"] is True


def test_format_fill_color_shorthand(styled_xlsx):
    """--fill-color shorthand sets solid fill without JSON."""
    result = runner.invoke(
        app,
        ["format", str(styled_xlsx), "A1", "--fill-color", "FFFF00"],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"

    # Verify fill was applied
    r = runner.invoke(app, ["format", str(styled_xlsx), "A1", "--read"])
    assert r.exit_code == 0
    fill = json.loads(r.stdout)["fill"]
    assert fill["type"] == "solid"
    # Color should contain FFFF00 (may have alpha prefix)
    assert "FFFF00" in fill["color"].upper()


def test_format_shorthand_combines_with_json(styled_xlsx):
    """Shorthand flags layer on top of --font JSON."""
    result = runner.invoke(
        app,
        ["format", str(styled_xlsx), "A1", "--font", '{"size": 14}', "--bold"],
    )
    assert result.exit_code == 0, result.stdout

    r = runner.invoke(app, ["format", str(styled_xlsx), "A1", "--read"])
    assert r.exit_code == 0
    font = json.loads(r.stdout)["font"]
    assert font["bold"] is True
    assert font["size"] == 14


def test_format_no_bold_shorthand(styled_xlsx):
    """--no-bold explicitly disables bold."""
    # First make it bold
    runner.invoke(app, ["format", str(styled_xlsx), "A1", "--bold"])
    # Then un-bold it
    result = runner.invoke(app, ["format", str(styled_xlsx), "A1", "--no-bold"])
    assert result.exit_code == 0, result.stdout

    r = runner.invoke(app, ["format", str(styled_xlsx), "A1", "--read"])
    assert r.exit_code == 0
    assert json.loads(r.stdout)["font"]["bold"] is not True


# ---------------------------------------------------------------------------
# Regression: multi-range --output preserves all ranges
# ---------------------------------------------------------------------------


def test_format_multi_range_output_preserves_all(styled_xlsx, tmp_path):
    """Multi-range format with --output preserves formatting from all ranges.

    Regression: multi-range format with --output always loaded from the original
    source file, so each iteration's save overwrote the previous. Only the last
    range's formatting survived. Fixed with working_path variable that switches
    to the output file after the first iteration.
    """
    from openpyxl import load_workbook

    out = tmp_path / "formatted.xlsx"
    result = runner.invoke(
        app,
        [
            "format",
            str(styled_xlsx),
            "A1:A1,B1:B1",
            "--bold",
            "--output",
            str(out),
        ],
    )
    assert result.exit_code == 0, result.stdout

    # Verify BOTH cells are bold in the output file
    wb = load_workbook(str(out))
    ws = wb.active
    assert ws["A1"].font.bold is True, "A1 should be bold (first range)"
    assert ws["B1"].font.bold is True, "B1 should be bold (second range)"
    wb.close()


def test_format_multi_range_output_non_writable_extension(styled_xlsx, tmp_path):
    """Multi-range format with --output using a non-writable extension (.xls).

    Regression: the adapter auto-converts non-writable extensions to .xlsx
    (e.g. out.xls → out.xlsx). The working_path must use the adapter's actual
    save path, not the raw user-provided output string, otherwise iteration 2
    tries to load a file that doesn't exist.
    """
    from openpyxl import load_workbook

    # Use .xls extension — adapter will auto-convert to .xlsx
    out = tmp_path / "formatted.xls"
    result = runner.invoke(
        app,
        [
            "format",
            str(styled_xlsx),
            "A1:A1,B1:B1",
            "--bold",
            "--output",
            str(out),
        ],
    )
    assert result.exit_code == 0, result.stdout

    # Adapter converts .xls → .xlsx; verify the actual file exists and has both ranges
    actual_out = tmp_path / "formatted.xlsx"
    assert actual_out.exists(), "Adapter should auto-convert .xls to .xlsx"
    wb = load_workbook(str(actual_out))
    ws = wb.active
    assert ws["A1"].font.bold is True, "A1 should be bold (first range)"
    assert ws["B1"].font.bold is True, "B1 should be bold (second range)"
    wb.close()


# ---------------------------------------------------------------------------
# Batch formatting (--batch / --batch-file)
# ---------------------------------------------------------------------------


def test_format_batch_basic(styled_xlsx):
    """--batch applies multiple format groups in one call."""
    batch_spec = '[{"range": "A1:C1", "bold": true}]'
    result = runner.invoke(
        app,
        ["format", str(styled_xlsx), "A1", "--batch", batch_spec],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert data["groups_applied"] == 1

    # Verify A1 is bold via --read
    r = runner.invoke(app, ["format", str(styled_xlsx), "A1", "--read"])
    assert r.exit_code == 0
    assert json.loads(r.stdout)["font"]["bold"] is True


def test_format_batch_file(styled_xlsx, tmp_path):
    """--batch-file reads format spec from a JSON file and applies all styles."""
    batch_spec = [{"range": "A1:C1", "bold": True, "fill_color": "FFFF00"}]
    batch_file = tmp_path / "batch.json"
    batch_file.write_text(json.dumps(batch_spec))

    result = runner.invoke(
        app,
        ["format", str(styled_xlsx), "A1", "--batch-file", str(batch_file)],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert data["groups_applied"] == 1

    # Readback: verify bold and fill were actually applied
    r = runner.invoke(app, ["format", str(styled_xlsx), "A1", "--read"])
    assert r.exit_code == 0
    fmt = json.loads(r.stdout)
    assert fmt["font"]["bold"] is True
    assert fmt["fill"]["type"] == "solid"
    assert "FFFF00" in fmt["fill"]["color"].upper()


def test_format_batch_comma_ranges(styled_xlsx):
    """--batch with comma-separated ranges in a single entry formats both ranges."""
    batch_spec = '[{"range": "A1:C1,A3:C3", "bold": true}]'
    result = runner.invoke(
        app,
        ["format", str(styled_xlsx), "A1", "--batch", batch_spec],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert data["total_cells_formatted"] == 6  # A1:C1 (3) + A3:C3 (3)

    # A1 and A3 should be bold, A2 should NOT
    from openpyxl import load_workbook

    wb = load_workbook(str(styled_xlsx))
    ws = wb.active
    assert ws["A1"].font.bold is True, "A1 should be bold (first range)"
    assert ws["A3"].font.bold is True, "A3 should be bold (second range)"
    assert ws["A2"].font.bold is not True, "A2 should NOT be bold (between ranges)"
    wb.close()


# ---------------------------------------------------------------------------
# P2 — Relative output path
# ---------------------------------------------------------------------------


def test_format_output_file_is_relative(styled_xlsx):
    """output_file in response must be present, relative, and match the source filename."""
    from pathlib import Path

    result = runner.invoke(
        app,
        ["format", str(styled_xlsx), "A1", "--bold"],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert "output_file" in data, "output_file must be present in format response"
    assert not data["output_file"].startswith("/"), (
        f"output_file should be relative, got: {data['output_file']}"
    )
    assert Path(data["output_file"]).name == styled_xlsx.name
