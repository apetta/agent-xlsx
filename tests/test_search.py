"""Tests for search command: --columns, --limit, --range, and composability."""

import json

import pytest
from openpyxl import Workbook
from typer.testing import CliRunner

from agent_xlsx.cli import app

runner = CliRunner()


@pytest.fixture
def wide_xlsx(tmp_path):
    """Workbook with 5 columns and 20 rows of known data for search tests."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    # Headers
    ws["A1"] = "Name"
    ws["B1"] = "Code"
    ws["C1"] = "Amount"
    ws["D1"] = "Description"
    ws["E1"] = "Category"
    # Data rows
    for i in range(2, 22):
        ws[f"A{i}"] = f"Item-{i - 1}"
        ws[f"B{i}"] = f"CODE-{i - 1:03d}"
        ws[f"C{i}"] = (i - 1) * 100
        ws[f"D{i}"] = f"This is a long description for item {i - 1} with lots of text"
        ws[f"E{i}"] = "TypeA" if i % 2 == 0 else "TypeB"
    # Add some searchable duplicates
    ws["A10"] = "TARGET"
    ws["B10"] = "TARGET"
    ws["D10"] = "TARGET"
    p = tmp_path / "wide.xlsx"
    wb.save(p)
    return p


@pytest.fixture
def multisheet_xlsx(tmp_path):
    """Workbook with two sheets for cross-sheet search tests."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "header"
    ws1["A2"] = "findme"
    ws1["B1"] = "other"
    ws1["B2"] = "data"

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "header"
    ws2["A2"] = "findme"
    ws2["B1"] = "other"
    ws2["B2"] = "unique_to_sheet2"

    p = tmp_path / "multi.xlsx"
    wb.save(p)
    return p


@pytest.fixture
def formula_xlsx(tmp_path):
    """Workbook with formulas for --in-formulas search tests."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Value"
    ws["B1"] = "Formula"
    ws["A2"] = 10
    ws["B2"] = "=SUM(A2:A10)"
    ws["A3"] = 20
    ws["B3"] = "=AVERAGE(A2:A10)"
    ws["C1"] = "Other"
    ws["C2"] = "=A2*2"
    p = tmp_path / "formulas.xlsx"
    wb.save(p)
    return p


# ---------------------------------------------------------------------------
# P1 #3 — --columns filter
# ---------------------------------------------------------------------------


def test_search_columns_by_letter(wide_xlsx):
    """--columns A only searches column A."""
    result = runner.invoke(app, ["search", str(wide_xlsx), "TARGET", "--columns", "A"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["match_count"] == 1
    assert data["matches"][0]["column"] == "A"


def test_search_columns_by_header_name(wide_xlsx):
    """--columns accepts header names."""
    result = runner.invoke(app, ["search", str(wide_xlsx), "TARGET", "--columns", "Name"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["match_count"] == 1
    assert data["matches"][0]["column"] == "A"


def test_search_columns_multiple(wide_xlsx):
    """--columns A,B searches both columns."""
    result = runner.invoke(app, ["search", str(wide_xlsx), "TARGET", "--columns", "A,B"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["match_count"] == 2
    cols = {m["column"] for m in data["matches"]}
    assert cols == {"A", "B"}


def test_search_columns_excludes_noisy_columns(wide_xlsx):
    """--columns A excludes matches from D (Description) which also contains TARGET."""
    result = runner.invoke(app, ["search", str(wide_xlsx), "TARGET", "--columns", "A"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    # Only A column should match, not D
    assert all(m["column"] == "A" for m in data["matches"])


def test_search_columns_invalid_raises_error(wide_xlsx):
    """Invalid column references produce INVALID_COLUMN error."""
    result = runner.invoke(app, ["search", str(wide_xlsx), "test", "--columns", "ZZZ"])
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "INVALID_COLUMN"


def test_search_columns_with_in_formulas(formula_xlsx):
    """--columns works with --in-formulas path."""
    result = runner.invoke(
        app,
        ["search", str(formula_xlsx), "SUM", "--in-formulas", "--columns", "B"],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["match_count"] == 1
    assert "SUM" in data["matches"][0]["formula"]


# ---------------------------------------------------------------------------
# P1 #4 — --limit flag
# ---------------------------------------------------------------------------


def test_search_limit_below_default(wide_xlsx):
    """--limit 1 returns at most 1 result."""
    result = runner.invoke(app, ["search", str(wide_xlsx), "Item", "--limit", "1"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["match_count"] <= 1


def test_search_limit_truncated_flag(wide_xlsx):
    """truncated=True when results hit the effective limit."""
    result = runner.invoke(app, ["search", str(wide_xlsx), "Item", "--limit", "1"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["truncated"] is True


def test_search_limit_not_truncated(wide_xlsx):
    """truncated=False when fewer results than limit."""
    result = runner.invoke(
        app, ["search", str(wide_xlsx), "TARGET", "--columns", "A", "--limit", "100"]
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["truncated"] is False


def test_search_limit_with_in_formulas(formula_xlsx):
    """--limit works with --in-formulas path."""
    result = runner.invoke(
        app, ["search", str(formula_xlsx), "A2", "--in-formulas", "--limit", "1"]
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["match_count"] <= 1


# ---------------------------------------------------------------------------
# P1 #5 — --range scoping
# ---------------------------------------------------------------------------


def test_search_range_restricts_rows(wide_xlsx):
    """--range restricts search to specified rows."""
    # TARGET is in row 10. Range A2:A5 should not find it.
    result = runner.invoke(app, ["search", str(wide_xlsx), "TARGET", "--range", "A2:A5"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["match_count"] == 0

    # Range A2:A15 should find it
    result2 = runner.invoke(app, ["search", str(wide_xlsx), "TARGET", "--range", "A2:A15"])
    assert result2.exit_code == 0, result2.stdout
    data2 = json.loads(result2.stdout)
    assert data2["match_count"] >= 1


def test_search_range_with_sheet_prefix(multisheet_xlsx):
    """--range Sheet2!A1:B5 searches Sheet2."""
    result = runner.invoke(
        app,
        ["search", str(multisheet_xlsx), "unique_to_sheet2", "--range", "Sheet2!A1:B5"],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["match_count"] == 1
    assert data["matches"][0]["sheet"] == "Sheet2"


def test_search_range_overrides_sheet_flag(multisheet_xlsx):
    """--range Sheet2!A1:B5 overrides --sheet Sheet1."""
    result = runner.invoke(
        app,
        [
            "search",
            str(multisheet_xlsx),
            "unique_to_sheet2",
            "--range",
            "Sheet2!A1:B5",
            "--sheet",
            "Sheet1",
        ],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["match_count"] == 1
    assert data["matches"][0]["sheet"] == "Sheet2"


# ---------------------------------------------------------------------------
# Composability
# ---------------------------------------------------------------------------


def test_search_columns_and_limit(wide_xlsx):
    """--columns A --limit 2 combines both filters."""
    result = runner.invoke(
        app, ["search", str(wide_xlsx), "Item", "--columns", "A", "--limit", "2"]
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["match_count"] <= 2
    assert all(m["column"] == "A" for m in data["matches"])


def test_search_all_three(wide_xlsx):
    """--columns, --range, and --limit compose together."""
    result = runner.invoke(
        app,
        [
            "search",
            str(wide_xlsx),
            "Item",
            "--columns",
            "A",
            "--range",
            "A2:E10",
            "--limit",
            "3",
        ],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["match_count"] <= 3
    assert all(m["column"] == "A" for m in data["matches"])


def test_search_columns_by_header_name_with_range(wide_xlsx):
    """--columns by header name works when combined with --range."""
    # "Name" is header of column A. Search within A2:D15 using header name.
    result = runner.invoke(
        app,
        ["search", str(wide_xlsx), "TARGET", "--columns", "Name", "--range", "A2:D15"],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["match_count"] == 1
    assert data["matches"][0]["column"] == "A"


def test_search_columns_header_name_with_range_in_formulas(formula_xlsx):
    """--columns by header name + --range works with --in-formulas."""
    result = runner.invoke(
        app,
        [
            "search",
            str(formula_xlsx),
            "SUM",
            "--in-formulas",
            "--columns",
            "Formula",
            "--range",
            "B1:C3",
        ],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["match_count"] == 1


def test_search_columns_header_name_single_column_range(wide_xlsx):
    """--columns by header name resolves with a single-column range (mirrors WDIEXCEL scenario)."""
    # B6 = "CODE-005" (not overwritten by TARGET fixtures at row 10)
    result = runner.invoke(
        app,
        ["search", str(wide_xlsx), "CODE-005", "--columns", "Code", "--range", "B2:B15"],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["match_count"] == 1
    assert data["matches"][0]["column"] == "B"


# ---------------------------------------------------------------------------
# Backward compatibility
# ---------------------------------------------------------------------------


def test_search_no_new_flags_unchanged(sample_xlsx):
    """Default behavior (no new flags) matches existing output structure."""
    result = runner.invoke(app, ["search", str(sample_xlsx), "value1"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert "query" in data
    assert "match_count" in data
    assert "matches" in data
    assert "truncated" in data
    assert "search_time_ms" in data
    assert "file_size_human" in data


def test_search_file_size_human_present(sample_xlsx):
    """file_size_human is present in search output."""
    result = runner.invoke(app, ["search", str(sample_xlsx), "value1"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert "file_size_human" in data
    assert isinstance(data["file_size_human"], str)


# ---------------------------------------------------------------------------
# Basic search, --regex, --ignore-case, unicode
# ---------------------------------------------------------------------------


def test_search_basic(sample_xlsx):
    """Basic search returns matching rows."""
    result = runner.invoke(app, ["search", str(sample_xlsx), "value1"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["match_count"] >= 1
    assert any("value1" in str(m.get("value", "")) for m in data["matches"])


def test_search_regex(wide_xlsx):
    """--regex treats query as a regex pattern."""
    result = runner.invoke(app, ["search", str(wide_xlsx), "Item-1\\d", "--regex"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    # Should match Item-10 through Item-19 (and possibly Item-1 depending on anchoring)
    assert data["match_count"] >= 1
    for m in data["matches"]:
        assert "Item-1" in str(m.get("value", ""))


def test_search_regex_invalid(sample_xlsx):
    """--regex with an invalid pattern returns INVALID_REGEX error."""
    result = runner.invoke(app, ["search", str(sample_xlsx), "[invalid(", "--regex"])
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "INVALID_REGEX"


def test_search_ignore_case(wide_xlsx):
    """--ignore-case matches regardless of case."""
    result = runner.invoke(app, ["search", str(wide_xlsx), "target", "--ignore-case"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    # "TARGET" is in the fixture at row 10 across columns A, B, D
    assert data["match_count"] >= 1


def test_search_unicode(unicode_xlsx):
    """Search finds Unicode/CJK content."""
    result = runner.invoke(app, ["search", str(unicode_xlsx), "\u5c71\u7530\u592a\u90ce"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["match_count"] >= 1
