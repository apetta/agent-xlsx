"""Tests for sheet command: list, create, rename, delete, copy, hide/unhide."""

import json
import shutil

from openpyxl import load_workbook
from typer.testing import CliRunner

from agent_xlsx.cli import app

runner = CliRunner()


# ---------------------------------------------------------------------------
# --list
# ---------------------------------------------------------------------------


def test_sheet_list(multisheet_xlsx):
    """--list returns all sheet names, states, and dimensions."""
    result = runner.invoke(app, ["sheet", str(multisheet_xlsx), "--list"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert data["action"] == "list"
    assert data["count"] == 3
    names = [s["name"] for s in data["sheets"]]
    assert names == ["Alpha", "Beta", "Gamma"]
    # Each sheet entry should have state and dimensions
    for sheet in data["sheets"]:
        assert "state" in sheet
        assert "dimensions" in sheet
        assert sheet["state"] == "visible"


# ---------------------------------------------------------------------------
# --create
# ---------------------------------------------------------------------------


def test_sheet_create(multisheet_xlsx, tmp_path):
    """--create adds a new sheet and saves the workbook."""
    test_file = tmp_path / "test.xlsx"
    shutil.copy(multisheet_xlsx, test_file)

    result = runner.invoke(app, ["sheet", str(test_file), "--create", "NewSheet"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert data["action"] == "create"
    assert data["sheet"] == "NewSheet"

    # Verify the sheet was actually created
    wb = load_workbook(str(test_file))
    assert "NewSheet" in wb.sheetnames
    assert len(wb.sheetnames) == 4
    wb.close()


# ---------------------------------------------------------------------------
# --rename
# ---------------------------------------------------------------------------


def test_sheet_rename(multisheet_xlsx, tmp_path):
    """--rename with --new-name renames an existing sheet."""
    test_file = tmp_path / "test.xlsx"
    shutil.copy(multisheet_xlsx, test_file)

    result = runner.invoke(
        app, ["sheet", str(test_file), "--rename", "Alpha", "--new-name", "AlphaRenamed"]
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert data["action"] == "rename"
    assert data["old_name"] == "Alpha"
    assert data["new_name"] == "AlphaRenamed"

    # Verify in workbook
    wb = load_workbook(str(test_file))
    assert "AlphaRenamed" in wb.sheetnames
    assert "Alpha" not in wb.sheetnames
    wb.close()


def test_sheet_rename_missing_new_name(multisheet_xlsx):
    """--rename without --new-name produces MISSING_ARGUMENT error."""
    result = runner.invoke(app, ["sheet", str(multisheet_xlsx), "--rename", "Alpha"])
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "MISSING_ARGUMENT"


# ---------------------------------------------------------------------------
# --delete
# ---------------------------------------------------------------------------


def test_sheet_delete(multisheet_xlsx, tmp_path):
    """--delete removes the named sheet."""
    test_file = tmp_path / "test.xlsx"
    shutil.copy(multisheet_xlsx, test_file)

    result = runner.invoke(app, ["sheet", str(test_file), "--delete", "Gamma"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert data["action"] == "delete"
    assert data["sheet"] == "Gamma"

    # Verify the sheet was removed
    wb = load_workbook(str(test_file))
    assert "Gamma" not in wb.sheetnames
    assert len(wb.sheetnames) == 2
    wb.close()


def test_sheet_delete_not_found(multisheet_xlsx, tmp_path):
    """--delete with a non-existent sheet name produces SHEET_NOT_FOUND."""
    test_file = tmp_path / "test.xlsx"
    shutil.copy(multisheet_xlsx, test_file)

    result = runner.invoke(app, ["sheet", str(test_file), "--delete", "NoSuchSheet"])
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "SHEET_NOT_FOUND"
    assert "NoSuchSheet" in data["message"]


# ---------------------------------------------------------------------------
# --copy
# ---------------------------------------------------------------------------


def test_sheet_copy_with_new_name(multisheet_xlsx, tmp_path):
    """--copy with --new-name creates a copy under the specified name."""
    test_file = tmp_path / "test.xlsx"
    shutil.copy(multisheet_xlsx, test_file)

    result = runner.invoke(
        app, ["sheet", str(test_file), "--copy", "Alpha", "--new-name", "AlphaCopy"]
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert data["action"] == "copy"
    assert data["source"] == "Alpha"
    assert data["copy"] == "AlphaCopy"

    # Verify both sheets exist
    wb = load_workbook(str(test_file))
    assert "Alpha" in wb.sheetnames
    assert "AlphaCopy" in wb.sheetnames
    assert len(wb.sheetnames) == 4
    wb.close()


def test_sheet_copy_without_new_name(multisheet_xlsx, tmp_path):
    """--copy without --new-name uses openpyxl's default naming (e.g. 'Alpha Copy')."""
    test_file = tmp_path / "test.xlsx"
    shutil.copy(multisheet_xlsx, test_file)

    result = runner.invoke(app, ["sheet", str(test_file), "--copy", "Alpha"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert data["action"] == "copy"
    assert data["source"] == "Alpha"
    # openpyxl default copy name
    assert "copy" in data["copy"].lower() or "Alpha" in data["copy"]

    # Verify the original still exists
    wb = load_workbook(str(test_file))
    assert "Alpha" in wb.sheetnames
    assert len(wb.sheetnames) == 4
    wb.close()


def test_sheet_copy_not_found(multisheet_xlsx, tmp_path):
    """--copy with a non-existent sheet name produces SHEET_NOT_FOUND."""
    test_file = tmp_path / "test.xlsx"
    shutil.copy(multisheet_xlsx, test_file)

    result = runner.invoke(app, ["sheet", str(test_file), "--copy", "NoSuchSheet"])
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "SHEET_NOT_FOUND"


# ---------------------------------------------------------------------------
# --hide / --unhide
# ---------------------------------------------------------------------------


def test_sheet_hide(multisheet_xlsx, tmp_path):
    """--hide sets the sheet state to hidden."""
    test_file = tmp_path / "test.xlsx"
    shutil.copy(multisheet_xlsx, test_file)

    result = runner.invoke(app, ["sheet", str(test_file), "--hide", "Beta"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert data["action"] == "hide"
    assert data["sheet"] == "Beta"

    # Verify the sheet is hidden
    wb = load_workbook(str(test_file))
    assert wb["Beta"].sheet_state == "hidden"
    wb.close()


def test_sheet_unhide(multisheet_xlsx, tmp_path):
    """--unhide restores a hidden sheet to visible."""
    test_file = tmp_path / "test.xlsx"
    shutil.copy(multisheet_xlsx, test_file)

    # First hide the sheet
    wb = load_workbook(str(test_file))
    wb["Beta"].sheet_state = "hidden"
    wb.save(str(test_file))
    wb.close()

    # Then unhide via CLI
    result = runner.invoke(app, ["sheet", str(test_file), "--unhide", "Beta"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert data["action"] == "unhide"
    assert data["sheet"] == "Beta"

    # Verify the sheet is visible again
    wb = load_workbook(str(test_file))
    assert wb["Beta"].sheet_state == "visible"
    wb.close()


def test_sheet_hide_not_found(multisheet_xlsx, tmp_path):
    """--hide with a non-existent sheet name produces SHEET_NOT_FOUND."""
    test_file = tmp_path / "test.xlsx"
    shutil.copy(multisheet_xlsx, test_file)

    result = runner.invoke(app, ["sheet", str(test_file), "--hide", "NoSuchSheet"])
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "SHEET_NOT_FOUND"


# ---------------------------------------------------------------------------
# --output (save to new file)
# ---------------------------------------------------------------------------


def test_sheet_create_with_output(multisheet_xlsx, tmp_path):
    """--output saves the result to a new file, leaving the original unchanged."""
    test_file = tmp_path / "test.xlsx"
    shutil.copy(multisheet_xlsx, test_file)
    output_file = tmp_path / "output.xlsx"

    result = runner.invoke(
        app, ["sheet", str(test_file), "--create", "OutputSheet", "--output", str(output_file)]
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert output_file.exists()

    # Output file should have the new sheet
    wb_out = load_workbook(str(output_file))
    assert "OutputSheet" in wb_out.sheetnames
    wb_out.close()

    # Original should NOT have the new sheet
    wb_orig = load_workbook(str(test_file))
    assert "OutputSheet" not in wb_orig.sheetnames
    wb_orig.close()


# ---------------------------------------------------------------------------
# No action specified
# ---------------------------------------------------------------------------


def test_sheet_no_action(multisheet_xlsx):
    """Invoking sheet with no action flags produces MISSING_ACTION error."""
    result = runner.invoke(app, ["sheet", str(multisheet_xlsx)])
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "MISSING_ACTION"
    assert "suggestions" in data


# ---------------------------------------------------------------------------
# SHEET_NOT_FOUND for rename and unhide
# ---------------------------------------------------------------------------


def test_sheet_rename_not_found(multisheet_xlsx, tmp_path):
    """--rename with a non-existent sheet name produces SHEET_NOT_FOUND."""
    test_file = tmp_path / "test.xlsx"
    shutil.copy(multisheet_xlsx, test_file)

    result = runner.invoke(
        app, ["sheet", str(test_file), "--rename", "NoSuchSheet", "--new-name", "NewName"]
    )
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "SHEET_NOT_FOUND"


def test_sheet_unhide_not_found(multisheet_xlsx, tmp_path):
    """--unhide with a non-existent sheet name produces SHEET_NOT_FOUND."""
    test_file = tmp_path / "test.xlsx"
    shutil.copy(multisheet_xlsx, test_file)

    result = runner.invoke(app, ["sheet", str(test_file), "--unhide", "NoSuchSheet"])
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "SHEET_NOT_FOUND"
