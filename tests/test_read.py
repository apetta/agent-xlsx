"""Tests for read command: --headers flag and file_size_human output."""

import json

import pytest
from openpyxl import Workbook
from typer.testing import CliRunner

from agent_xlsx.cli import app

runner = CliRunner()


@pytest.fixture
def tabular_xlsx(tmp_path):
    """Workbook with headers in row 1 and data in rows 2+."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws["A1"] = "Product"
    ws["B1"] = "Revenue"
    ws["C1"] = "Region"
    for i in range(2, 12):
        ws[f"A{i}"] = f"Product-{i - 1}"
        ws[f"B{i}"] = (i - 1) * 1000
        ws[f"C{i}"] = "North" if i % 2 == 0 else "South"
    p = tmp_path / "tabular.xlsx"
    wb.save(p)
    return p


# ---------------------------------------------------------------------------
# P2 #8 — --headers flag
# ---------------------------------------------------------------------------


def test_read_headers_resolves_column_names(tabular_xlsx):
    """--headers resolves letter headers to row-1 names in range reads."""
    result = runner.invoke(app, ["read", str(tabular_xlsx), "A5:C5", "--headers"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    # Headers should be resolved from row 1, not letters
    assert "Product" in data["headers"]
    assert "Revenue" in data["headers"]
    assert "Region" in data["headers"]
    # column_map should map letters to names
    assert "column_map" in data
    assert data["column_map"]["A"] == "Product"


def test_read_headers_with_no_header_ignored(tabular_xlsx):
    """--headers + --no-header: --no-header takes precedence."""
    result = runner.invoke(app, ["read", str(tabular_xlsx), "A5:C5", "--headers", "--no-header"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    # Should still have letter headers since --no-header wins
    assert "column_map" not in data


def test_read_headers_without_range_is_noop(tabular_xlsx):
    """--headers without a range is a no-op (full reads already have headers)."""
    result = runner.invoke(app, ["read", str(tabular_xlsx), "--headers"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    # Full read: headers are already from row 1
    assert "Product" in data["headers"]
    # No column_map needed since headers were already resolved
    assert "column_map" not in data


def test_read_range_without_headers_uses_letters(tabular_xlsx):
    """Range reads without --headers use column letters (existing behavior)."""
    result = runner.invoke(app, ["read", str(tabular_xlsx), "A5:C5"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    # Should be letter headers: A, B, C
    assert data["headers"] == ["A", "B", "C"]
    assert "column_map" not in data


# ---------------------------------------------------------------------------
# P2 #7 — file_size_human in output
# ---------------------------------------------------------------------------


def test_read_file_size_human_present(tabular_xlsx):
    """file_size_human field is present in read output."""
    result = runner.invoke(app, ["read", str(tabular_xlsx)])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert "file_size_human" in data
    assert isinstance(data["file_size_human"], str)
    # Should be in KB range for a small test file
    assert "KB" in data["file_size_human"] or "B" in data["file_size_human"]


def test_probe_file_size_human_present(tabular_xlsx):
    """file_size_human field is present in probe output."""
    result = runner.invoke(app, ["probe", str(tabular_xlsx)])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert "file_size_human" in data
    assert "size_bytes" in data  # backward compat: size_bytes still present


# ---------------------------------------------------------------------------
# P2 #3 — --headers on multi-range reads
# ---------------------------------------------------------------------------


def test_read_headers_multi_range(tabular_xlsx):
    """--headers resolves names on multi-range reads."""
    result = runner.invoke(app, ["read", str(tabular_xlsx), "A5:C5,A8:C8", "--headers"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    # Multi-range returns results array
    for r in data["results"]:
        assert "Product" in r["headers"]
        assert "Revenue" in r["headers"]
        assert "column_map" in r
        assert r["column_map"]["A"] == "Product"


def test_read_headers_multi_range_wide_columns(tabular_xlsx):
    """--headers resolves names for non-A columns on multi-range reads."""
    result = runner.invoke(app, ["read", str(tabular_xlsx), "B5:C5,B8:C8", "--headers"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    for r in data["results"]:
        assert "Revenue" in r["headers"]
        assert "Region" in r["headers"]
        assert "column_map" in r
        assert r["column_map"]["B"] == "Revenue"


def test_read_headers_multi_range_no_header_wins(tabular_xlsx):
    """--headers + --no-header on multi-range: --no-header wins."""
    result = runner.invoke(
        app, ["read", str(tabular_xlsx), "A5:C5,A8:C8", "--headers", "--no-header"]
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    for r in data["results"]:
        assert "column_map" not in r


# ---------------------------------------------------------------------------
# Issue #8 — OOB column warning
# ---------------------------------------------------------------------------


def test_read_oob_columns_returns_warning(tabular_xlsx):
    """Reading beyond the sheet's data range returns a warning."""
    # tabular_xlsx has columns A-C (3 columns). Requesting A1:Z2 should warn.
    result = runner.invoke(app, ["read", str(tabular_xlsx), "A1:Z2"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert "warning" in data
    assert "omitted" in data["warning"].lower() or "column" in data["warning"].lower()


def test_read_within_bounds_no_warning(tabular_xlsx):
    """Reading within the sheet's data range produces no warning."""
    result = runner.invoke(app, ["read", str(tabular_xlsx), "A1:C2"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert "warning" not in data


# ---------------------------------------------------------------------------
# EmptyCell handling in --formulas mode
# ---------------------------------------------------------------------------


@pytest.fixture
def sparse_xlsx(tmp_path):
    """Workbook with data gaps that produce EmptyCell in read_only mode."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sparse"
    ws["A1"] = "Val"
    ws["B1"] = "Formula"
    ws["A2"] = 10
    ws["B2"] = "=SUM(A2:A3)"
    ws["A3"] = 20
    ws["B3"] = "=AVERAGE(A2:A3)"
    # Row 4 intentionally empty — creates EmptyCell objects in read_only mode
    ws["B5"] = "=A2*2"
    p = tmp_path / "sparse.xlsx"
    wb.save(p)
    return p


def test_read_formulas_with_empty_cells(sparse_xlsx):
    """--formulas handles EmptyCell objects without crashing."""
    result = runner.invoke(app, ["read", str(sparse_xlsx), "--formulas"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["backend"] == "openpyxl"
    # Should have cells from non-empty positions only (compact mode strips empties)
    assert data["cell_count"] > 0
    # Verify formulas are captured
    formulas = [c for c in data["cells"] if c.get("formula")]
    assert len(formulas) >= 2  # B2, B3, B5 have formulas


# ---------------------------------------------------------------------------
# --sheet flag
# ---------------------------------------------------------------------------


def test_read_sheet_flag(multisheet_xlsx):
    """--sheet selects the specified sheet for reading."""
    result = runner.invoke(app, ["read", str(multisheet_xlsx), "--sheet", "Beta"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["headers"] == ["ID", "Value"]
    assert data["row_count"] == 10


def test_read_sheet_not_found(multisheet_xlsx):
    """--sheet with a non-existent name produces a SHEET_NOT_FOUND error."""
    result = runner.invoke(app, ["read", str(multisheet_xlsx), "--sheet", "DoesNotExist"])
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "SHEET_NOT_FOUND"


# ---------------------------------------------------------------------------
# --limit flag
# ---------------------------------------------------------------------------


def test_read_limit(multisheet_xlsx):
    """--limit caps the number of returned rows."""
    result = runner.invoke(app, ["read", str(multisheet_xlsx), "--sheet", "Beta", "--limit", "3"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["row_count"] == 3
    assert len(data["data"]) == 3


# ---------------------------------------------------------------------------
# --offset flag
# ---------------------------------------------------------------------------


def test_read_offset(multisheet_xlsx):
    """--offset skips the specified number of rows."""
    # Beta has 10 data rows. Offset 5 should leave 5 rows.
    result = runner.invoke(app, ["read", str(multisheet_xlsx), "--sheet", "Beta", "--offset", "5"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["row_count"] == 5


# ---------------------------------------------------------------------------
# --limit + --offset combined
# ---------------------------------------------------------------------------


def test_read_limit_and_offset(multisheet_xlsx):
    """--limit + --offset work together: skip first N rows then cap at M."""
    result = runner.invoke(
        app,
        ["read", str(multisheet_xlsx), "--sheet", "Beta", "--offset", "2", "--limit", "3"],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["row_count"] == 3
    assert len(data["data"]) == 3


# ---------------------------------------------------------------------------
# --sort ascending
# ---------------------------------------------------------------------------


def test_read_sort_ascending(multisheet_xlsx):
    """--sort sorts rows by the specified column in ascending order."""
    result = runner.invoke(
        app, ["read", str(multisheet_xlsx), "--sheet", "Alpha", "--sort", "Score"]
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    score_idx = data["headers"].index("Score")
    scores = [row[score_idx] for row in data["data"]]
    assert scores == sorted(scores)


# ---------------------------------------------------------------------------
# --sort --descending
# ---------------------------------------------------------------------------


def test_read_sort_descending(multisheet_xlsx):
    """--sort --descending sorts rows in descending order."""
    result = runner.invoke(
        app,
        ["read", str(multisheet_xlsx), "--sheet", "Alpha", "--sort", "Score", "--descending"],
    )
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    score_idx = data["headers"].index("Score")
    scores = [row[score_idx] for row in data["data"]]
    assert scores == sorted(scores, reverse=True)


# ---------------------------------------------------------------------------
# --compact (default) and --no-compact
# ---------------------------------------------------------------------------


def test_read_compact_default(compact_xlsx):
    """Default compact mode drops the fully-null NullCol column."""
    result = runner.invoke(app, ["read", str(compact_xlsx)])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert "NullCol" not in data["headers"]
    assert "Name" in data["headers"]


def test_read_no_compact(compact_xlsx):
    """--no-compact preserves fully-null columns."""
    result = runner.invoke(app, ["read", str(compact_xlsx), "--no-compact"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert "NullCol" in data["headers"]
    assert len(data["headers"]) == 4


# ---------------------------------------------------------------------------
# --all-sheets
# ---------------------------------------------------------------------------


def test_read_all_sheets(multisheet_xlsx):
    """--all-sheets reads every sheet and returns a results array."""
    result = runner.invoke(app, ["read", str(multisheet_xlsx), "--all-sheets"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert "results" in data
    assert data["total_ranges"] == 3
    sheet_names = [r["sheet"] for r in data["results"]]
    assert "Alpha" in sheet_names
    assert "Beta" in sheet_names
    assert "Gamma" in sheet_names


# ---------------------------------------------------------------------------
# --format csv
# ---------------------------------------------------------------------------


def test_read_format_csv(multisheet_xlsx):
    """--format csv outputs raw CSV to stdout."""
    result = runner.invoke(
        app, ["read", str(multisheet_xlsx), "--sheet", "Alpha", "--format", "csv"]
    )
    assert result.exit_code == 0, result.stdout
    lines = result.stdout.strip().split("\n")
    # First line is header row
    assert "Name" in lines[0]
    assert "Score" in lines[0]
    # 5 data rows + 1 header = 6 lines
    assert len(lines) == 6


# ---------------------------------------------------------------------------
# truncated flag
# ---------------------------------------------------------------------------


def test_read_truncated_flag_when_at_limit(multisheet_xlsx):
    """truncated=true when row_count equals the effective limit."""
    # Beta has 10 rows; setting limit=5 should truncate
    result = runner.invoke(app, ["read", str(multisheet_xlsx), "--sheet", "Beta", "--limit", "5"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["truncated"] is True


def test_read_truncated_flag_when_below_limit(multisheet_xlsx):
    """truncated=false when all rows fit within the limit."""
    # Alpha has 5 rows; default limit is 100, so not truncated
    result = runner.invoke(app, ["read", str(multisheet_xlsx), "--sheet", "Alpha"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["truncated"] is False
