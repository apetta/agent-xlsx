"""Tests for write command: auto-create, size guard, formula support, --value, --from-json."""

import json

from openpyxl import load_workbook
from typer.testing import CliRunner

from agent_xlsx.cli import app

runner = CliRunner()


# ---------------------------------------------------------------------------
# P0 #1 — Auto-create new files
# ---------------------------------------------------------------------------


def test_write_creates_new_xlsx(tmp_path):
    """Writing to a non-existent .xlsx file auto-creates it."""
    out = tmp_path / "new.xlsx"
    result = runner.invoke(app, ["write", str(out), "A1", "hello"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert data["created"] is True
    assert out.exists()


def test_write_creates_new_xlsx_with_sheet(tmp_path):
    """Auto-created file uses the --sheet name for the default sheet."""
    out = tmp_path / "new.xlsx"
    result = runner.invoke(app, ["write", str(out), "A1", "hello", "--sheet", "Dashboard"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["created"] is True

    # Verify sheet name
    read_result = runner.invoke(app, ["read", str(out), "--sheet", "Dashboard"])
    assert read_result.exit_code == 0, read_result.stdout


def test_write_creates_new_xlsx_with_json(tmp_path):
    """Auto-create works with --json array data."""
    out = tmp_path / "new.xlsx"
    result = runner.invoke(app, ["write", str(out), "A1", "--json", "[[1,2],[3,4]]"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["created"] is True
    assert data["cells_written"] == 4


def test_write_creates_new_xlsx_with_from_csv(tmp_path):
    """Auto-create works with --from-csv."""
    csv_file = tmp_path / "data.csv"
    csv_file.write_text("a,b\n1,2\n")
    out = tmp_path / "new.xlsx"
    result = runner.invoke(app, ["write", str(out), "A1", "--from-csv", str(csv_file)])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["created"] is True


def test_write_rejects_non_writable_extension(tmp_path):
    """Writing to a non-existent .xls file fails (not writable)."""
    out = tmp_path / "new.xls"
    result = runner.invoke(app, ["write", str(out), "A1", "hello"])
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "INVALID_FORMAT"


def test_write_existing_file_no_created_flag(sample_xlsx):
    """Writing to an existing file does NOT include 'created' in output."""
    result = runner.invoke(app, ["write", str(sample_xlsx), "C1", "test"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert "created" not in data


def test_write_output_creates_new_file_directly(tmp_path):
    """--output with non-existent source creates the output file directly."""
    source = tmp_path / "nonexistent.xlsx"
    output = tmp_path / "output.xlsx"
    result = runner.invoke(app, ["write", str(source), "A1", "test", "--output", str(output)])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["created"] is True
    assert output.exists()
    # Source file should NOT be created (we write directly to output)
    assert not source.exists()


# ---------------------------------------------------------------------------
# P0 #2 — Size guard for large files
# ---------------------------------------------------------------------------


def test_write_output_rejects_large_source(sample_xlsx, monkeypatch):
    """--output with a large source file fails fast with FILE_TOO_LARGE."""
    monkeypatch.setattr(
        "agent_xlsx.adapters.openpyxl_adapter.file_size_bytes",
        lambda _: 30 * 1024 * 1024,  # 30MB
    )
    result = runner.invoke(app, ["write", str(sample_xlsx), "A1", "test", "-o", "/tmp/out.xlsx"])
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "FILE_TOO_LARGE"


def test_write_inplace_warns_large_file(sample_xlsx, monkeypatch):
    """In-place write on large file succeeds with a warning."""
    monkeypatch.setattr(
        "agent_xlsx.adapters.openpyxl_adapter.file_size_bytes",
        lambda _: 30 * 1024 * 1024,  # 30MB
    )
    result = runner.invoke(app, ["write", str(sample_xlsx), "A1", "test"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert "warning" in data
    assert "30MB" in data["warning"]


def test_write_normal_size_no_warning(sample_xlsx):
    """Small files produce no warning."""
    result = runner.invoke(app, ["write", str(sample_xlsx), "A1", "test"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert "warning" not in data


# ---------------------------------------------------------------------------
# P1 #6 — Formula support in --json and --from-csv
# ---------------------------------------------------------------------------


def test_write_json_formula_autodetect(tmp_path):
    """Strings starting with '=' in --json are auto-detected as formulas."""
    out = tmp_path / "new.xlsx"
    runner.invoke(app, ["write", str(out), "A1", "--json", '[[10,20],["=SUM(A1:B1)",""]]'])
    result = runner.invoke(app, ["read", str(out), "--formulas"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    formulas = [c for c in data["cells"] if c.get("formula")]
    assert any("SUM" in f["formula"] for f in formulas)


def test_write_json_formula_flag(tmp_path):
    """--formula with --json: strings starting with '=' are written as formulas."""
    out = tmp_path / "new.xlsx"
    runner.invoke(
        app,
        [
            "write",
            str(out),
            "A1",
            "--json",
            '[["=SUM(A2:A5)","=AVERAGE(B2:B5)"]]',
            "--formula",
        ],
    )
    result = runner.invoke(app, ["read", str(out), "--formulas"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    formulas = [c["formula"] for c in data["cells"] if c.get("formula")]
    assert "=SUM(A2:A5)" in formulas
    assert "=AVERAGE(B2:B5)" in formulas


def test_write_json_formula_flag_no_double_equals(tmp_path):
    """--formula flag does not double-prefix strings that already start with '='."""
    out = tmp_path / "new.xlsx"
    runner.invoke(
        app,
        [
            "write",
            str(out),
            "A1",
            "--json",
            '[["=SUM(A2:A5)"]]',
            "--formula",
        ],
    )
    result = runner.invoke(app, ["read", str(out), "--formulas"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    formulas = [c["formula"] for c in data["cells"] if c.get("formula")]
    assert "=SUM(A2:A5)" in formulas
    # No "==SUM(A2:A5)"
    assert not any(f.startswith("==") for f in formulas)


def test_write_csv_formula_flag(tmp_path):
    """--formula with --from-csv: '=' prefixed values are written as formulas."""
    csv_file = tmp_path / "formulas.csv"
    csv_file.write_text("=SUM(A1:A5)\n=AVERAGE(B1:B5)\n")
    out = tmp_path / "new.xlsx"
    runner.invoke(
        app,
        ["write", str(out), "A1", "--from-csv", str(csv_file), "--formula"],
    )
    result = runner.invoke(app, ["read", str(out), "--formulas"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    formulas = [c["formula"] for c in data["cells"] if c.get("formula")]
    assert "=SUM(A1:A5)" in formulas
    assert "=AVERAGE(B1:B5)" in formulas


def test_write_json_formula_mixed_content(tmp_path):
    """--formula with --json: plain strings are NOT converted to formulas."""
    out = tmp_path / "new.xlsx"
    runner.invoke(
        app,
        [
            "write",
            str(out),
            "A1",
            "--json",
            '[["United Kingdom","GDP growth (%)","=AVERAGE(C2:M2)"]]',
            "--formula",
        ],
    )
    result = runner.invoke(app, ["read", str(out), "--formulas"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    # Plain text should NOT be formulas
    plain_cells = [c for c in data["cells"] if c["cell"] in ("A1", "B1")]
    for c in plain_cells:
        assert c["formula"] is None, f"{c['cell']} should not be a formula"
    # Formula should be preserved
    formula_cells = [c for c in data["cells"] if c["cell"] == "C1"]
    assert formula_cells[0]["formula"] == "=AVERAGE(C2:M2)"


def test_write_csv_formula_mixed_content(tmp_path):
    """--formula with --from-csv: only '=' prefixed values become formulas."""
    csv_file = tmp_path / "mixed.csv"
    csv_file.write_text("United Kingdom,GDP growth (%),=AVERAGE(C2:M2)\n")
    out = tmp_path / "new.xlsx"
    runner.invoke(
        app,
        ["write", str(out), "A1", "--from-csv", str(csv_file), "--formula"],
    )
    result = runner.invoke(app, ["read", str(out), "--formulas"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    # Plain text should NOT be formulas
    plain_cells = [c for c in data["cells"] if c["cell"] in ("A1", "B1")]
    for c in plain_cells:
        assert c["formula"] is None, f"{c['cell']} should not be a formula"
    # Formula should be preserved
    formula_cells = [c for c in data["cells"] if c["cell"] == "C1"]
    assert formula_cells[0]["formula"] == "=AVERAGE(C2:M2)"


# ---------------------------------------------------------------------------
# Issue #1 — --value option for negative numbers
# ---------------------------------------------------------------------------


def test_write_negative_number_via_value_option(tmp_path):
    """--value handles negative numbers that Click would misparse as flags."""
    out = tmp_path / "neg.xlsx"
    result = runner.invoke(app, ["write", str(out), "C10", "--value", "-4.095"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"

    # Verify the value was written correctly
    wb = load_workbook(str(out))
    ws = wb.active
    assert ws["C10"].value == -4.095
    wb.close()


def test_write_value_option_overrides_positional(tmp_path):
    """--value takes precedence over the positional value argument."""
    out = tmp_path / "override.xlsx"
    result = runner.invoke(app, ["write", str(out), "A1", "positional", "--value", "option-wins"])
    assert result.exit_code == 0, result.stdout

    wb = load_workbook(str(out))
    ws = wb.active
    assert ws["A1"].value == "option-wins"
    wb.close()


def test_write_negative_number_double_dash(tmp_path):
    """-- sentinel allows negative numbers as positional args."""
    out = tmp_path / "sentinel.xlsx"
    result = runner.invoke(app, ["write", str(out), "A1", "--", "-99.5"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"

    wb = load_workbook(str(out))
    ws = wb.active
    assert ws["A1"].value == -99.5
    wb.close()


# ---------------------------------------------------------------------------
# Issue #5/6 — --from-json file input
# ---------------------------------------------------------------------------


def test_write_from_json_file(tmp_path):
    """--from-json reads 2D array data from a JSON file."""
    json_file = tmp_path / "data.json"
    json_file.write_text("[[1, 2], [3, 4]]")
    out = tmp_path / "from_json.xlsx"

    result = runner.invoke(app, ["write", str(out), "A1", "--from-json", str(json_file)])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert data["status"] == "success"
    assert data["cells_written"] == 4

    # Verify written data
    wb = load_workbook(str(out))
    ws = wb.active
    assert ws["A1"].value == 1
    assert ws["B1"].value == 2
    assert ws["A2"].value == 3
    assert ws["B2"].value == 4
    wb.close()


def test_write_from_json_file_not_found(tmp_path):
    """--from-json with non-existent file produces a clear error."""
    out = tmp_path / "test.xlsx"
    result = runner.invoke(app, ["write", str(out), "A1", "--from-json", "/tmp/nonexistent.json"])
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "FILE_NOT_FOUND"


# ---------------------------------------------------------------------------
# P2 — Relative output path
# ---------------------------------------------------------------------------


def test_write_output_file_is_relative(tmp_path):
    """output_file in response must be present, relative, and match the target filename."""
    from pathlib import Path

    out = tmp_path / "out.xlsx"
    result = runner.invoke(app, ["write", str(out), "A1", "test"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)
    assert "output_file" in data, "output_file must be present in write response"
    assert not data["output_file"].startswith("/"), (
        f"output_file should be relative, got: {data['output_file']}"
    )
    assert Path(data["output_file"]).name == "out.xlsx"
