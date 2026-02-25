"""Shared pytest fixtures for agent-xlsx tests."""

import pytest
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink


@pytest.fixture
def sample_xlsx(tmp_path):
    """Minimal .xlsx with known cell content for command output tests."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "header"
    ws["A2"] = "value1"
    ws["B1"] = "amount"
    ws["B2"] = 100
    p = tmp_path / "sample.xlsx"
    wb.save(p)
    return p


@pytest.fixture
def sample_xlsm(tmp_path):
    """Minimal macro-enabled .xlsm for VBA gate tests (no actual VBA content needed)."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "test"
    p = tmp_path / "sample.xlsm"
    wb.save(p)
    return p


@pytest.fixture
def rich_xlsx(tmp_path):
    """Multi-purpose workbook with formulas, comments, CF, DV, hyperlinks, merged cells.

    Used by: test_probe, test_overview, test_inspect, test_export
    - Sheet "Sales": 5 cols (Date, Product, Revenue, Quantity, Region), 15 data rows
    - Sheet "Summary": formulas, merged cell, comment, hyperlink, CF, DV
    """
    wb = Workbook()

    # --- Sheet: Sales ---
    ws_sales = wb.active
    ws_sales.title = "Sales"
    headers = ["Date", "Product", "Revenue", "Quantity", "Region"]
    for col_idx, h in enumerate(headers, 1):
        ws_sales.cell(row=1, column=col_idx, value=h)

    # Apply date number format to column A (Date)
    regions = ["North", "South", "East", "West"]
    products = ["Widget", "Gadget", "Doohickey"]
    for i in range(2, 17):  # 15 data rows
        cell_date = ws_sales.cell(row=i, column=1, value=45000 + i)  # Excel serial dates
        cell_date.number_format = "yyyy-mm-dd"
        ws_sales.cell(row=i, column=2, value=products[(i - 2) % len(products)])
        ws_sales.cell(row=i, column=3, value=(i - 1) * 1000.50)
        ws_sales.cell(row=i, column=4, value=(i - 1) * 10)
        ws_sales.cell(row=i, column=5, value=regions[(i - 2) % len(regions)])

    # --- Sheet: Summary ---
    ws_summary = wb.create_sheet("Summary")

    # Merged cell
    ws_summary.merge_cells("A1:C1")
    ws_summary["A1"] = "Sales Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)

    # Labels and formulas
    ws_summary["A3"] = "Total Revenue"
    ws_summary["B3"] = "=SUM(Sales!C2:C16)"
    ws_summary["A4"] = "Average Revenue"
    ws_summary["B4"] = "=AVERAGE(Sales!C2:C16)"
    ws_summary["A5"] = "Total Quantity"
    ws_summary["B5"] = "=SUM(Sales!D2:D16)"

    # Comment on A1
    ws_summary["A1"].comment = Comment("This is the summary header", "TestAuthor")

    # Hyperlink on C5
    ws_summary["C5"] = "More Info"
    ws_summary["C5"].hyperlink = "https://example.com"
    hl = Hyperlink(ref="C5", target="https://example.com", display="More Info")
    ws_summary._hyperlinks.append(hl)

    # Conditional formatting on B3:B5 (highlight if > 10000)
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws_summary.conditional_formatting.add(
        "B3:B5",
        CellIsRule(operator="greaterThan", formula=["10000"], fill=red_fill),
    )

    # Data validation on D2 (list validation)
    dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    dv.error = "Please select a valid priority"
    dv.errorTitle = "Invalid Priority"
    dv.prompt = "Choose a priority level"
    ws_summary.add_data_validation(dv)
    dv.add("D2")
    ws_summary["D1"] = "Priority"
    ws_summary["D2"] = "High"

    p = tmp_path / "rich.xlsx"
    wb.save(p)
    return p


@pytest.fixture
def multisheet_xlsx(tmp_path):
    """3-sheet workbook for --sheet and --all-sheets tests.

    - "Alpha": 3 cols (Name, Score, Grade), 5 rows
    - "Beta": 2 cols (ID, Value), 10 rows
    - "Gamma": headers only (empty data)
    """
    wb = Workbook()

    ws_alpha = wb.active
    ws_alpha.title = "Alpha"
    ws_alpha["A1"] = "Name"
    ws_alpha["B1"] = "Score"
    ws_alpha["C1"] = "Grade"
    for i in range(2, 7):
        ws_alpha[f"A{i}"] = f"Student-{i - 1}"
        ws_alpha[f"B{i}"] = 60 + (i - 1) * 5
        ws_alpha[f"C{i}"] = "A" if i > 4 else "B"

    ws_beta = wb.create_sheet("Beta")
    ws_beta["A1"] = "ID"
    ws_beta["B1"] = "Value"
    for i in range(2, 12):
        ws_beta[f"A{i}"] = i - 1
        ws_beta[f"B{i}"] = (i - 1) * 100

    ws_gamma = wb.create_sheet("Gamma")
    ws_gamma["A1"] = "Empty1"
    ws_gamma["B1"] = "Empty2"
    # No data rows — headers only

    p = tmp_path / "multisheet.xlsx"
    wb.save(p)
    return p


@pytest.fixture
def formula_error_xlsx(tmp_path):
    """Workbook with deliberate formula errors for recalc --check-only.

    Uses cached error string values (data_only=True reads these).
    Also has normal formulas for total_formulas counting.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Errors"

    # Normal data for formula references
    ws["A1"] = "Value"
    ws["A2"] = 10
    ws["A3"] = 20
    ws["A4"] = 30
    ws["A5"] = 0

    # Normal formulas
    ws["B1"] = "Formula"
    ws["B2"] = "=SUM(A2:A5)"
    ws["B3"] = "=A2+A3"

    # Cached error values (simulates what openpyxl reads in data_only=True)
    ws["C1"] = "Errors"
    ws["C2"] = "#REF!"
    ws["C3"] = "#DIV/0!"
    ws["C4"] = "#NAME?"
    ws["C5"] = "#REF!"  # Second #REF! to test count > 1

    p = tmp_path / "formula_errors.xlsx"
    wb.save(p)
    return p


@pytest.fixture
def compact_xlsx(tmp_path):
    """Workbook with a fully-null column for compact/no-compact testing.

    4 columns: Name, Value, NullCol (all None), Category. 5 data rows.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["C1"] = "NullCol"
    ws["D1"] = "Category"
    for i in range(2, 7):
        ws[f"A{i}"] = f"Item-{i - 1}"
        ws[f"B{i}"] = (i - 1) * 10
        # C column intentionally left empty (null)
        ws[f"D{i}"] = "CatA" if i % 2 == 0 else "CatB"

    p = tmp_path / "compact.xlsx"
    wb.save(p)
    return p


@pytest.fixture
def unicode_xlsx(tmp_path):
    """Workbook with Unicode content for edge case testing.

    Headers with accented characters, CJK values.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Nom"
    ws["B1"] = "Montant (EUR)"
    ws["C1"] = "Region"
    ws["A2"] = "Jean-Pierre"
    ws["B2"] = 1500.75
    ws["C2"] = "Ile-de-France"
    ws["A3"] = "\u5c71\u7530\u592a\u90ce"  # Japanese name
    ws["B3"] = 2300.00
    ws["C3"] = "\u6771\u4eac"  # Tokyo
    ws["A4"] = "Muller"
    ws["B4"] = 980.50
    ws["C4"] = "Baden-Wurttemberg"

    p = tmp_path / "unicode.xlsx"
    wb.save(p)
    return p
