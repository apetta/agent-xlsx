"""Tests for the inspect command: metadata, sheet, range, names, charts, vba,
format, comments, conditional, validation, hyperlinks, and error handling."""

import json

import pytest
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from typer.testing import CliRunner

from agent_xlsx.cli import app

runner = CliRunner()


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def inspect_xlsx(tmp_path):
    """Workbook with formulas, merged cells, comments, and DV for full inspection.

    Omits conditional formatting and _hyperlinks because openpyxl 3.1.5
    does not reliably round-trip these through _cf_rules iteration and
    _hyperlinks after save/reload.
    """
    wb = Workbook()

    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Value"
    ws["A2"] = 10
    ws["A3"] = 20

    ws_summary = wb.create_sheet("Summary")

    # Merged cell
    ws_summary.merge_cells("A1:C1")
    ws_summary["A1"] = "Report Title"
    ws_summary["A1"].font = Font(bold=True, size=14)

    # Formulas
    ws_summary["A3"] = "Total"
    ws_summary["B3"] = "=SUM(Data!A2:A3)"
    ws_summary["A4"] = "Average"
    ws_summary["B4"] = "=AVERAGE(Data!A2:A3)"

    # Comment
    ws_summary["A1"].comment = Comment("Header comment", "Tester")

    # Data validation
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv.error = "Pick Yes or No"
    dv.errorTitle = "Invalid"
    dv.prompt = "Select option"
    ws_summary.add_data_validation(dv)
    dv.add("D2")

    p = tmp_path / "inspect.xlsx"
    wb.save(p)
    return p


# ---------------------------------------------------------------------------
# Default metadata (no flags)
# ---------------------------------------------------------------------------


def test_inspect_default_metadata(rich_xlsx):
    """Default inspect returns workbook-level metadata for all sheets."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx)])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    # Workbook-level keys from get_workbook_metadata
    assert "sheets" in data
    assert "named_ranges" in data
    assert "named_range_count" in data
    assert "has_vba" in data
    assert "total_formula_count" in data
    assert "total_chart_count" in data

    # rich_xlsx has Sales and Summary sheets
    sheet_names = [s["name"] for s in data["sheets"]]
    assert "Sales" in sheet_names
    assert "Summary" in sheet_names

    # Summary has 3 formulas (B3, B4, B5)
    assert data["total_formula_count"] == 3
    assert data["has_vba"] is False


def test_inspect_default_sheet_metadata_shape(rich_xlsx):
    """Each sheet in default inspect has the expected per-sheet keys."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx)])
    data = json.loads(result.stdout)

    for sheet in data["sheets"]:
        assert "name" in sheet
        assert "dimensions" in sheet
        assert "has_formulas" in sheet
        assert "formula_count" in sheet
        assert "has_charts" in sheet
        assert "chart_count" in sheet
        assert "has_tables" in sheet
        assert "table_names" in sheet


# ---------------------------------------------------------------------------
# --sheet (full sheet inspection)
# ---------------------------------------------------------------------------


def test_inspect_sheet_full_inspection(inspect_xlsx):
    """--sheet returns comprehensive inspection data for a single sheet."""
    result = runner.invoke(app, ["inspect", str(inspect_xlsx), "--sheet", "Summary"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert data["sheet"] == "Summary"
    assert "dimensions" in data

    # Formulas section (B3, B4)
    assert data["formulas"]["count"] == 2
    assert len(data["formulas"]["sample"]) == 2
    assert data["formulas"]["truncated"] is False

    # Merged cells section (A1:C1 is merged)
    assert data["merged_cells"]["count"] == 1
    assert "A1:C1" in data["merged_cells"]["regions"]

    # Comments section (A1 has a comment)
    assert data["comments"]["count"] == 1
    assert data["comments"]["items"][0]["cell"] == "A1"
    assert data["comments"]["items"][0]["author"] == "Tester"
    assert "Header comment" in data["comments"]["items"][0]["text"]

    # Conditional formatting section (none in this fixture)
    assert data["conditional_formatting"]["count"] == 0

    # Data validation section (D2 has list validation)
    assert data["data_validation"]["count"] == 1
    dv_rule = data["data_validation"]["rules"][0]
    assert dv_rule["type"] == "list"
    assert dv_rule["allow_blank"] is True
    assert dv_rule["error_message"] == "Pick Yes or No"
    assert dv_rule["error_title"] == "Invalid"
    assert dv_rule["prompt"] == "Select option"

    # Hyperlinks section (none in this fixture)
    assert data["hyperlinks"]["count"] == 0


def test_inspect_sheet_full_with_all_elements(rich_xlsx):
    """--sheet on rich_xlsx Summary has CF, hyperlinks, comments, DV, merged cells."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--sheet", "Summary"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert data["sheet"] == "Summary"
    assert data["formulas"]["count"] == 3
    assert data["merged_cells"]["count"] == 1
    assert "A1:C1" in data["merged_cells"]["regions"]
    assert data["comments"]["count"] == 1
    assert data["conditional_formatting"]["count"] == 1
    assert data["conditional_formatting"]["rules"][0]["type"] == "cellIs"
    assert data["data_validation"]["count"] == 1
    assert data["hyperlinks"]["count"] == 1
    assert data["hyperlinks"]["items"][0]["target"] == "https://example.com"


def test_inspect_sheet_no_special_elements(inspect_xlsx):
    """Data sheet has no formulas, comments, CF, DV, or hyperlinks."""
    result = runner.invoke(app, ["inspect", str(inspect_xlsx), "--sheet", "Data"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert data["sheet"] == "Data"
    assert data["formulas"]["count"] == 0
    assert data["comments"]["count"] == 0
    assert data["conditional_formatting"]["count"] == 0
    assert data["data_validation"]["count"] == 0
    assert data["hyperlinks"]["count"] == 0


# ---------------------------------------------------------------------------
# --range (inspect formulas in a cell range)
# ---------------------------------------------------------------------------


def test_inspect_range_with_formulas(rich_xlsx):
    """--range returns formula summary for a range containing formulas."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--range", "Summary!B3:B5"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert data["range"] == "Summary!B3:B5"
    assert data["sheet"] == "Summary"
    # formula_count includes all cells in the range (3 cells, all formulas)
    assert data["formula_count"] == 3
    assert len(data["sample_formulas"]) == 3

    # Verify actual formulas are present
    formulas = [f["formula"] for f in data["sample_formulas"] if f.get("formula")]
    assert any("SUM" in f for f in formulas)
    assert any("AVERAGE" in f for f in formulas)


def test_inspect_range_data_only(rich_xlsx):
    """--range on a data-only range returns cells with no formulas."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--range", "Sales!A1:E1"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    # formula_count reflects total cells in the range (5 header cells)
    assert data["formula_count"] == 5
    # None of the cells should have actual formulas
    for cell in data["sample_formulas"]:
        assert cell["formula"] is None


def test_inspect_range_defaults_to_first_sheet(rich_xlsx):
    """--range without sheet prefix defaults to the first sheet."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--range", "A1:B2"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    # First sheet in rich_xlsx is "Sales"
    assert data["sheet"] == "Sales"


def test_inspect_range_output_shape(rich_xlsx):
    """--range output has the expected keys from summarise_formulas."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--range", "Summary!B3:B5"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert "formula_count" in data
    assert "formula_columns" in data
    assert "sample_formulas" in data
    assert "truncated" in data
    assert "range" in data
    assert "sheet" in data


# ---------------------------------------------------------------------------
# --names (named ranges)
# ---------------------------------------------------------------------------


def test_inspect_names(rich_xlsx):
    """--names returns named ranges info."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--names"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert "named_ranges" in data
    assert "count" in data
    # rich_xlsx has no named ranges defined
    assert data["count"] == 0
    assert data["named_ranges"] == []


# ---------------------------------------------------------------------------
# --charts
# ---------------------------------------------------------------------------


def test_inspect_charts(rich_xlsx):
    """--charts returns chart info across all sheets."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--charts"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert "charts" in data
    assert "total_chart_count" in data
    # rich_xlsx has no charts
    assert data["total_chart_count"] == 0
    assert data["charts"] == []


# ---------------------------------------------------------------------------
# --vba
# ---------------------------------------------------------------------------


def test_inspect_vba_no_macros(rich_xlsx):
    """--vba on a regular .xlsx reports no VBA."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--vba"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert "has_vba" in data
    assert data["has_vba"] is False


# ---------------------------------------------------------------------------
# --format (cell formatting)
# ---------------------------------------------------------------------------


def test_inspect_format_cell_with_sheet_ref(rich_xlsx):
    """--format with Sheet!Cell returns detailed formatting."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--format", "Summary!A1"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert data["cell"] == "A1"
    # A1 in Summary has bold=True, size=14
    assert data["font"]["bold"] is True
    assert data["font"]["size"] == 14

    # Standard formatting keys present
    assert "fill" in data
    assert "border" in data
    assert "alignment" in data
    assert "number_format" in data


def test_inspect_format_with_sheet_flag(rich_xlsx):
    """--format combined with --sheet resolves the correct sheet."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--sheet", "Summary", "--format", "A1"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    # Should match Summary!A1 formatting
    assert data["font"]["bold"] is True
    assert data["font"]["size"] == 14


def test_inspect_format_defaults_to_first_sheet(rich_xlsx):
    """--format without sheet prefix or --sheet defaults to first sheet."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--format", "A1"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    # First sheet is Sales; A1 value is "Date"
    assert data["cell"] == "A1"
    assert data["value"] == "Date"


def test_inspect_format_output_shape(rich_xlsx):
    """--format output has font, fill, border, alignment, and number_format."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--format", "Summary!A1"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    # Font sub-keys
    assert "name" in data["font"]
    assert "size" in data["font"]
    assert "bold" in data["font"]
    assert "italic" in data["font"]
    assert "underline" in data["font"]
    assert "color" in data["font"]

    # Fill sub-keys
    assert "type" in data["fill"]
    assert "color" in data["fill"]

    # Border sub-keys
    assert "top" in data["border"]
    assert "bottom" in data["border"]
    assert "left" in data["border"]
    assert "right" in data["border"]

    # Alignment sub-keys
    assert "horizontal" in data["alignment"]
    assert "vertical" in data["alignment"]
    assert "wrap_text" in data["alignment"]
    assert "text_rotation" in data["alignment"]


# ---------------------------------------------------------------------------
# --comments
# ---------------------------------------------------------------------------


def test_inspect_comments(rich_xlsx):
    """--comments returns comment data for the target sheet."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--sheet", "Summary", "--comments"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert data["sheet"] == "Summary"
    assert data["comment_count"] == 1
    assert data["comments"][0]["cell"] == "A1"
    assert data["comments"][0]["author"] == "TestAuthor"
    assert data["comments"][0]["text"] == "This is the summary header"
    assert data["truncated"] is False


def test_inspect_comments_defaults_to_first_sheet(rich_xlsx):
    """--comments without --sheet defaults to the first sheet (Sales)."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--comments"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    # Sales sheet has no comments
    assert data["sheet"] == "Sales"
    assert data["comment_count"] == 0


def test_inspect_comments_output_shape(rich_xlsx):
    """--comments output has the expected capped-list keys."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--sheet", "Summary", "--comments"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert "sheet" in data
    assert "comments" in data
    assert "comment_count" in data
    assert "truncated" in data

    # Each comment has cell, author, text
    comment = data["comments"][0]
    assert "cell" in comment
    assert "author" in comment
    assert "text" in comment


# ---------------------------------------------------------------------------
# --conditional
# ---------------------------------------------------------------------------


def test_inspect_conditional(rich_xlsx):
    """--conditional returns CF rules for a sheet with conditional formatting."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--conditional", "Summary"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert data["sheet"] == "Summary"
    assert data["rule_count"] == 1
    rule = data["rules"][0]
    assert rule["type"] == "cellIs"
    assert rule["operator"] == "greaterThan"
    assert "formula" in rule
    assert data["truncated"] is False


def test_inspect_conditional_no_rules(rich_xlsx):
    """--conditional on a sheet with no CF returns empty results."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--conditional", "Sales"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert data["sheet"] == "Sales"
    assert data["rule_count"] == 0
    assert data["rules"] == []
    assert data["truncated"] is False


# ---------------------------------------------------------------------------
# --validation
# ---------------------------------------------------------------------------


def test_inspect_validation(rich_xlsx):
    """--validation returns data validation rules for a sheet."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--validation", "Summary"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert data["sheet"] == "Summary"
    assert data["validation_count"] == 1
    val = data["validations"][0]
    assert val["type"] == "list"
    assert val["allow_blank"] is True
    assert val["error_message"] == "Please select a valid priority"
    assert val["error_title"] == "Invalid Priority"
    assert val["prompt"] == "Choose a priority level"
    assert data["truncated"] is False


def test_inspect_validation_no_rules(rich_xlsx):
    """--validation on a sheet with no DV returns empty results."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--validation", "Sales"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert data["validation_count"] == 0
    assert data["validations"] == []


def test_inspect_validation_output_shape(rich_xlsx):
    """--validation output has the expected capped-list keys."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--validation", "Summary"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert "sheet" in data
    assert "validations" in data
    assert "validation_count" in data
    assert "truncated" in data


# ---------------------------------------------------------------------------
# --hyperlinks
# ---------------------------------------------------------------------------


def test_inspect_hyperlinks(rich_xlsx):
    """--hyperlinks returns hyperlink data for a sheet with links."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--hyperlinks", "Summary"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert data["sheet"] == "Summary"
    assert data["hyperlink_count"] == 1
    link = data["hyperlinks"][0]
    assert link["cell"] == "C5"
    assert link["target"] == "https://example.com"
    assert data["truncated"] is False


def test_inspect_hyperlinks_no_links(rich_xlsx):
    """--hyperlinks on a sheet with no hyperlinks returns empty results."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--hyperlinks", "Sales"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert data["hyperlink_count"] == 0
    assert data["hyperlinks"] == []


# ---------------------------------------------------------------------------
# Error cases
# ---------------------------------------------------------------------------


def test_inspect_file_not_found():
    """inspect with a missing file produces a structured error."""
    result = runner.invoke(app, ["inspect", "/tmp/nonexistent_inspect.xlsx"])
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "FILE_NOT_FOUND"


def test_inspect_invalid_format(tmp_path):
    """inspect with a non-Excel file produces INVALID_FORMAT error."""
    txt = tmp_path / "data.txt"
    txt.write_text("not excel")
    result = runner.invoke(app, ["inspect", str(txt)])
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "INVALID_FORMAT"


def test_inspect_sheet_not_found(rich_xlsx):
    """--sheet with a non-existent sheet name exits with a non-zero code."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--sheet", "NonExistent"])
    assert result.exit_code != 0


def test_inspect_invalid_range(rich_xlsx):
    """--range with an invalid range string produces RANGE_INVALID error."""
    result = runner.invoke(app, ["inspect", str(rich_xlsx), "--range", "not-a-range!!!"])
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "RANGE_INVALID"
