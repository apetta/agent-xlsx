"""Tests for the probe CLI command — workbook profiling."""

import json

import pytest
from typer.testing import CliRunner

from agent_xlsx.cli import app

runner = CliRunner()


# ---------------------------------------------------------------------------
# Default lean output (no flags)
# ---------------------------------------------------------------------------


def test_probe_default_lean_output(rich_xlsx):
    """Default probe returns sheet names, dimensions, headers — no data parsing."""
    result = runner.invoke(app, ["probe", str(rich_xlsx)])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    # Top-level metadata
    assert data["_data_origin"] == "untrusted_spreadsheet"
    assert data["file"] == "rich.xlsx"
    assert "size_bytes" in data
    assert "file_size_human" in data
    assert data["format"] == "xlsx"
    assert "probe_time_ms" in data
    assert isinstance(data["sheets"], list)
    assert len(data["sheets"]) == 2  # Sales, Summary

    # Sheet-level structure
    sales = data["sheets"][0]
    assert sales["name"] == "Sales"
    assert sales["rows"] > 0
    assert sales["cols"] == 5
    assert "headers" in sales
    assert "Date" in sales["headers"]
    assert sales["visible"] is True
    assert "last_col" in sales

    # Lean mode should NOT include profiling detail
    assert "column_types" not in sales
    assert "sample" not in sales
    assert "numeric_summary" not in sales


def test_probe_default_includes_column_map(sample_xlsx):
    """Default probe includes column_map for header-to-letter mapping."""
    result = runner.invoke(app, ["probe", str(sample_xlsx)])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    sheet = data["sheets"][0]
    assert "column_map" in sheet
    assert sheet["column_map"]["header"] == "A"
    assert sheet["column_map"]["amount"] == "B"


def test_probe_default_named_ranges_and_tables(rich_xlsx):
    """Default probe includes named_ranges, tables, and has_vba fields."""
    result = runner.invoke(app, ["probe", str(rich_xlsx)])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert "named_ranges" in data
    assert isinstance(data["named_ranges"], list)
    assert "tables" in data
    assert isinstance(data["tables"], list)
    assert data["has_vba"] is False


# ---------------------------------------------------------------------------
# --types flag
# ---------------------------------------------------------------------------


def test_probe_types_includes_column_types(rich_xlsx):
    """--types adds column_types and null_counts to each sheet."""
    result = runner.invoke(app, ["probe", str(rich_xlsx), "--types"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    sales = data["sheets"][0]
    assert "column_types" in sales
    assert "null_counts" in sales
    # Sales sheet has 5 columns — types should include at least some of them
    assert len(sales["column_types"]) > 0
    # No sample data without --sample
    assert "sample" not in sales


# ---------------------------------------------------------------------------
# --sample flag
# ---------------------------------------------------------------------------


def test_probe_sample_includes_head_tail(rich_xlsx):
    """--sample N returns head and tail sample rows."""
    result = runner.invoke(app, ["probe", str(rich_xlsx), "--sample", "2"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    sales = data["sheets"][0]
    assert "sample" in sales
    assert "head" in sales["sample"]
    assert "tail" in sales["sample"]
    assert len(sales["sample"]["head"]) == 2
    assert len(sales["sample"]["tail"]) == 2


# ---------------------------------------------------------------------------
# --stats flag
# ---------------------------------------------------------------------------


def test_probe_stats_includes_numeric_summary(rich_xlsx):
    """--stats includes numeric_summary with min/max/mean/median/std."""
    result = runner.invoke(app, ["probe", str(rich_xlsx), "--stats"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    sales = data["sheets"][0]
    # --stats implies --types
    assert "column_types" in sales
    assert "numeric_summary" in sales
    # Revenue and Quantity are numeric columns in Sales
    assert any(col in sales["numeric_summary"] for col in ["Revenue", "Quantity"])
    # Check summary structure
    for col_summary in sales["numeric_summary"].values():
        assert "min" in col_summary
        assert "max" in col_summary
        assert "mean" in col_summary


def test_probe_stats_includes_string_summary(rich_xlsx):
    """--stats includes string_summary with unique counts and top_values."""
    result = runner.invoke(app, ["probe", str(rich_xlsx), "--stats"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    sales = data["sheets"][0]
    # Product and Region are string columns in Sales
    assert "string_summary" in sales
    assert any(col in sales["string_summary"] for col in ["Product", "Region"])
    for col_summary in sales["string_summary"].values():
        assert "unique" in col_summary
        assert "top_values" in col_summary


# ---------------------------------------------------------------------------
# --full flag
# ---------------------------------------------------------------------------


def test_probe_full_enables_all_detail(rich_xlsx):
    """--full enables types + stats + sample(3)."""
    result = runner.invoke(app, ["probe", str(rich_xlsx), "--full"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    sales = data["sheets"][0]
    # --full implies types, stats, sample(3)
    assert "column_types" in sales
    assert "null_counts" in sales
    assert "numeric_summary" in sales
    assert "sample" in sales
    # Default sample from --full is 3
    assert len(sales["sample"]["head"]) == 3


# ---------------------------------------------------------------------------
# --sheet flag
# ---------------------------------------------------------------------------


def test_probe_sheet_filters_to_single_sheet(multisheet_xlsx):
    """--sheet filters output to the specified sheet only."""
    result = runner.invoke(app, ["probe", str(multisheet_xlsx), "--sheet", "Beta"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert len(data["sheets"]) == 1
    assert data["sheets"][0]["name"] == "Beta"
    assert data["sheets"][0]["cols"] == 2
    assert "ID" in data["sheets"][0]["headers"]


# ---------------------------------------------------------------------------
# --no-header flag
# ---------------------------------------------------------------------------


def test_probe_no_header_uses_column_letters(sample_xlsx):
    """--no-header uses column letters (A, B, C) instead of row-1 values."""
    result = runner.invoke(app, ["probe", str(sample_xlsx), "--no-header", "--types"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    sheet = data["sheets"][0]
    assert sheet["headers"][0] == "A"
    assert sheet["headers"][1] == "B"
    # column_map should NOT be present in no-header mode
    assert "column_map" not in sheet


def test_probe_no_header_lean_mode(sample_xlsx):
    """--no-header in lean mode (no --types) still uses letter headers."""
    result = runner.invoke(app, ["probe", str(sample_xlsx), "--no-header"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    sheet = data["sheets"][0]
    assert sheet["headers"][0] == "A"
    assert sheet["headers"][1] == "B"
    assert "column_map" not in sheet


# ---------------------------------------------------------------------------
# --head-cols flag
# ---------------------------------------------------------------------------


def test_probe_head_cols_limits_profiled_columns(rich_xlsx):
    """--head-cols N limits profiling to first N columns."""
    result = runner.invoke(app, ["probe", str(rich_xlsx), "--types", "--head-cols", "2"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    sales = data["sheets"][0]
    # Full header list should still include all columns
    assert len(sales["headers"]) == 5
    # But profiled detail is limited
    assert sales.get("profiled_columns") == 2
    assert sales.get("columns_truncated") is True
    # column_types should only have profiled columns
    assert len(sales["column_types"]) <= 2


# ---------------------------------------------------------------------------
# Empty sheet handling
# ---------------------------------------------------------------------------


def test_probe_empty_sheet(multisheet_xlsx):
    """Probe handles headers-only (empty data) sheets gracefully."""
    result = runner.invoke(app, ["probe", str(multisheet_xlsx), "--sheet", "Gamma"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    gamma = data["sheets"][0]
    assert gamma["name"] == "Gamma"
    assert "headers" in gamma
    assert "Empty1" in gamma["headers"]
    assert "Empty2" in gamma["headers"]


def test_probe_empty_sheet_with_stats(multisheet_xlsx):
    """--stats on an empty (headers-only) sheet does not crash."""
    result = runner.invoke(app, ["probe", str(multisheet_xlsx), "--sheet", "Gamma", "--stats"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    gamma = data["sheets"][0]
    # Should succeed but numeric_summary may be absent (no data rows)
    assert gamma["name"] == "Gamma"
    assert "column_types" in gamma


# ---------------------------------------------------------------------------
# Error cases — structured JSON errors
# ---------------------------------------------------------------------------


def test_probe_file_not_found(tmp_path):
    """Non-existent file returns FILE_NOT_FOUND structured error."""
    fake_path = str(tmp_path / "nonexistent.xlsx")
    result = runner.invoke(app, ["probe", fake_path])
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "FILE_NOT_FOUND"


def test_probe_invalid_format(tmp_path):
    """Non-Excel file returns INVALID_FORMAT structured error."""
    bad_file = tmp_path / "notes.txt"
    bad_file.write_text("not excel")
    result = runner.invoke(app, ["probe", str(bad_file)])
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "INVALID_FORMAT"


def test_probe_sheet_not_found(sample_xlsx):
    """--sheet with non-existent sheet name returns SHEET_NOT_FOUND error."""
    result = runner.invoke(app, ["probe", str(sample_xlsx), "--sheet", "DoesNotExist"])
    assert result.exit_code == 1
    data = json.loads(result.stdout)
    assert data["error"] is True
    assert data["code"] == "SHEET_NOT_FOUND"
    assert "suggestions" in data


# ---------------------------------------------------------------------------
# Multi-sheet default output
# ---------------------------------------------------------------------------


def test_probe_multisheet_all_sheets(multisheet_xlsx):
    """Probe without --sheet returns all sheets in order."""
    result = runner.invoke(app, ["probe", str(multisheet_xlsx)])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    assert len(data["sheets"]) == 3
    names = [s["name"] for s in data["sheets"]]
    assert names == ["Alpha", "Beta", "Gamma"]


# ---------------------------------------------------------------------------
# --full with --sample override
# ---------------------------------------------------------------------------


def test_probe_full_with_explicit_sample(rich_xlsx):
    """--full + --sample 5 uses the larger sample count."""
    result = runner.invoke(app, ["probe", str(rich_xlsx), "--full", "--sample", "5"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    sales = data["sheets"][0]
    assert "sample" in sales
    # --full sets sample=max(sample, 3), so --sample 5 wins
    assert len(sales["sample"]["head"]) == 5


# ---------------------------------------------------------------------------
# --brief flag — condensed profile
# ---------------------------------------------------------------------------


def test_probe_brief_includes_types_and_nulls(rich_xlsx):
    """--brief includes column_types and null_counts."""
    result = runner.invoke(app, ["probe", str(rich_xlsx), "--brief"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    sales = data["sheets"][0]
    assert "column_types" in sales, "--brief should include column_types"
    assert "null_counts" in sales, "--brief should include null_counts"


def test_probe_brief_excludes_sample_and_stats(rich_xlsx):
    """--brief excludes sample, numeric_summary, and string_summary."""
    result = runner.invoke(app, ["probe", str(rich_xlsx), "--brief"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    sales = data["sheets"][0]
    assert "sample" not in sales, "--brief should NOT include sample"
    assert "numeric_summary" not in sales, "--brief should NOT include numeric_summary"
    assert "string_summary" not in sales, "--brief should NOT include string_summary"


# ---------------------------------------------------------------------------
# String truncation in probe output
# ---------------------------------------------------------------------------


@pytest.fixture()
def freetext_xlsx(tmp_path):
    """Workbook with a free-text column (long strings) for truncation tests."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "ID"
    ws["B1"] = "Description"
    for i in range(2, 12):  # 10 data rows
        ws[f"A{i}"] = i - 1
        # 200+ char strings — avg well above FREETEXT_AVG_LENGTH_THRESHOLD (100)
        ws[f"B{i}"] = f"This is a very long description for item {i - 1}. " * 5
    p = tmp_path / "freetext.xlsx"
    wb.save(p)
    return p


def test_probe_stats_free_text_columns_are_compact(freetext_xlsx):
    """Free-text columns (avg > 100 chars) emit type: free_text instead of top_values."""
    result = runner.invoke(app, ["probe", str(freetext_xlsx), "--full"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    sheet = data["sheets"][0]
    assert "string_summary" in sheet
    desc_summary = sheet["string_summary"].get("Description")
    assert desc_summary is not None, "Description should be in string_summary"
    assert desc_summary["type"] == "free_text", "Long strings should be classified as free_text"
    assert "top_values" not in desc_summary, "free_text columns should not have top_values"
    assert "avg_length" in desc_summary


def test_probe_sample_long_strings_are_truncated(freetext_xlsx):
    """Sample row string values are capped at SAMPLE_VALUE_MAX_CHARS (100) + '...'."""
    result = runner.invoke(app, ["probe", str(freetext_xlsx), "--sample", "3"])
    assert result.exit_code == 0, result.stdout
    data = json.loads(result.stdout)

    sheet = data["sheets"][0]
    assert "sample" in sheet
    # Check all sample rows — description values should be truncated
    for row in sheet["sample"]["head"]:
        desc_val = row.get("Description")
        if desc_val and isinstance(desc_val, str):
            assert len(desc_val) <= 103, (  # 100 chars + "..."
                f"Sample string should be truncated to 103 chars, got {len(desc_val)}"
            )
